"""
validate_fea.py - INVERSE DESIGN PHASE 4: one real FEA per family. GATE 7.

Phase 2 produced geometry whose properties came from the model ALONE. This is
the only step that finds out whether the model was right. One homogenisation per
family on the design Phase 3 named, then the measured stiffness against the
predicted one.

    GATE 7   |measured - predicted| / measured  <=  15 %

WHAT IT DOES NOT DO: reimplement your pipeline.
    Every metric here comes from ntop_batch's OWN functions -
    voigt_reuss_hill and tensorial_anisotropy_index - imported, not copied.
    Those are what produced every E_over_Es and TAI in the training labels, so
    the GATE 7 comparison is like-for-like by construction. Reimplementing an
    nTop parser is what cost a failed run on 13/08.

    Both notebooks are driven exactly as production drives them:
        -v2 -s -j in.json -o out.json <notebook>
    with the notebook COPIED and refreshed from source before each run, because
    -s writes the result back into whichever notebook file is driven.

THE TWO NOTEBOOKS
    ADMS  Share\\ADMS\\ADMS_<Variant>_only_FEA.ntop   8 inputs:
          Density | Thickness (mm) | Seed | Out Path | Inner Size (mm) |
          Size Multi | Stress Path | Shear Stress Path
          Same design parameters as the only_STL run, so it rebuilds the SAME
          geometry and homogenises it. Es = 200 GPa (steel), nu_s = 0.3.

    TPMS  Share\\TPMS\\TPMS FEA\\FEA-P1\\<Type>_density0.20.ntop   4 inputs:
          t_max | t_min | Edge length (mm) | Path
          t is the value Phase 2 used. Es = 1.8 GPa (polymer), tensor in kPa.
          Both material constants come from ML_settings.xlsx, never hardcoded.

    Edge length IS THE FE MESH SIZE - in the notebook it feeds both `Feature
    size` and `Edge length` of the FE Volume Mesh block. It is set to
    EDGE_C * t.

    ⭐ EDGE_C = 0.4, MEASURED FROM THE NOTEBOOK, NOT FROM tpms_batch_run.
      That script has EDGE_C = 2 with a comment saying it is set by hand per
      model, and 2 is NOT what was used here. `FRD Generator_density0.20.ntop`
      was saved by `-s` after a real run and carries the values it ran with:
          t_max 0.7672   Edge length 0.3069 mm   ->  0.3069 / 0.7672 = 0.4000
      At 0.4 the mesh is 26 elements across a 10 mm cell and ~1.4 through the
      wall. At 2 it would be 5 across the cell and 0.28 through the wall - the
      wall would not be resolved at all and GATE 7 would be meaningless.
      Override with --edge-c only with evidence as good as that.

ONE-WAY RULE - THE WHOLE POINT OF THIS FOLDER
    Everything lands in Share\\Inverse design\\validation\\. These rows are a
    TEST of the model and must NEVER be merged into the training dataset: a row
    the model was tested on is not a row it can then be trained on. Nothing here
    writes to Share\\ADMS, Share\\TPMS, or Results_summary.xlsx.

Usage
    python validate_fea.py                 both families, from Phase 3
    python validate_fea.py --family adms
    python validate_fea.py --dry-run
"""

from __future__ import annotations
import os, sys, json, glob, time, shutil, argparse, subprocess
import numpy as np
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
ID_ROOT = os.path.dirname(HERE)
SHARE = os.path.normpath(os.path.join(ID_ROOT, ".."))
ML_DIR = os.environ.get("ML_DIR") or os.path.join(SHARE, "ML")
ADMS_DIR = os.environ.get("ADMS_DIR") or os.path.join(SHARE, "ADMS")
TPMS_DIR = os.environ.get("TPMS_DIR") or os.path.join(SHARE, "TPMS")
ML_SCRIPTS = os.path.join(ML_DIR, "scripts")
if not os.path.isdir(ML_SCRIPTS):
    sys.exit(f"[E] cannot find the forward ML scripts at {ML_SCRIPTS}")
sys.path.insert(0, ML_SCRIPTS)

RESULTS = os.path.join(ID_ROOT, "search", "results")
VAL_ROOT = os.path.join(ID_ROOT, "validation")          # the one-way folder
TPMS_FEA_DIR = os.path.join(TPMS_DIR, "TPMS FEA")

NTOPCL_CANDIDATES = [
    r"C:\Program Files\nTopology\nTopology\ntopcl.exe",
    r"C:\Program Files\nTopology\ntopcl.exe",
    r"D:\Program Files\nTopology\nTopology\ntopcl.exe",
]

GATE7_TOL = 15.0        # per cent, the approved plan's threshold
REUSE = [False]         # set by --from-existing


def die(m):
    print(f"\n[E] {m}")
    sys.exit(1)


def find_ntopcl():
    p = os.environ.get("NTOPCL")
    if p and os.path.exists(p):
        return p
    for c in NTOPCL_CANDIDATES:
        if os.path.exists(c):
            return c
    die("ntopcl.exe not found. Set NTOPCL=<full path> and re-run.")


def production():
    """ntop_batch's OWN homogenisation functions. Imported, never copied - they
    are what produced every label the model was trained on."""
    if ADMS_DIR not in sys.path:
        sys.path.append(ADMS_DIR)
    import ntop_batch
    return ntop_batch


def sheet(name, key, default):
    """A material constant from ML_settings.xlsx. Nothing is hardcoded."""
    import openpyxl
    x = os.path.join(ML_DIR, "ML_settings.xlsx")
    if not os.path.exists(x):
        return default
    wb = openpyxl.load_workbook(x, data_only=True)
    if name not in wb.sheetnames:
        return default
    for r in wb[name].iter_rows(min_row=2, values_only=True):
        if r and r[0] and str(r[0]).strip() == key and r[1] not in (None, ""):
            return r[1]
    return default


def sanity(m, tag):
    """A lattice cannot be stiffer than the solid it is made of.

    0 < E/Es < 1 always. Anything outside that is a unit error, not physics -
    and it is the ONLY automatic way to catch one, because GATE 7 compares a
    prediction against a measurement and a scale error moves neither relative
    to the other. On 13/08 the ADMS run returned E/Es = 99.46 and every gate
    failed except TAI, which is a ratio and therefore immune. That asymmetry
    IS the signature: if the scale-invariant metric passes while the others
    fail together, suspect units before suspecting the model.
    """
    e = m.get("E_over_Es", float("nan"))
    if not (0.0 < e < 1.0):
        print(f"\n    *** {tag}: E/Es = {e:.6g}. A lattice CANNOT be stiffer")
        print(f"    *** than its solid, so this is a UNIT ERROR, not a result.")
        print(f"    *** Ratio to a plausible value suggests a factor of "
              f"{10 ** round(np.log10(abs(e))) if e else 0:g}.")
        print(f"    *** Do NOT report GATE 7 from this run.\n")
        return False
    return True


def metrics(C_gpa, es_gpa, nu_s):
    """E/Es, G/Gs, TAI, C11_n... exactly as ntop_batch computes the labels."""
    nb = production()
    h = nb.voigt_reuss_hill(C_gpa)
    gs = es_gpa / (2.0 * (1.0 + nu_s))
    return dict(
        E_over_Es=h["E_iso"] / es_gpa,
        G_over_Gs=h["G_H"] / gs,
        nu_iso=h["nu_iso"],
        TAI=nb.tensorial_anisotropy_index(C_gpa),
        C11_n=C_gpa[0, 0] / es_gpa,
        C12_n=C_gpa[0, 1] / es_gpa,
        C44_n=C_gpa[3, 3] / es_gpa,
        E_iso_GPa=h["E_iso"],
    )


def run_ntop(ntopcl, src_nb, work_nb, inputs, rd, dry):
    """One nTop run, driven exactly as production drives it."""
    os.makedirs(rd, exist_ok=True)
    inj, outj = os.path.join(rd, "in.json"), os.path.join(rd, "out.json")
    with open(inj, "w") as f:
        json.dump({"inputs": inputs}, f, indent=2)
    if dry:
        return None, 0.0
    if REUSE[0]:
        print("      [--from-existing: not re-running nTop, "
              "reading the CSV already on disk]")
        return 0, 0.0
    os.makedirs(os.path.dirname(work_nb), exist_ok=True)
    shutil.copy2(src_nb, work_nb)          # -s writes back: always start pristine
    t0 = time.time()
    args = [ntopcl, "-v2", "-s", "-j", inj, "-o", outj, work_nb]
    with open(os.path.join(rd, "log.txt"), "w") as lf:
        lf.write(" ".join(args) + "\n\n")
        lf.flush()
        rc = subprocess.run(args, stdout=lf, stderr=subprocess.STDOUT,
                            text=True).returncode
    return rc, time.time() - t0


def do_adms(ntopcl, w, dry, variant_hint=None):
    stem = str(w["stem"])
    topo = variant_hint or stem.split("_")[2]           # ID2_ADMS_<topo>_...
    name = {"df": "ADMS_DF_only_FEA.ntop", "flow": "ADMS_Flow_only_FEA.ntop",
            "raw": "ADMS_Raw_only_FEA.ntop"}.get(topo.lower())
    src = os.path.join(ADMS_DIR, name) if name else None
    if not (src and os.path.exists(src)):
        c = [p for p in glob.glob(os.path.join(ADMS_DIR, "ADMS_*_only_FEA.ntop"))
             if topo.lower() in os.path.basename(p).lower()]
        if len(c) != 1:
            die(f"cannot resolve the only_FEA notebook for {topo!r}: {c}")
        src = c[0]
    rd = os.path.join(VAL_ROOT, "adms", stem)
    out_csv = os.path.join(rd, stem + ".csv")
    inputs = [
        {"description": "", "name": "Density", "type": "real",
         "value": float(w["D"])},
        {"description": "", "name": "Thickness", "type": "real", "units": "mm",
         "value": float(w["t"])},
        {"description": "", "name": "Seed", "type": "integer",
         "value": int(w["seed"])},
        {"description": "", "name": "Out Path", "type": "file_path",
         "value": out_csv.replace("\\", "/")},
        {"description": "", "name": "Inner Size", "type": "real", "units": "mm",
         "value": float(w["inner"])},
        {"description": "", "name": "Size Multi", "type": "real",
         "value": float(w["size_multi"])},
        {"description": "", "name": "Stress Path", "type": "file_path",
         "value": os.path.join(rd, stem + "_stress.csv").replace("\\", "/")},
        {"description": "", "name": "Shear Stress Path", "type": "file_path",
         "value": os.path.join(rd, stem + "_shear_stress.csv").replace("\\", "/")},
    ]
    print(f"    notebook  {os.path.basename(src)}")
    print(f"    inputs    Density {w['D']}  Thickness {w['t']}  Seed {int(w['seed'])}"
          f"  Inner Size {w['inner']}  Size Multi {w['size_multi']}")
    print(f"    writes    {rd}")
    rc, secs = run_ntop(ntopcl, src, os.path.join(rd, "_working.ntop"),
                        inputs, rd, dry)
    if dry:
        return None, None, 0.0, out_csv
    if rc != 0 or not os.path.exists(out_csv):
        print(f"      FAILED (exit {rc}, {secs / 60:.1f} min) - see {rd}\\log.txt")
        return None, None, secs, out_csv
    # ntop_batch.read_c_tensor_csv returns the tensor in MPa. The /1000 that
    # makes it GPa lives at the CALL SITE - ntop_batch.py line 581, inside
    # analyse_c_tensor:
    #       C = read_c_tensor_csv(csv_path) / 1000.0     # MPa -> GPa
    # Calling the reader without that division is what produced E/Es = 99.46 on
    # 13/08: a 1000x error that every scale-dependent metric inherited and that
    # only TAI survived, because TAI is a ratio. Importing a production function
    # is not enough - you have to import its call path too.
    try:
        C = production().read_c_tensor_csv(out_csv) / 1000.0
    except Exception as ex:
        print(f"      could not read the C tensor: {type(ex).__name__}: {ex}")
        C = None
    if C is None:
        print(f"      could not parse a 6x6 from {out_csv}")
        return None, None, secs, out_csv
    es = float(sheet("labels", "es_gpa", 200.0))
    nu = float(sheet("labels", "nu_s", 0.3))
    print(f"    material  Es {es} GPa   nu_s {nu}   (ML_settings 'labels' sheet)")
    return C, metrics(np.asarray(C, float), es, nu), secs, out_csv


def do_tpms(ntopcl, w, dry, edge_c):
    stem = str(w["stem"])
    topo = stem.replace("ID2_", "").rsplit("_", 1)[0].replace("_", " ")
    c = glob.glob(os.path.join(TPMS_FEA_DIR, "*", f"{topo}*.ntop"))
    if len(c) != 1:
        die(f"cannot resolve ONE TPMS FEA notebook for {topo!r} under "
            f"{TPMS_FEA_DIR}: {c}")
    src = c[0]
    rd = os.path.join(VAL_ROOT, "tpms", stem)
    out_csv = os.path.join(rd, stem + ".csv")
    t = float(w["t_max"])
    edge = edge_c * t
    inputs = [
        {"name": "t_max", "type": "scalar", "values": t},
        {"name": "t_min", "type": "scalar", "values": -t},
        {"name": "Edge length", "type": "scalar", "units": "mm", "values": edge},
        {"name": "Path", "type": "file_path", "value": out_csv.replace("\\", "/"),
         "values": out_csv.replace("\\", "/")},
    ]
    print(f"    notebook  {os.path.relpath(src, TPMS_DIR)}")
    print(f"    inputs    t_max {t:.6f}  t_min {-t:.6f}  "
          f"Edge length {edge:.6f} mm")
    print(f"    mesh      Edge length = {edge_c:g} x t  ->  {edge:.6f} mm")
    print(f"              {10.0 / edge:.0f} elements across a 10 mm cell.")
    print(f"              EDGE_C {edge_c:g} is measured from the notebook saved")
    print(f"              at density 0.20 (t_max 0.7672, Edge length 0.3069).")
    print(f"    writes    {rd}")
    rc, secs = run_ntop(ntopcl, src, os.path.join(rd, "_working.ntop"),
                        inputs, rd, dry)
    if dry:
        return None, None, 0.0, out_csv
    if rc != 0 or not os.path.exists(out_csv):
        print(f"      FAILED (exit {rc}, {secs / 60:.1f} min) - see {rd}\\log.txt")
        return None, None, secs, out_csv
    import collect_tensors
    C = collect_tensors._read_raw_6x6(out_csv)          # partner's raw 6x6
    if C is None:
        print(f"      could not parse a 6x6 from {out_csv}")
        return None, None, secs, out_csv
    unit = str(sheet("tpms_labels", "tensor_unit", "kPa")).strip().lower()
    to_gpa = {"pa": 1e-9, "kpa": 1e-6, "mpa": 1e-3, "gpa": 1.0}.get(unit, 1e-6)
    es = float(sheet("tpms_labels", "es", 1800.0))
    eu = str(sheet("tpms_labels", "es_unit", "MPa")).strip().lower()
    es_gpa = es * {"pa": 1e-9, "kpa": 1e-6, "mpa": 1e-3, "gpa": 1.0}.get(eu, 1e-3)
    nu = float(sheet("tpms_labels", "nu", 0.3))
    print(f"    material  Es {es_gpa} GPa   nu_s {nu}   tensor in {unit}   "
          f"(ML_settings 'tpms_labels' sheet)")
    return (np.asarray(C, float) * to_gpa,
            metrics(np.asarray(C, float) * to_gpa, es_gpa, nu), secs, out_csv)


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--family", choices=["adms", "tpms", "both"], default="both")
    ap.add_argument("--target", type=float, default=0.10)
    ap.add_argument("--objective", default="E_over_Es")
    ap.add_argument("--edge-c", type=float, default=0.4,
                    help="TPMS FE mesh: Edge length = EDGE_C x t. 0.4 is "
                         "MEASURED from FRD Generator_density0.20.ntop "
                         "(t_max 0.7672 / Edge length 0.3069). NOT the 2 in "
                         "tpms_batch_run, which is a different model's value.")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--from-existing", action="store_true",
                    help="re-derive GATE 7 from the FEA CSVs already in "
                         "validation\\, without re-running nTop. The tensor is "
                         "the expensive part and it is already on disk.")
    a = ap.parse_args()

    fj = sorted(glob.glob(os.path.join(
        RESULTS, f"phase3_finalists_{a.objective}_*.json")))
    if not fj:
        die("no Phase 3 finalists found - run ID_4_RANK.bat first")
    fin = json.load(open(fj[-1], encoding="utf-8"))["finalists"]

    pa = sorted(glob.glob(os.path.join(RESULTS, f"phase2_adms_{a.objective}_*.csv")))
    pt = sorted(glob.glob(os.path.join(RESULTS, f"phase2_tpms_{a.objective}_*.csv")))
    rows = {}
    if pa:
        d = pd.read_csv(pa[-1])
        rows["ADMS"] = d
    if pt:
        rows["TPMS"] = pd.read_csv(pt[-1])

    REUSE[0] = bool(a.from_existing)
    ntopcl = find_ntopcl()
    os.makedirs(VAL_ROOT, exist_ok=True)

    print("=" * 78)
    print("  INVERSE DESIGN - PHASE 4   one real FEA per family   GATE 7")
    print("=" * 78)
    print(f"  ntopcl      {ntopcl}")
    print(f"  finalists   {os.path.basename(fj[-1])}")
    print(f"  writes into {VAL_ROOT}")
    print("              ONE-WAY: these rows TEST the model. They must never be")
    print("              merged into the training dataset.")
    print(f"  GATE 7      |measured - predicted| / measured <= {GATE7_TOL:g} %")

    out = {}
    for fam in ("ADMS", "TPMS"):
        if a.family != "both" and a.family.upper() != fam:
            continue
        if fam not in fin:
            print(f"\n  {fam}: no finalist, skipping")
            continue
        stem = fin[fam]["stem"]
        df = rows.get(fam)
        if df is None or stem not in set(df.stem):
            die(f"{stem} is not in the Phase 2 {fam} results")
        w = df[df.stem == stem].iloc[0]
        print("\n" + "-" * 78)
        print(f"  {fam}   {stem}")
        print("-" * 78)
        if fam == "ADMS":
            C, m, secs, csvp = do_adms(ntopcl, w, a.dry_run)
        else:
            C, m, secs, csvp = do_tpms(ntopcl, w, a.dry_run, a.edge_c)
        if a.dry_run:
            print("    [dry-run] nothing run")
            continue
        if m is None:
            out[fam] = dict(stem=stem, ok=False)
            continue
        if not sanity(m, fam):
            out[fam] = dict(stem=stem, ok=False, reason="unit error")
            continue
        print(f"\n    FEA completed in {secs / 60:.1f} min"
              if secs else "\n    read from the existing FEA output")
        print(f"    {'target':<12} {'predicted':>12} {'MEASURED':>12} {'diff %':>9}  GATE 7")
        rec = dict(stem=stem, ok=True, secs=round(secs, 1), csv=csvp,
                   measured={k: float(v) for k, v in m.items()},
                   predicted={}, gate7={})
        for k in ("E_over_Es", "G_over_Gs", "TAI", "C11_n", "C12_n", "C44_n"):
            pk = f"pred_{k}"
            if pk not in w or k not in m:
                continue
            pv, mv = float(w[pk]), float(m[k])
            d = (pv - mv) / mv * 100.0 if mv else float("nan")
            ok7 = abs(d) <= GATE7_TOL
            rec["predicted"][k] = pv
            rec["gate7"][k] = bool(ok7)
            mark = "PASS" if ok7 else "FAIL"
            star = "  <-- the objective" if k == a.objective else ""
            print(f"    {k:<12} {pv:12.6f} {mv:12.6f} {d:+9.2f}  {mark}{star}")
        primary = rec["gate7"].get(a.objective)
        rec["gate7_primary"] = bool(primary)
        print(f"\n    GATE 7 on {a.objective}: "
              f"{'PASS' if primary else 'FAIL'}")
        out[fam] = rec

    if a.dry_run:
        print("\n  [dry-run] nothing run")
        return

    fout = os.path.join(VAL_ROOT, f"phase4_gate7_{a.objective}_{a.target:g}.json")
    with open(fout, "w") as f:
        json.dump(dict(built_utc=time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()),
                       gate7_tol_pct=GATE7_TOL, edge_c=a.edge_c,
                       objective=a.objective, target=a.target,
                       families=out), f, indent=2)

    print("\n" + "=" * 78)
    print("  GATE 7 SUMMARY")
    print("=" * 78)
    for fam, r in out.items():
        if not r.get("ok"):
            print(f"    {fam:<5} FEA FAILED")
            continue
        pv = r["predicted"].get(a.objective)
        mv = r["measured"].get(a.objective)
        d = (pv - mv) / mv * 100.0
        print(f"    {fam:<5} predicted {pv:.6f}   measured {mv:.6f}   "
              f"{d:+.2f} %   {'PASS' if r['gate7_primary'] else 'FAIL'}")
    print(f"\n  -> {fout}")
    print("\n  This is the only number in the whole inverse design that is not a")
    print("  prediction. Report it as the validation of the method, and report it")
    print("  whichever way it came out.")
    print("=" * 78)


if __name__ == "__main__":
    main()
