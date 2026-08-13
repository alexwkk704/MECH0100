"""
chart_builder.py
================
Rebuild the master 'charts' + 'chart_config' sheets with matplotlib-rendered
PNGs. Called from:

  - refresh_charts.py  (via REFRESH_CHARTS.bat)  -> label-only edits
  - ntop_batch.py update_master()                 -> after every sim batch

Alex-editable knobs live on the 'chart_config' sheet of the master xlsx.
Blank cell = use default. Edit the cell, save, run REFRESH_CHARTS.bat.
"""
from __future__ import annotations
import math, os, re
from collections import defaultdict
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XlImage
import matplotlib
matplotlib.use("Agg")  # no display needed
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.colors as mcolors

# ------------------------------------------------------------------
# style
# ------------------------------------------------------------------
plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "font.size":         12,
    "axes.titlesize":    14,
    "axes.titleweight":  "bold",
    "axes.labelsize":    13,
    "axes.labelweight":  "medium",
    "xtick.labelsize":   11,
    "ytick.labelsize":   11,
    "legend.fontsize":   10,
    "legend.frameon":    True,
    "legend.framealpha": 0.9,
    "figure.dpi":        140,
    "savefig.dpi":       160,
    "savefig.bbox":      "tight",
    "axes.grid":         True,
    "grid.alpha":        0.25,
    "grid.linestyle":    "--",
    "axes.spines.top":   False,
    "axes.spines.right": False,
})

# ------------------------------------------------------------------
# solid-material modulus used to turn normalised E/Es back into GPa.
# NOT a constant: ntop_batch overwrites this from the 'settings' sheet of
# runs_input.xlsx (es_gpa) before calling rebuild_charts, so a different
# base material (e.g. a polymer) plots correctly without editing this file.
# 200.0 GPa = steel, the value used for the ADMS dataset.
# ------------------------------------------------------------------
ES_GPA = 200.0

# ------------------------------------------------------------------
# default chart config (Alex overrides via 'chart_config' sheet)
# ------------------------------------------------------------------
CHART_DEFAULTS = [
    # id, title, x_label, y_label, subtitle
    ("stiffness_E_over_Es",
     "Stiffness scaling: how E depends on relative density",
     "Relative density ρ_rel",
     "Normalised stiffness  E_iso / E_s",
     "",  # subtitle filled with live fit equation at render time
     "Colour = wall thickness; marker shape = ADMS variant. One fitted law per variant. "
     "Fits use the standard family only (Inner_Size 15, Size_Multi 1.6); off-family runs "
     "(i9/i12, other multipliers) are PLOTTED but excluded from the fits — they follow "
     "different curves due to RVE / boundary-cut effects."),
    ("shear_G_over_Gs",
     "Shear stiffness: how G depends on relative density",
     "Relative density ρ_rel",
     "Normalised shear modulus  G_iso / G_s",
     "",
     "Same ρ-exponent as stiffness → isotropic scaling."),
    ("yield_onset",
     "Yield strength: the applied load a lattice can carry before permanent deformation begins",
     "Relative density ρ_rel",
     "Onset stress (MPa)",
     "",
     "Gibson-Ashby plastic-collapse exponent (≈1.5)."),
    ("rve_convergence",
     "RVE convergence: how the analysed cube size affects stiffness and isotropy",
     "Inner envelope size (mm)",
     "Normalised metric",
     "Baseline: D=20.83, t=0.4, m=1.6, seed 2",
     "Left: E vs scaling-law prediction. Right: TAI. Standard 5x5x5 = 15 mm."),
    ("parameter_effects",
     "Which design parameter drives stiffness the most? — fold ΔE per parameter",
     "Fold ΔE (max / min)",
     "",
     "At fixed baseline of the other four parameters.",
     "Thickness ≈ Density (both ρ-driving) > Inner_Size (structural) > Size_Multi (weak) > Seed (INERT)."),
]

HDR_FONT = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="500778")


# ------------------------------------------------------------------
# data / fit helpers
# ------------------------------------------------------------------
def _load_rows(ws_sum):
    h = [c.value for c in ws_sum[1]]
    rows = []
    for r in range(2, ws_sum.max_row + 1):
        d = {k: ws_sum.cell(r, i + 1).value for i, k in enumerate(h)}
        if d.get("Run"):
            rows.append(d)
    return rows


def _dedup_highest(rs):
    g = defaultdict(list)
    for r in rs:
        g[_cfg_key(r)].append(r)
    return [sorted(gv, key=lambda r: int(str(r.get("Version") or "v0")[1:] or 0))[-1]
            for gv in g.values()]


def _adms_type(r):
    """ADMS variant for a row: 'df' | 'raw' | 'flow'.

    Parsed from the Run stem, which the batch script writes as
    ADMS_<type>_i..._d..._t..._s..._m...  Legacy rows (5cube_*, pre-multi-type)
    have no type token and are DF by definition.
    """
    m = re.search(r"ADMS[_ ]?(DF|raw|flow)", str(r.get("Run") or ""), re.I)
    return m.group(1).lower() if m else "df"


def _cfg_key(r):
    """Config identity INCLUDING ADMS type.

    Type must be in the key: DF and flow can share the same
    (Density, Thickness, Inner_Size, Size_Multi, Seed) yet be completely
    different geometries. Without the type, the version-dedup below would
    treat them as re-runs of one config and silently drop one of them.
    """
    return (_adms_type(r), r.get("Density"), r.get("Thickness"),
            r.get("Inner_Size"), r.get("Size_Multi"), r.get("Seed"))


def _dedup_distinct(rs):
    """Alex's rule (2026-07-17): within one 5-param config, drop bit-identical
    repeat runs (they add no information) but KEEP repeats whose results
    differ (FE re-mesh wiggle = independent measurements)."""
    g = defaultdict(list)
    for r in rs:
        g[_cfg_key(r)].append(r)
    out = []
    for gv in g.values():
        seen = set()
        for r in gv:
            sig = (r.get("Rho_rel"), r.get("E_iso_GPa"),
                   r.get("G_iso_GPa"), r.get("Yield_onset_MPa"))
            if sig in seen:
                continue
            seen.add(sig)
            out.append(r)
    return out


def _fit_family(rs, adms_type="df"):
    """Standard family for the quotable scaling laws: 5x5x5 RVE
    (Inner_Size=15) at Size_Multi=1.6, for ONE ADMS variant.

    Off-family runs (i9/i12, other multipliers) are valid results but follow
    different curves (RVE / boundary-cut effects) and would bias the law.

    ADMS TYPE FILTER (added 2026-07-18): raw/flow runs are also i15 & m1.6, so
    without this they were fitted together with DF and shifted the published
    law (0.863 -> 0.848). Each variant has its own law:
        DF   E/Es = 0.863*rho^1.83
        flow E/Es = 0.609*rho^1.66
    Pass adms_type=None to fit across all variants deliberately.
    """
    out = [r for r in rs
           if r.get("Inner_Size") == 15
           and r.get("Size_Multi") is not None
           and abs(r.get("Size_Multi") - 1.6) < 0.01]
    if adms_type is not None:
        out = [r for r in out if _adms_type(r) == adms_type.lower()]
    return out


def _powerfit(xs, ys):
    if len(xs) < 3:
        return None, None, None, len(xs)
    xs, ys = np.array(xs), np.array(ys)
    lx, ly = np.log(xs), np.log(ys)
    n, c = np.polyfit(lx, ly, 1)
    yh = np.exp(n * lx + c)
    ss_tot = float(np.sum((ys - ys.mean()) ** 2))
    r2 = 1 - float(np.sum((ys - yh) ** 2)) / ss_tot if ss_tot else 0
    return math.exp(c), float(n), r2, len(xs)


def _safe_pairs(rows, key):
    """Return xs, ys, thicknesses, inner_sizes, adms_types (all same length)."""
    xs, ys, ts, i_s, tp = [], [], [], [], []
    for r in rows:
        rho, y = r.get("Rho_rel"), r.get(key)
        if rho is None or y is None or rho <= 0:
            continue
        xs.append(rho); ys.append(y)
        ts.append(r.get("Thickness")); i_s.append(r.get("Inner_Size"))
        tp.append(_adms_type(r))
    return xs, ys, ts, i_s, tp


# ------------------------------------------------------------------
# chart_config sheet
# ------------------------------------------------------------------
def _ensure_chart_config(wb):
    """Create or read the 'chart_config' sheet. Return a dict:
       chart_id -> {title, xlabel, ylabel, subtitle, notes}.
    Preserves user overrides; fills blanks with defaults."""
    existing = {}
    if "chart_config" in wb.sheetnames:
        ws_old = wb["chart_config"]
        for r in range(2, ws_old.max_row + 1):
            cid = ws_old.cell(r, 1).value
            if not cid:
                continue
            existing[cid] = {
                "title":    ws_old.cell(r, 2).value or "",
                "xlabel":   ws_old.cell(r, 3).value or "",
                "ylabel":   ws_old.cell(r, 4).value or "",
                "subtitle": ws_old.cell(r, 5).value or "",
                "notes":    ws_old.cell(r, 6).value or "",
            }
        del wb["chart_config"]

    ws = wb.create_sheet("chart_config", 1)
    headers = ["Chart ID", "Chart title (editable)", "X-axis label (editable)",
               "Y-axis label (editable)", "Subtitle / footnote (editable)",
               "Notes / description"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h); c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    cfg = {}
    for i, (cid, t0, xl0, yl0, sub0, notes0) in enumerate(CHART_DEFAULTS, 2):
        e = existing.get(cid, {})
        row = {
            "title":    e.get("title")    or t0,
            "xlabel":   e.get("xlabel")   or xl0,
            "ylabel":   e.get("ylabel")   or yl0,
            "subtitle": e.get("subtitle") or sub0,
            "notes":    e.get("notes")    or notes0,
        }
        ws.cell(i, 1, cid); ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 2, row["title"])
        ws.cell(i, 3, row["xlabel"])
        ws.cell(i, 4, row["ylabel"])
        ws.cell(i, 5, row["subtitle"])
        ws.cell(i, 6, row["notes"])
        cfg[cid] = row

    # instructions
    ir = len(CHART_DEFAULTS) + 4
    ws.cell(ir, 1, "HOW TO USE").font = Font(bold=True, color="500778", size=12)
    ws.cell(ir + 1, 1, "1) Edit any cell above to change how a chart is labelled.").font = Font(italic=True)
    ws.cell(ir + 2, 1, "2) Save this file.").font = Font(italic=True)
    ws.cell(ir + 3, 1, "3) Double-click REFRESH_CHARTS.bat to re-render (~5 seconds).").font = Font(italic=True)
    ws.cell(ir + 4, 1, "4) Or wait — charts auto-refresh after every RUN_SIMULATION.bat.").font = Font(italic=True)
    ws.cell(ir + 6, 1, "Reset a cell to default: clear it (delete contents). The default text will fill in next refresh.").font = Font(italic=True, color="666666")

    for col_letter, width in zip("ABCDEF", [22, 55, 28, 32, 45, 60]):
        ws.column_dimensions[col_letter].width = width
    for r in range(2, 2 + len(CHART_DEFAULTS)):
        ws.row_dimensions[r].height = 42
        for cl in "BCDEF":
            ws[f"{cl}{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    return cfg


# ------------------------------------------------------------------
# chart renderers
# ------------------------------------------------------------------
def _viridis_for(t_vals):
    ts = sorted(set(t for t in t_vals if t is not None))
    if not ts: return {}
    if len(ts) == 1: return {ts[0]: cm.viridis(0.5)}
    tmin, tmax = min(ts), max(ts)
    return {t: cm.viridis((t - tmin) / (tmax - tmin)) for t in ts}


# marker shape encodes the ADMS variant; colour still encodes wall thickness
TYPE_MARKER = {"df": "o", "flow": "^", "raw": "s"}
TYPE_LABEL = {"df": "ADMS+DF", "flow": "ADMS flow", "raw": "ADMS raw"}


def _render_scaling(fname, cfg_row, xs, ys, ts, C, n, R2, N, tps=None,
                    type_fits=None):
    fig, ax = plt.subplots(figsize=(11, 7.2))
    if tps is None:
        tps = ["df"] * len(xs)
    # Colour = wall thickness (CONTINUOUS colourbar, not one legend entry per
    # value — with 3 variants x ~15 thicknesses that produced a 40-entry legend
    # that swamped the plot). Marker SHAPE = ADMS variant.
    tvals = [t for t in ts if t is not None]
    tmin, tmax = (min(tvals), max(tvals)) if tvals else (0.0, 1.0)
    norm = mcolors.Normalize(vmin=tmin, vmax=tmax if tmax > tmin else tmin + 1)
    by_tp = defaultdict(list)
    for x, y, t, tp in zip(xs, ys, ts, tps):
        by_tp[tp].append((x, y, t))
    for tp in ("df", "flow", "raw"):
        pts = by_tp.get(tp)
        if not pts:
            continue
        ax.scatter([p[0] for p in pts], [p[1] for p in pts],
                   s=90,
                   c=[cm.viridis(norm(p[2])) if p[2] is not None else "#666" for p in pts],
                   marker=TYPE_MARKER.get(tp, "o"),
                   edgecolor="white", linewidth=0.7,
                   label=TYPE_LABEL[tp], alpha=0.95, zorder=3)
    sm = cm.ScalarMappable(cmap=cm.viridis, norm=norm)
    sm.set_array([])
    cbar = fig.colorbar(sm, ax=ax, pad=0.02)
    cbar.set_label("wall thickness  t (mm)")

    # One fitted line PER ADMS variant. Plotting points while fitting only a
    # subset (without saying so) misleads the reader; fitting a single line
    # through several geometry families produces a law describing none of them.
    # So: every plotted family gets its own visible, labelled law.
    TYPE_LINE = {"df": "--", "flow": "--", "raw": "--"}
    TYPE_LINECOLOR = {"df": "black", "flow": "#b34700", "raw": "#00666e"}
    eq_lines = []
    if type_fits:
        for tp in ("df", "flow", "raw"):
            f = type_fits.get(tp)
            if not f:
                continue
            Ct, nt, R2t, Nt = f
            if Ct is None:
                continue
            xs_t = [x for x, p in zip(xs, tps) if p == tp]
            if not xs_t:
                continue
            x_fit = np.linspace(max(min(xs_t) * 0.9, 0.03), max(xs_t) * 1.05, 200)
            ax.plot(x_fit, Ct * x_fit ** nt, TYPE_LINE.get(tp, "--"),
                    color=TYPE_LINECOLOR.get(tp, "black"), linewidth=1.8,
                    label=f"{TYPE_LABEL[tp]} fit", zorder=2)
            eq_lines.append(f"{TYPE_LABEL[tp]}:  y = {Ct:.3f}·ρ^{nt:.2f}"
                            f"   R² = {R2t:.3f}, N = {Nt}")
    elif C is not None and xs:
        x_fit = np.linspace(max(min(xs) * 0.9, 0.03), max(xs) * 1.05, 200)
        ax.plot(x_fit, C * x_fit ** n, "k--", linewidth=1.8,
                label="fitted law (dashed)", zorder=2)

    ax.set_title(cfg_row["title"])
    ax.set_xlabel(cfg_row["xlabel"])
    ax.set_ylabel(cfg_row["ylabel"])

    # subtitle = user override, ELSE one equation line per fitted variant
    # (never a single unlabelled equation when several families are plotted —
    # the reader must be able to see which law belongs to which data).
    if cfg_row["subtitle"]:
        subtitle = cfg_row["subtitle"]
    elif eq_lines:
        subtitle = "\n".join(eq_lines)
    elif C is not None:
        subtitle = f"Fit: y = {C:.3f}·ρ^{n:.2f}    R² = {R2:.3f}, N = {N}"
    else:
        subtitle = f"N = {N} (fit unavailable)"
    ax.text(0.98, 0.02, subtitle, transform=ax.transAxes, fontsize=10.5,
            verticalalignment="bottom", horizontalalignment="right",
            style="italic", color="#333333",
            bbox=dict(boxstyle="round,pad=0.4", facecolor="white",
                      edgecolor="#cccccc", alpha=0.85))

    ax.legend(loc="upper left", framealpha=0.9)
    if xs:
        ax.set_xlim(0.04, max(xs) * 1.08)
    ax.set_ylim(bottom=0)
    plt.tight_layout(); plt.savefig(fname); plt.close(fig)


def _render_rve(fname, cfg_row, dedup, C_E, n_E):
    baseline = {"Density": 20.83, "Thickness": 0.4, "Size_Multi": 1.6, "Seed": 2}
    ms = [r for r in dedup
          if all(r.get(k) == v for k, v in baseline.items())
          and r.get("Inner_Size") is not None]
    ms.sort(key=lambda r: r["Inner_Size"])
    if not ms:
        return  # no baseline data yet

    xs   = [r["Inner_Size"] for r in ms]
    e_ac = [r["E_iso_GPa"]  for r in ms]
    rhos = [r["Rho_rel"]    for r in ms]
    tais = [r.get("TAI")    for r in ms]

    if C_E is None:
        return
    Es = ES_GPA
    e_pred = [C_E * (rho ** n_E) * Es for rho in rhos]
    e_ratio = [ea / ep for ea, ep in zip(e_ac, e_pred)]

    fig, (a1, a2) = plt.subplots(1, 2, figsize=(14, 5.5))
    fig.suptitle(cfg_row["title"], fontsize=14, fontweight="bold", y=1.02)
    if cfg_row["subtitle"]:
        fig.text(0.5, 0.96, cfg_row["subtitle"], ha="center",
                 fontsize=10.5, style="italic", color="#333333")

    a1.plot(xs, e_ratio, "o-", color="#500778", linewidth=2, markersize=11,
            markeredgecolor="white", markeredgewidth=0.9)
    for x, y in zip(xs, e_ratio):
        a1.annotate(f"{y:.2f}", (x, y), xytext=(10, 5),
                    textcoords="offset points", fontsize=10)
    a1.axhline(1.0, linestyle="--", color="grey", alpha=0.6, linewidth=1)
    a1.set_title("Stiffness convergence to scaling law", fontsize=13)
    a1.set_xlabel(cfg_row["xlabel"])
    a1.set_ylabel("E_iso / scaling-law prediction")
    a1.set_xticks(sorted(set(xs) | {9, 12, 15, 18}))
    a1.set_xlim(min(xs) - 1, max(max(xs), 18) + 1)

    a2.plot(xs, tais, "s-", color="#AC1414", linewidth=2, markersize=11,
            markeredgecolor="white", markeredgewidth=0.9)
    for x, y in zip(xs, tais):
        if y is not None:
            a2.annotate(f"{y:.3f}", (x, y), xytext=(10, 5),
                        textcoords="offset points", fontsize=10)
    a2.set_title("Anisotropy vs RVE size", fontsize=13)
    a2.set_xlabel(cfg_row["xlabel"])
    a2.set_ylabel("TAI (lower = more isotropic)")
    a2.set_xticks(sorted(set(xs) | {9, 12, 15, 18}))
    a2.set_xlim(min(xs) - 1, max(max(xs), 18) + 1)

    plt.tight_layout(); plt.savefig(fname); plt.close(fig)


def _render_parameter_effects(fname, cfg_row, dedup):
    SLICES = {
        "Density":    {"Thickness": 0.4,   "Inner_Size": 15, "Size_Multi": 1.6, "Seed": 2},
        "Thickness":  {"Density": 20.83,   "Inner_Size": 15, "Size_Multi": 1.6, "Seed": 2},
        "Inner_Size": {"Density": 20.83,   "Thickness": 0.4, "Size_Multi": 1.6, "Seed": 2},
        "Size_Multi": {"Density": 20.83,   "Thickness": 0.4, "Inner_Size": 15,  "Seed": 2},
        "Seed":       {"Density": 20.83,   "Thickness": 0.4, "Inner_Size": 15,  "Size_Multi": 1.6},
    }
    P_COLOURS = {"Density":"#500778","Thickness":"#D66A00","Inner_Size":"#007E67",
                 "Size_Multi":"#0072CE","Seed":"#AC145A"}
    folds = []
    for param, base in SLICES.items():
        ms = [r for r in dedup
              if all(r.get(k) == v for k, v in base.items())
              and r.get(param) is not None and r.get("E_iso_GPa") is not None]
        if not ms: continue
        es = [r["E_iso_GPa"] for r in ms]
        folds.append((param, max(es) / min(es)))
    folds.sort(key=lambda p: -p[1])
    if not folds:
        return
    names = [p[0] for p in folds]; vals = [p[1] for p in folds]
    colours = [P_COLOURS[n] for n in names]

    fig, ax = plt.subplots(figsize=(11, 5.5))
    bars = ax.barh(names[::-1], vals[::-1], color=colours[::-1],
                   edgecolor="white", linewidth=1.5)
    for bar, v in zip(bars, vals[::-1]):
        ax.text(bar.get_width() + 0.15, bar.get_y() + bar.get_height() / 2,
                f"{v:.2f}×", va="center", fontsize=12, fontweight="bold")
    ax.set_title(cfg_row["title"])
    ax.set_xlabel(cfg_row["xlabel"])
    ax.set_xlim(0, max(vals) * 1.18)
    if cfg_row["subtitle"]:
        ax.text(0.98, 0.05, cfg_row["subtitle"], transform=ax.transAxes,
                ha="right", fontsize=10.5, style="italic", color="#333333")
    plt.tight_layout(); plt.savefig(fname); plt.close(fig)


# ------------------------------------------------------------------
# public entry point
# ------------------------------------------------------------------
def rebuild_charts(master_xlsx_path):
    """Read summary + chart_config from master, re-render all PNGs, embed
    in 'charts' sheet. Idempotent, safe to call after every batch."""
    master = Path(master_xlsx_path)
    if not master.exists():
        print(f"[chart] master not found: {master}")
        return
    charts_dir = master.parent / "_charts"
    charts_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(master)
    if "summary" not in wb.sheetnames:
        print("[chart] no 'summary' sheet, skipping")
        return
    all_rows = _load_rows(wb["summary"])
    dedup = _dedup_highest(all_rows)          # latest-version rows (RVE / param sheets)
    distinct = _dedup_distinct(all_rows)      # all distinct runs (scatter)
    # Fitted laws = standard family of ONE variant (DF). raw/flow are plotted in
    # the scatter but must not enter the DF law (they have their own exponents).
    fit_rows = _fit_family(distinct, adms_type="df")
    _tcount = defaultdict(int)
    for r in distinct:
        _tcount[_adms_type(r)] += 1
    if len(_tcount) > 1:
        print(f"[chart] ADMS variants present: {dict(_tcount)} — "
              f"fitted law uses DF only (N={len(fit_rows)})")

    cfg = _ensure_chart_config(wb)

    E_xs, E_ys, E_ts, _, E_tp = _safe_pairs(distinct, "E_over_Es")
    G_xs, G_ys, G_ts, _, G_tp = _safe_pairs(distinct, "G_over_Gs")
    O_xs, O_ys, O_ts, _, O_tp = _safe_pairs(distinct, "Yield_onset_MPa")
    fE_xs, fE_ys, _, _, _ = _safe_pairs(fit_rows, "E_over_Es")
    fG_xs, fG_ys, _, _, _ = _safe_pairs(fit_rows, "G_over_Gs")
    fO_xs, fO_ys, _, _, _ = _safe_pairs(fit_rows, "Yield_onset_MPa")
    C_E, n_E, R2_E, N_E = _powerfit(fE_xs, fE_ys)
    C_G, n_G, R2_G, N_G = _powerfit(fG_xs, fG_ys)
    C_O, n_O, R2_O, N_O = _powerfit(fO_xs, fO_ys)

    # Per-variant laws: each geometry family gets its OWN fitted line, so no
    # plotted data is silently excluded from fitting and no single line is
    # passed off as describing several families at once.
    def _fits_by_type(key):
        out = {}
        for tp in ("df", "flow", "raw"):
            fam = _fit_family(distinct, adms_type=tp)
            xs_t, ys_t, _, _, _ = _safe_pairs(fam, key)
            if len(xs_t) >= 3:
                out[tp] = _powerfit(xs_t, ys_t)
        return out

    tf_E = _fits_by_type("E_over_Es")
    tf_G = _fits_by_type("G_over_Gs")
    tf_O = _fits_by_type("Yield_onset_MPa")
    for tp, (Ct, nt, R2t, Nt) in sorted(tf_E.items()):
        print(f"[chart] law {tp:>4}: E/Es = {Ct:.3f}·ρ^{nt:.2f}  "
              f"R²={R2t:.3f}  N={Nt}")

    png_files = []
    _render_scaling(charts_dir / "01_stiffness.png",
                    cfg["stiffness_E_over_Es"],
                    E_xs, E_ys, E_ts, C_E, n_E, R2_E, N_E, tps=E_tp, type_fits=tf_E)
    png_files.append("01_stiffness.png")

    _render_scaling(charts_dir / "02_shear_stiffness.png",
                    cfg["shear_G_over_Gs"],
                    G_xs, G_ys, G_ts, C_G, n_G, R2_G, N_G, tps=G_tp, type_fits=tf_G)
    png_files.append("02_shear_stiffness.png")

    _render_scaling(charts_dir / "03_yield_onset.png",
                    cfg["yield_onset"],
                    O_xs, O_ys, O_ts, C_O, n_O, R2_O, N_O, tps=O_tp, type_fits=tf_O)
    png_files.append("03_yield_onset.png")

    _render_rve(charts_dir / "04_rve_convergence.png",
                cfg["rve_convergence"], dedup, C_E, n_E)
    if (charts_dir / "04_rve_convergence.png").exists():
        png_files.append("04_rve_convergence.png")

    _render_parameter_effects(charts_dir / "05_parameter_effects.png",
                               cfg["parameter_effects"], dedup)
    if (charts_dir / "05_parameter_effects.png").exists():
        png_files.append("05_parameter_effects.png")

    # embed in charts sheet
    if "charts" in wb.sheetnames:
        del wb["charts"]
    ws_ch = wb.create_sheet("charts", 2)
    ws_ch["A1"] = ("Master charts — matplotlib PNGs. Edit labels on 'chart_config' sheet "
                   "then run REFRESH_CHARTS.bat.")
    ws_ch["A1"].font = Font(bold=True, color="500778", size=12)
    ws_ch["A2"] = f"Standalone PNGs in {charts_dir.name}/ for direct use in slides."
    ws_ch["A2"].font = Font(italic=True, size=10)

    row = 4
    for fname in png_files:
        img = XlImage(str(charts_dir / fname))
        img.width = 900; img.height = 600
        ws_ch.add_image(img, f"A{row}")
        row += 40

    wb.save(master)
    print(f"[chart] {len(png_files)} PNGs rebuilt in {charts_dir}")
    print(f"[chart] fits — E/Es={C_E:.3f}·ρ^{n_E:.2f}  G/Gs={C_G:.3f}·ρ^{n_G:.2f}  "
          f"onset={C_O:.1f}·ρ^{n_O:.2f}" if C_E else "[chart] fits skipped (insufficient data)")


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        str(Path(__file__).resolve().parent / "Results" / "Results_summary.xlsx")
    rebuild_charts(path)