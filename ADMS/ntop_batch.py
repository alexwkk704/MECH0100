"""
ntop_batch.py - self-contained nTop Automate batch runner + post-processor.

Package prepared by Alex Wong (UCL MECH0100) for sharing.
Everything is relative to THIS folder:

    <this folder>/
      RUN_SIMULATION.bat        <- double-click to run
      ntop_batch.py             <- this script
      runs_input.xlsx           <- type your parameters here (sheet 'runs')
                                   material settings on sheet 'settings'
      ADMS_DF_generic_v1.ntop   <- nTop notebook (homogenisation + static compression + shear)
      input_template_<type>.json <- notebook input schema (auto-regenerated when
                                   missing OR when the .ntop is newer; force
                                   with --refresh-schema)
      Results/                                 <- created on first run
        Results_summary.xlsx                   <- MASTER: cumulative across every
                                                  batch, append-only, Version
                                                  column (v1/v2/...) preserves
                                                  re-runs of the same geometry
        YYYYMMDD_HH-MM_RUN/                    <- one folder per .bat invocation
          Results_summary.xlsx                 <- summary for this batch only
          Data/<run_name>/                     <- per-run: C tensor CSV, stress
                                                  point-maps (compression + shear),
                                                  cached .ntop, log, JSONs

Requirements: nTop installed + licensed (sign in to the desktop app first),
Python 3.9+ with numpy and openpyxl (the .bat installs these automatically).

Method notes (matches Week 7 work-progress presentation + Week 8 shear addition):
  * Homogenisation: 6x6 C tensor from nTop (MPa) -> Voigt-Reuss-Hill
    isotropic E, nu, G; TAI, Zener, universal anisotropy, directional E range.

  * Static COMPRESSION: bottom fully fixed, top displaced by -0.667% strain
    (TopDisp = InnerSize * -0.00667, along Z). Exports von Mises point map (Pa).
      sigma_applied  = E_iso * strain                (macroscopic normal stress)
      sigma_p99      = 99th percentile of nodal von Mises (robust peak;
                       raw max is mesh-sensitive at cut edges - reported too)
      Yield_onset    = sigma_applied * (sigma_ys / sigma_p99)
                       -> macroscopic compressive stress at which the p99
                          local stress reaches yield (linear scaling).

  * Static SHEAR (NEW): bottom fully fixed, top face imposed Ux = InnerSize *
    +0.00667, Uy=Uz=0 (simple shear along X). Exports von Mises point map (Pa).
      tau_applied    = G_iso * strain                (macroscopic shear stress)
      shear_p99      = 99th percentile of nodal von Mises under shear load
      Shear_onset    = tau_applied * (sigma_ys / shear_p99)
                       -> macroscopic shear stress at which the p99 local
                          von Mises stress reaches yield.

  Homogenize + Compression + Shear all use the SAME FE volume mesh, so the
  three analyses are directly comparable. Runtime per config is roughly
  Homogenize (~18 min) + Compression (~3 min) + Shear (~3 min) = ~24 min.
"""

from __future__ import annotations
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
from pathlib import Path

import numpy as np

# ============================================================
# locations - everything relative to this file
# ============================================================
SCRIPT_DIR   = Path(__file__).resolve().parent
NTOP_FILE    = SCRIPT_DIR / "ADMS_DF_generic_v1.ntop"
TEMPLATE_IN  = SCRIPT_DIR / "input_template.json"
INPUT_XLSX   = SCRIPT_DIR / "runs_input.xlsx"
RESULTS_DIR         = SCRIPT_DIR / "Results"
MASTER_SUMMARY_XLSX = RESULTS_DIR / "Results_summary.xlsx"  # cumulative across ALL batches

# Per-batch (this .bat invocation) subfolder + files. Set in main() before any run.
BATCH_DIR     = None   # e.g. Results/20260713_17-24_RUN
DATA_DIR      = None   # BATCH_DIR / "Data"
SUMMARY_XLSX  = None   # BATCH_DIR / "Results_summary.xlsx" (batch-only summary)


def _init_batch_dir():
    """Create Results/YYYYMMDD_HH-MM_RUN/ for this invocation and set globals."""
    global BATCH_DIR, DATA_DIR, SUMMARY_XLSX
    stamp = time.strftime("%Y%m%d_%H-%M_RUN")
    BATCH_DIR    = RESULTS_DIR / stamp
    DATA_DIR     = BATCH_DIR / "Data"
    SUMMARY_XLSX = BATCH_DIR / "Results_summary.xlsx"
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    BATCH_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print(f"[batch] output folder: {BATCH_DIR}")

# Usual install locations, tried in order. These are fallbacks, not assumptions:
# if nTop is installed elsewhere, set 'ntopcl_path' in the settings sheet of
# runs_input.xlsx (or the NTOPCL environment variable) and that wins.
NTOPCL_CANDIDATES = [
    r"C:\Program Files\nTopology\nTopology\ntopcl.exe",
    r"C:\Program Files\nTopology\ntopcl.exe",
]

NAME_OUTPATH          = "Out Path"
NAME_STRESSPATH       = "Stress Path"           # compression von Mises CSV
NAME_SHEAR_STRESSPATH = "Shear Stress Path"     # shear von Mises CSV (new)
NAME_STLPATH          = "STL Path"              # exported mesh STL (Export Mesh block; ignored if notebook lacks it)
NAME_TABLEPATH        = "Table Path"            # curvature table CSV (Export Table block; ignored if notebook lacks it)
FILE_PREFIX           = "ADMS_DF"

# defaults if the settings sheet is missing values
DEFAULT_SETTINGS = {"sigma_ys_mpa": 250.0, "es_gpa": 200.0, "nu_s": 0.3, "strain": 0.00667,
                    "adms_type": "DF", "ntop_file": "ADMS_DF_generic_v1.ntop"}

# Live copies of the material properties, set by main() from the settings sheet.
# The summary/chart builders run outside main()'s scope, so they read these
# instead of re-hardcoding steel. Change the material in runs_input.xlsx, NOT here.
ES_GPA_ACTIVE = DEFAULT_SETTINGS["es_gpa"]
# Stiffness exponent used only to draw the reference line on the RVE chart.
# ADMS-DF fitted value; override via the settings sheet key 'n_stiff'.
N_STIFF_ACTIVE = 1.83


def die(msg):
    print(f"\n[ERROR] {msg}")
    sys.exit(1)


def find_ntopcl(settings=None):
    """Locate ntopcl.exe. Priority: settings sheet > env var > usual installs > PATH."""
    from_sheet = (settings or {}).get("ntopcl_path")
    if from_sheet:
        if Path(from_sheet).exists():
            return str(from_sheet)
        die(f"ntopcl_path in runs_input.xlsx points at a file that does not "
            f"exist:\n  {from_sheet}")
    env = os.environ.get("NTOPCL")
    if env and Path(env).exists():
        return env
    for c in NTOPCL_CANDIDATES:
        if Path(c).exists():
            return c
    w = shutil.which("ntopcl")
    if w:
        return w
    die("ntopcl.exe not found. Either set 'ntopcl_path' in the settings sheet of\n"
        "runs_input.xlsx to the full path of ntopcl.exe, or set the NTOPCL\n"
        "environment variable, or install nTop to the default location.")


def tag_num(x, prefix, digits=2):
    s = f"{float(x):.{digits}f}".rstrip("0").rstrip(".")
    if "." not in s:
        s += "p0"
    return f"{prefix}{s}".replace(".", "p")


def stream_process(args, log_path=None):
    t0 = time.time()
    logf = open(log_path, "a", encoding="utf-8") if log_path else None
    proc = subprocess.Popen(
        args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
        text=True, bufsize=1, encoding="utf-8", errors="replace")
    stop_hb = threading.Event()

    def hb():
        last = time.time()
        while not stop_hb.is_set():
            time.sleep(2)
            if time.time() - last >= 30:
                print(f"  [heartbeat] {int(time.time() - t0)}s elapsed", flush=True)
                last = time.time()

    threading.Thread(target=hb, daemon=True).start()
    try:
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                print(f"  [{int(time.time() - t0):>4}s] {line}", flush=True)
                if logf:
                    logf.write(line + "\n")
    except KeyboardInterrupt:
        print("\n[ABORT] Ctrl+C - terminating ntopcl...")
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        stop_hb.set()
        if logf:
            logf.close()
        sys.exit(130)
    proc.wait()
    stop_hb.set()
    if logf:
        logf.close()
    return proc.returncode, time.time() - t0


# ============================================================
# CLI overrides (type / notebook)
# ============================================================
def _arg_value(flag):
    """--flag value  or  --flag=value. Returns None if the flag is absent."""
    for i, a in enumerate(sys.argv):
        if a == flag and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
        if a.startswith(flag + "="):
            return a.split("=", 1)[1]
    return None


def resolve_notebook_for_type(adms_type):
    """Find the .ntop next to this script whose name contains <adms_type>.

    Discovery, not a hardcoded map: any file matching ADMS_*generic*.ntop is a
    candidate and the type token is matched case-insensitively. Errors if the
    match is not unique, so a typo can never silently run the wrong notebook.
    """
    tok = str(adms_type).strip().lower()
    cands = [p for p in sorted(SCRIPT_DIR.glob("*.ntop"))
             if "generic" in p.name.lower() and tok in p.name.lower()]
    if len(cands) == 1:
        return cands[0].name
    have = [p.name for p in sorted(SCRIPT_DIR.glob("*.ntop")) if "generic" in p.name.lower()]
    die(f"--type {adms_type}: expected exactly one matching notebook, found "
        f"{len(cands)} ({[c.name for c in cands]}). Notebooks next to the "
        f"script: {have}")


# ============================================================
# notebook schema
# ============================================================
def ensure_schema(ntopcl, force=False):
    """Make sure input_template_<type>.json matches the CURRENT notebook.

    The template is the notebook's input SCHEMA: build_input_json() loads it,
    overwrites the values it wants, and hands the whole thing to ntopcl. Any
    input NOT overwritten keeps whatever the template says.

    So a stale template is silently wrong. Add an input to the .ntop (e.g.
    `Cell Size`, added 2026-08-05) and an old template simply will not contain
    it. The original code regenerated ONLY when the file was absent, so an
    edited notebook kept running against a months-old schema with no warning.
    Now the template is regenerated whenever the .ntop is NEWER than it.
    """
    stale = False
    if TEMPLATE_IN.exists() and not force:
        try:
            if NTOP_FILE.exists() and NTOP_FILE.stat().st_mtime > TEMPLATE_IN.stat().st_mtime:
                stale = True
                print(f"[schema] {NTOP_FILE.name} is NEWER than {TEMPLATE_IN.name} "
                      f"-> the notebook changed; REGENERATING the input schema.")
            else:
                return
        except OSError:
            return
    elif TEMPLATE_IN.exists() and force:
        stale = True
        print(f"[schema] --refresh-schema: regenerating {TEMPLATE_IN.name}")
    else:
        print(f"[schema] {TEMPLATE_IN.name} missing - generating with ntopcl -t ...")

    before = None
    if stale:
        try:
            before = [e["name"] for e in json.loads(
                TEMPLATE_IN.read_text(encoding="utf-8"))["inputs"]]
        except Exception:
            before = None
        TEMPLATE_IN.unlink()          # so the move below cannot be skipped

    rc, _ = stream_process([ntopcl, "-t", str(NTOP_FILE)])
    # ntopcl -t writes input_template.json into the CWD; move if needed
    cwd_tpl = Path.cwd() / "input_template.json"
    if not TEMPLATE_IN.exists() and cwd_tpl.exists():
        shutil.move(str(cwd_tpl), TEMPLATE_IN)
    if not TEMPLATE_IN.exists():
        die(f"could not generate input_template.json (ntopcl exit {rc})")

    try:
        after = [e["name"] for e in json.loads(
            TEMPLATE_IN.read_text(encoding="utf-8"))["inputs"]]
        print(f"[schema] {TEMPLATE_IN.name}: {len(after)} inputs -> {after}")
        if before is not None:
            added = [n for n in after if n not in before]
            gone = [n for n in before if n not in after]
            if added:
                print(f"[schema] NEW inputs since the old template: {added}")
            if gone:
                print(f"[schema] inputs REMOVED since the old template: {gone}")
    except Exception:
        pass


def load_schema():
    return json.loads(TEMPLATE_IN.read_text(encoding="utf-8"))


def build_input_json(input_values, dest_path, out_csv_path,
                     stress_csv_path, shear_stress_csv_path, stl_path=None,
                     table_path=None):
    data = load_schema()
    entries = {e["name"]: e for e in data["inputs"]}
    wanted = dict(input_values)
    wanted[NAME_OUTPATH] = str(out_csv_path).replace("\\", "/")
    wanted[NAME_STRESSPATH] = str(stress_csv_path).replace("\\", "/")
    wanted[NAME_SHEAR_STRESSPATH] = str(shear_stress_csv_path).replace("\\", "/")
    if stl_path is not None:
        # Robust: match the notebook's STL export input by name even if it
        # isn't exactly "STL Path" (any file_path input containing 'stl').
        stl_name = NAME_STLPATH if NAME_STLPATH in entries else next(
            (e["name"] for e in data["inputs"]
             if e.get("type") == "file_path" and "stl" in e["name"].lower()),
            NAME_STLPATH)
        wanted[stl_name] = str(stl_path).replace("\\", "/")
    if table_path is not None:
        # Curvature Export Table block. Matched by name, else any file_path
        # input containing 'table'. Ignored if the notebook lacks such an input.
        tbl_name = NAME_TABLEPATH if NAME_TABLEPATH in entries else next(
            (e["name"] for e in data["inputs"]
             if e.get("type") == "file_path" and "table" in e["name"].lower()),
            NAME_TABLEPATH)
        wanted[tbl_name] = str(table_path).replace("\\", "/")
    fed = {}
    for name, value in wanted.items():
        entry = entries.get(name)
        if not entry:
            continue
        key = "values" if "values" in entry else "value"
        entry[key] = value
        fed[name] = value
    Path(dest_path).write_text(json.dumps(data, indent=2), encoding="utf-8")
    return fed


def _first_number(obj):
    if isinstance(obj, dict):
        for k in ("val", "value", "values"):
            v = obj.get(k)
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                return float(v)
            if isinstance(v, list) and v and isinstance(v[0], (int, float)):
                return float(v[0])
        for v in obj.values():
            r = _first_number(v)
            if r is not None:
                return r
    elif isinstance(obj, list):
        for v in obj:
            r = _first_number(v)
            if r is not None:
                return r
    return None


def read_output_values(out_json_path):
    try:
        data = json.loads(Path(out_json_path).read_text(encoding="utf-8"))
    except Exception:
        return {}
    out = {}
    if isinstance(data, list):
        for entry in data:
            if not isinstance(entry, dict):
                continue
            # Table Output (the curvature Table block): columns p_rel/MeanCurv/
            # GaussCurv, values stored as strings. Map to rho_rel/meancurv/gauss.
            if entry.get("type") == "table":
                tbl = ((entry.get("value") or {}).get("table") or {})
                for col in tbl.get("columns", []):
                    nm = str(col.get("name", "")).strip().lower()
                    rws = col.get("rows") or []
                    if not rws:
                        continue
                    try:
                        v = float(rws[0])
                    except (TypeError, ValueError):
                        continue
                    # Accept BOTH naming conventions:
                    #   design-notebook style : GaussCurv / MeanCurv / p_rel
                    #   backfill style        : avg_Mn_K (Gaussian) / avg_Mn_H (mean)
                    # K = Gaussian curvature, H = mean curvature - the standard
                    # differential-geometry symbols, and the names used in
                    # ML/Curvature_Check/curvature_ntop.csv. Order matters:
                    # "GaussCurv" also contains "curv", so test gauss FIRST.
                    if "gauss" in nm or nm.endswith("_k") or "_k_" in nm:
                        out["gausscurv"] = v
                    elif "mean" in nm or "curv" in nm or nm.endswith("_h") or "_h_" in nm:
                        out["meancurv"] = v
                    elif "rel" in nm or "rho" in nm:
                        out["rho_rel"] = v
                continue
            name = str(entry.get("name", "")).strip()
            name = name.replace("ρ", "rho").replace("σ", "sigma")
            name = name.lower().replace(" ", "_")
            val = _first_number(entry.get("value"))
            if name and val is not None:
                out[name] = val
                # Curvature may also arrive as PLAIN NAMED SCALARS rather than a
                # Table (e.g. a notebook that exposes Average_Mn_K / Average_Mn_H
                # as outputs instead of building a Table block). Map those too, so
                # both notebook styles work. Only the NORMALISED "mn_" names are
                # accepted: Average_M_K (mm^-2) and Average_Mn_K (dimensionless)
                # both end in "_k", and silently taking the un-normalised one
                # would be wrong by a factor of Cell Size^2.
                if "mn_k" in name or "gausscurv" in name:
                    out["gausscurv"] = val
                elif "mn_h" in name or "meancurv" in name:
                    out["meancurv"] = val
    if not out:
        v = _first_number(data)
        if v is not None:
            out["rho_rel"] = v
    return out


def read_curv_csv(path):
    """Read the curvature Export-Table CSV: columns p_rel, MeanCurv, GaussCurv.

    Written by the notebook's Export Table block (Table Path input). Tolerant of
    header presence, column order (mapped by name when a header row exists) and
    ',' or ';' delimiters. Falls back to positional p_rel, MeanCurv, GaussCurv.
    Returns {'rho':.., 'mean':.., 'gauss':..} as floats or None.
    """
    out = {"rho": None, "mean": None, "gauss": None}
    try:
        raw = Path(path).read_text(encoding="utf-8", errors="replace")
    except Exception:
        return out

    def _num(t):
        try:
            return float(t)
        except Exception:
            return None

    rows = []
    for line in raw.splitlines():
        parts = [p.strip() for p in line.replace(";", ",").split(",") if p.strip()]
        if parts:
            rows.append(parts)
    if not rows:
        return out

    header = data = None
    for r in rows:
        if all(_num(t) is None for t in r):
            header = [t.lower() for t in r]
        elif data is None and any(_num(t) is not None for t in r):
            data = r
    if data is None:
        return out
    vals = [_num(t) for t in data]

    mapped = False
    if header and len(header) == len(vals):
        for h, v in zip(header, vals):
            if v is None:
                continue
            # same two conventions as the Table parser: GaussCurv/MeanCurv, or
            # avg_Mn_K (Gaussian) / avg_Mn_H (mean). gauss is tested FIRST.
            if "gauss" in h or h.endswith("_k") or "_k_" in h:
                out["gauss"] = v; mapped = True
            elif "mean" in h or "curv" in h or h.endswith("_h") or "_h_" in h:
                out["mean"] = v; mapped = True
            elif "rel" in h or "rho" in h:
                out["rho"] = v; mapped = True
    if not mapped:
        # Positional fallback assumes p_rel, MeanCurv, GaussCurv. If the notebook
        # writes a different ORDER (e.g. p_rel, avg_Mn_K, avg_Mn_H) this silently
        # SWAPS Gaussian and mean, and nothing downstream would ever notice.
        # Say so loudly rather than guess quietly.
        print(f"[!] curvature CSV headers not recognised ({header}); falling back to "
              "positional p_rel, MeanCurv, GaussCurv - VERIFY the column order.")
        if len(vals) >= 1:
            out["rho"] = vals[0]
        if len(vals) >= 2:
            out["mean"] = vals[1]
        if len(vals) >= 3:
            out["gauss"] = vals[2]
    return out


# ============================================================
# homogenisation post-processing (Voigt-Reuss-Hill)
# ============================================================
def read_c_tensor_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        for line in f:
            vals = [v.strip() for v in line.replace(";", ",").split(",") if v.strip()]
            if not vals:
                continue
            try:
                rows.append([float(v) for v in vals])
            except ValueError:
                continue
    if len(rows) < 6:
        raise ValueError(f"{path}: expected 6 rows of C tensor, got {len(rows)}")
    C = np.array(rows[:6])
    if C.shape != (6, 6):
        raise ValueError(f"{path}: C tensor shape {C.shape}, expected (6,6)")
    return 0.5 * (C + C.T)


def voigt_reuss_hill(C):
    K_V = (C[0, 0] + C[1, 1] + C[2, 2] + 2 * (C[0, 1] + C[0, 2] + C[1, 2])) / 9.0
    G_V = ((C[0, 0] + C[1, 1] + C[2, 2]) - (C[0, 1] + C[0, 2] + C[1, 2])
           + 3 * (C[3, 3] + C[4, 4] + C[5, 5])) / 15.0
    S = np.linalg.inv(C)
    K_R = 1.0 / ((S[0, 0] + S[1, 1] + S[2, 2]) + 2 * (S[0, 1] + S[0, 2] + S[1, 2]))
    G_R = 15.0 / (4 * (S[0, 0] + S[1, 1] + S[2, 2])
                  - 4 * (S[0, 1] + S[0, 2] + S[1, 2])
                  + 3 * (S[3, 3] + S[4, 4] + S[5, 5]))
    K_H = 0.5 * (K_V + K_R)
    G_H = 0.5 * (G_V + G_R)
    E_iso = 9 * K_H * G_H / (3 * K_H + G_H)
    nu_iso = (3 * K_H - 2 * G_H) / (2 * (3 * K_H + G_H))
    return dict(K_V=K_V, K_R=K_R, K_H=K_H, G_V=G_V, G_R=G_R, G_H=G_H,
                E_iso=E_iso, nu_iso=nu_iso)


def zener_ratio(C):
    Z1 = 2 * C[3, 3] / (C[0, 0] - C[0, 1])
    Z2 = 2 * C[4, 4] / (C[1, 1] - C[1, 2])
    Z3 = 2 * C[5, 5] / (C[2, 2] - C[0, 2])
    return float(np.mean([Z1, Z2, Z3]))


def tensorial_anisotropy_index(C):
    vrh = voigt_reuss_hill(C)
    lam = vrh["K_H"] - 2 * vrh["G_H"] / 3.0
    mu = vrh["G_H"]
    C_iso = np.zeros((6, 6))
    for i in range(3):
        for j in range(3):
            C_iso[i, j] = lam + (2 * mu if i == j else 0.0)
    for i in range(3, 6):
        C_iso[i, i] = mu
    return float(np.linalg.norm(C - C_iso, ord="fro") / np.linalg.norm(C, ord="fro"))


def _sphere_points(n=8000):
    phi = np.pi * (3.0 - np.sqrt(5.0))
    idx = np.arange(n)
    z = 1 - (2 * idx + 1) / n
    r = np.sqrt(1 - z * z)
    theta = phi * idx
    return np.stack([r * np.cos(theta), r * np.sin(theta), z], axis=1)


def _voigt_to_full_compliance(S6):
    voigt = {(0, 0): 0, (1, 1): 1, (2, 2): 2, (1, 2): 3, (2, 1): 3,
             (0, 2): 4, (2, 0): 4, (0, 1): 5, (1, 0): 5}
    S = np.zeros((3, 3, 3, 3))
    for i in range(3):
        for j in range(3):
            for k in range(3):
                for l in range(3):
                    a, b = voigt[(i, j)], voigt[(k, l)]
                    f = 1.0
                    if a >= 3:
                        f *= 0.5
                    if b >= 3:
                        f *= 0.5
                    S[i, j, k, l] = S6[a, b] * f
    return S


def directional_modulus(C, n=8000):
    S = _voigt_to_full_compliance(np.linalg.inv(C))
    pts = _sphere_points(n)
    Snn = np.einsum("pi,pj,pk,pl,ijkl->p", pts, pts, pts, pts, S)
    E_n = 1.0 / Snn
    return float(E_n.min()), float(E_n.max())


def analyse_c_tensor(csv_path):
    C = read_c_tensor_csv(csv_path) / 1000.0     # MPa -> GPa
    vrh = voigt_reuss_hill(C)
    E_min, E_max = directional_modulus(C)
    return dict(E_iso=vrh["E_iso"], nu_iso=vrh["nu_iso"], G_iso=vrh["G_H"],
                TAI=tensorial_anisotropy_index(C), Z=zener_ratio(C),
                AU=5 * (vrh["G_V"] / vrh["G_R"]) + (vrh["K_V"] / vrh["K_R"]) - 6.0,
                E_min=E_min, E_max=E_max, dE=E_max - E_min,
                C11=C[0, 0], C12=C[0, 1], C44=C[3, 3])


def stress_stats(stress_csv_path):
    """Von Mises point map (Pa) -> raw max + p99, in MPa."""
    p = Path(stress_csv_path)
    if not p.exists() or p.stat().st_size == 0:
        return None
    vals = []
    with open(p, encoding="utf-8", errors="replace") as f:
        for line in f:
            parts = line.replace(";", ",").split(",")
            for s in reversed(parts):
                s = s.strip()
                if not s:
                    continue
                try:
                    vals.append(float(s))
                except ValueError:
                    pass
                break
    if not vals:
        return None
    a = np.array(vals)
    return dict(max_raw=float(a.max()) / 1e6,
                p99=float(np.percentile(a, 99)) / 1e6,
                n_nodes=len(a))


# ============================================================
# input workbook
# ============================================================
def load_input_xlsx():
    import openpyxl
    if not INPUT_XLSX.exists():
        die(f"runs_input.xlsx not found next to the script:\n  {INPUT_XLSX}")
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)

    # settings sheet (optional)
    settings = dict(DEFAULT_SETTINGS)
    if "settings" in [s.lower() for s in wb.sheetnames]:
        ws = wb[[s for s in wb.sheetnames if s.lower() == "settings"][0]]
        for r in range(1, ws.max_row + 1):
            k = ws.cell(r, 1).value
            v = ws.cell(r, 2).value
            if k is None or v is None:
                continue
            k = str(k).strip().lower()
            if "yield" in k or "sigma_ys" in k:
                settings["sigma_ys_mpa"] = float(v)
            elif k in ("nu_s", "nu", "poisson", "solid_nu", "poisson_ratio"):
                settings["nu_s"] = float(v)
            elif k.startswith("es") or "solid" in k:
                settings["es_gpa"] = float(v)
            elif "strain" in k:
                settings["strain"] = float(v)
            elif "type" in k:                      # adms_type: DF | raw | flow
                settings["adms_type"] = str(v).strip()
            elif "ntop" in k or "notebook" in k:   # ntop_file: which .ntop to run
                settings["ntop_file"] = str(v).strip()
            elif "n_stiff" in k or "exponent" in k:  # reference-line exponent
                settings["n_stiff"] = float(v)
            elif "ntopcl" in k or "exe" in k:       # full path to ntopcl.exe
                settings["ntopcl_path"] = str(v).strip()

    # runs sheet
    name = next((s for s in wb.sheetnames if s.lower() == "runs"), wb.sheetnames[0])
    ws = wb[name]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    hidx = {h.lower(): i for i, h in enumerate(headers) if h}
    col_alias = {
        "density": "Density", "thickness": "Thickness",
        "inner size": "Inner Size", "size multi": "Size Multi", "seed": "Seed",
    }
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = {}
        for key, nb in col_alias.items():
            if key not in hidx:
                continue
            v = ws.cell(r, hidx[key] + 1).value
            if v is None or v == "":
                continue
            try:
                vals[nb] = int(v) if nb == "Seed" else float(v)
            except (TypeError, ValueError):
                pass
        if any(k in vals for k in ("Density", "Thickness", "Inner Size")):
            rows.append(vals)
    return rows, settings


def run_stem(iv):
    parts = [FILE_PREFIX]
    for name, prefix, digits in (("Inner Size", "i", 2), ("Density", "d", 2),
                                 ("Thickness", "t", 2), ("Seed", "s", 0),
                                 ("Size Multi", "m", 2)):
        v = iv.get(name)
        if v is None:
            continue
        parts.append(f"{prefix}{int(v)}" if digits == 0 else tag_num(v, prefix, digits))
    return "_".join(parts)


# ============================================================
# one run
# ============================================================
def run_one(ntopcl, iv, defaults):
    """Execute one nTop run into <batch>/Data/<stem>/. Returns run record.

    Each .bat invocation gets its own timestamped batch folder (created in
    main() via _init_batch_dir), so runs are never skipped across batches.
    """
    merged = dict(defaults)
    merged.update(iv)
    stem = run_stem(merged)
    rdir = DATA_DIR / stem
    out_csv = rdir / f"{stem}.csv"
    stress_csv = rdir / f"{stem}_stress.csv"
    shear_stress_csv = rdir / f"{stem}_shear_stress.csv"
    stl_file = rdir / f"{stem}.stl"
    curv_csv = rdir / f"{stem}_curv.csv"
    out_json = rdir / "outputs.json"
    log_txt = rdir / "log.txt"
    info_json = rdir / "run_info.json"

    rdir.mkdir(parents=True, exist_ok=True)
    in_json = rdir / "inputs.json"
    fed = build_input_json(merged, in_json, out_csv, stress_csv, shear_stress_csv,
                           stl_path=stl_file, table_path=curv_csv)

    cached_ntop = rdir / f"{stem}.ntop"
    shutil.copy(NTOP_FILE, cached_ntop)

    print(f"\n{'=' * 62}\n  RUN {stem}")
    for k, v in fed.items():
        if k not in (NAME_OUTPATH, NAME_STRESSPATH, NAME_SHEAR_STRESSPATH,
                     NAME_STLPATH, NAME_TABLEPATH):
            print(f"    {k} = {v}")
    print(f"    Start {time.strftime('%H:%M:%S')}\n{'=' * 62}")

    args = [ntopcl, "-v2", "-s", "-j", str(in_json), "-o", str(out_json),
            str(cached_ntop)]
    rc, secs = stream_process(args, log_path=log_txt)

    ok = out_csv.exists() and out_csv.stat().st_size > 100
    # STL export bookkeeping (Export Mesh block; only if notebook has "STL Path")
    stl_ok = stl_file.exists() and stl_file.stat().st_size > 1000
    stl_fed = any("stl" in k.lower() for k in fed)
    if not stl_fed:
        print("  [!] notebook schema has no 'STL Path' input - STL not exported "
              "(run with --refresh-schema to rebuild the schema from the notebook)")
    elif not stl_ok:
        print("  [!] STL Path was set but no STL was written - check Export Mesh block")
    else:
        stl_lib = SCRIPT_DIR / "ADMS_STL"
        if stl_lib.is_dir():
            try:
                shutil.copy(stl_file, stl_lib / stl_file.name)
                print(f"  STL exported + copied to ADMS_STL ({stl_file.stat().st_size / 1e6:.1f} MB)")
            except OSError as ex:
                print(f"  [!] STL copy to ADMS_STL failed: {ex}")
    outputs = read_output_values(out_json) if ok else {}
    info = {"inputs": merged, "ok": ok, "exit_code": rc,
            "time_s": round(secs, 1), "outputs": outputs,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S")}
    info_json.write_text(json.dumps(info, indent=2), encoding="utf-8")
    rho = outputs.get("rho_rel")
    print(f"  exit={rc}  time={secs / 60:.1f} min  CSV={'OK' if ok else 'MISSING'}"
          f"  rho_rel={f'{rho:.4f}' if rho is not None else 'n/a'}")
    return {"stem": stem, "ok": ok, "secs": secs, "skipped": False}


# ============================================================
# summary workbook (rebuilt from Results/Data every time)
# ============================================================
SUMMARY_COLS = [
    "Run", "Version", "Batch",
    "Density", "Thickness", "Inner_Size", "Size_Multi", "Seed",
    # The curvature columns below are NORMALISED by this length
    # (avg_Mn_K = mean Gaussian x CellSize^2, avg_Mn_H = mean H x CellSize),
    # so two rows are only comparable if it matches. It became a notebook
    # input on 2026-08-05; every row before that was run at 3 mm and is
    # left blank here rather than back-dated with a value nobody measured.
    "Cell_Size_mm",
    # NAMES MATTER. The master Results_summary.xlsx already carries these two
    # columns as MeanCurv_STL / GaussCurv_STL (145 of 159 rows), and the ML
    # label_map sheet reads those exact names. update_master() rebuilds the
    # workbook from this list by NAME, so a different spelling here would have
    # blanked every existing value. Both names mean the same thing: nTop's
    # curvature evaluated on the exported STL mesh.
    "Rho_rel", "MeanCurv_STL", "GaussCurv_STL",
    # ------ elastic (from homogenisation) ------
    "E_iso_GPa", "E_over_Es", "nu_iso", "G_iso_GPa", "G_over_Gs",
    "TAI", "Zener_Z", "A_U", "E_min_GPa", "E_max_GPa", "dE_GPa",
    "C11_GPa", "C12_GPa", "C44_GPa",
    # ------ compression yield (static compression run) ------
    "Sigma_applied_MPa", "Sigma_p99_MPa", "Sigma_max_raw_MPa",
    "Yield_onset_MPa", "SCF_p99",
    # ------ shear yield (static shear run - NEW) ------
    "Tau_applied_MPa", "Shear_p99_MPa", "Shear_max_raw_MPa",
    "Shear_onset_MPa", "SCF_shear_p99",
    # ------ material + status ------
    "Sigma_ys_MPa",
    "Status", "Time_min",
]


def build_summary(settings, stem_versions=None):
    """Build this batch's Results_summary.xlsx. Returns list of row dicts.

    stem_versions maps each stem to the Version label ('v1'/'v2'/...) it should
    have based on how many entries with that stem already exist in the master.
    Defaults to 'v1' for any stem not in the dict.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    sigma_ys = settings["sigma_ys_mpa"]
    es = settings["es_gpa"]
    nu_s = float(settings.get("nu_s", DEFAULT_SETTINGS["nu_s"]))
    strain = settings["strain"]
    stem_versions = stem_versions or {}
    batch_name = BATCH_DIR.name if BATCH_DIR is not None else ""

    rows = []
    for rdir in sorted(DATA_DIR.iterdir() if DATA_DIR.exists() else []):
        if not rdir.is_dir():
            continue
        stem = rdir.name
        info = {}
        info_p = rdir / "run_info.json"
        if info_p.exists():
            try:
                info = json.loads(info_p.read_text(encoding="utf-8"))
            except Exception:
                info = {}
        iv = info.get("inputs", {})
        out_csv = rdir / f"{stem}.csv"
        stress_csv = rdir / f"{stem}_stress.csv"
        shear_stress_csv = rdir / f"{stem}_shear_stress.csv"

        row = {c: None for c in SUMMARY_COLS}
        row["Run"] = stem
        row["Version"] = stem_versions.get(stem, "v1")
        row["Batch"] = batch_name
        row["Density"] = iv.get("Density")
        row["Thickness"] = iv.get("Thickness")
        row["Inner_Size"] = iv.get("Inner Size")
        row["Size_Multi"] = iv.get("Size Multi")
        row["Seed"] = iv.get("Seed")
        row["Cell_Size_mm"] = iv.get("Cell Size")
        row["Sigma_ys_MPa"] = sigma_ys
        row["Time_min"] = (round(info["time_s"] / 60, 1)
                           if info.get("time_s") else None)
        # Curvature + p_rel. Prefer the Output table (parsed by read_output_values
        # into outputs.meancurv/gausscurv/rho_rel); fall back to the Export Table
        # CSV (<stem>_curv.csv) if the notebook still writes it. Either source
        # works, so deleting the Export Table block leaves curvature intact.
        curv = read_curv_csv(rdir / f"{stem}_curv.csv")
        outs = info.get("outputs", {})
        row["Rho_rel"] = outs.get("rho_rel") if outs.get("rho_rel") is not None else curv.get("rho")
        row["MeanCurv_STL"] = outs.get("meancurv") if outs.get("meancurv") is not None else curv.get("mean")
        row["GaussCurv_STL"] = outs.get("gausscurv") if outs.get("gausscurv") is not None else curv.get("gauss")

        ok = out_csv.exists() and out_csv.stat().st_size > 100
        if not ok:
            row["Status"] = "FAIL (no C tensor CSV - see log.txt)"
            rows.append(row)
            continue

        try:
            h = analyse_c_tensor(out_csv)
        except Exception as e:
            row["Status"] = f"FAIL (C tensor parse: {e})"
            rows.append(row)
            continue

        # solid shear modulus for G/Gs (isotropic solid): Gs = Es / 2(1+nu_s).
        # nu_s comes from the settings sheet - it must match the solid Poisson
        # ratio set in the .ntop Isotropic Linear Elastic block, otherwise
        # G_over_Gs is not comparable across materials.
        gs_gpa = es / (2.0 * (1.0 + nu_s))
        row.update({
            "E_iso_GPa": round(h["E_iso"], 4),
            "E_over_Es": round(h["E_iso"] / es, 5),
            "nu_iso": round(h["nu_iso"], 4),
            "G_iso_GPa": round(h["G_iso"], 4),
            "G_over_Gs": round(h["G_iso"] / gs_gpa, 5),
            "TAI": round(h["TAI"], 4),
            "Zener_Z": round(h["Z"], 4),
            "A_U": round(h["AU"], 4),
            "E_min_GPa": round(h["E_min"], 3),
            "E_max_GPa": round(h["E_max"], 3),
            "dE_GPa": round(h["dE"], 3),
            "C11_GPa": round(h["C11"], 3),
            "C12_GPa": round(h["C12"], 3),
            "C44_GPa": round(h["C44"], 3),
        })

        # ---- compression onset (static compression CSV) ----
        sigma_applied = h["E_iso"] * 1000.0 * strain     # GPa -> MPa
        row["Sigma_applied_MPa"] = round(sigma_applied, 2)

        ss = stress_stats(stress_csv)
        comp_ok = False
        if ss:
            comp_ok = True
            row["Sigma_p99_MPa"] = round(ss["p99"], 1)
            row["Sigma_max_raw_MPa"] = round(ss["max_raw"], 1)
            if ss["p99"] > 0:
                row["Yield_onset_MPa"] = round(
                    sigma_applied * sigma_ys / ss["p99"], 3)
                row["SCF_p99"] = round(ss["p99"] / sigma_applied, 2)

        # ---- shear onset (static shear CSV, NEW) ----
        tau_applied = h["G_iso"] * 1000.0 * strain       # GPa -> MPa
        row["Tau_applied_MPa"] = round(tau_applied, 2)

        sh = stress_stats(shear_stress_csv)
        shear_ok = False
        if sh:
            shear_ok = True
            row["Shear_p99_MPa"] = round(sh["p99"], 1)
            row["Shear_max_raw_MPa"] = round(sh["max_raw"], 1)
            if sh["p99"] > 0:
                row["Shear_onset_MPa"] = round(
                    tau_applied * sigma_ys / sh["p99"], 3)
                row["SCF_shear_p99"] = round(sh["p99"] / tau_applied, 2)

        # ---- status ----
        if comp_ok and shear_ok:
            row["Status"] = "OK"
        elif comp_ok:
            row["Status"] = "OK (no shear CSV)"
        elif shear_ok:
            row["Status"] = "OK (no compression CSV)"
        else:
            row["Status"] = "OK (stiffness only - no stress CSVs)"
        rows.append(row)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(SUMMARY_COLS)
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="500778")   # UCL purple
    for c in ws[1]:
        c.font = hdr_font
        c.fill = hdr_fill
    for row in rows:
        ws.append([row.get(c) for c in SUMMARY_COLS])
    for i, col in enumerate(SUMMARY_COLS, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = \
            max(11, len(col) + 2)
    ws.freeze_panes = "B2"

    # method notes sheet
    ws2 = wb.create_sheet("method")
    gs_gpa_note = es / (2.0 * (1.0 + nu_s))
    notes = [
        ["Column", "Meaning"],
        ["Rho_rel", "Relative density measured by nTop (solid fraction of the cube)"],
        ["-- ELASTIC (homogenisation) --", ""],
        ["E_iso_GPa / nu_iso / G_iso_GPa", "Voigt-Reuss-Hill isotropic moduli from the homogenised 6x6 C tensor"],
        ["E_over_Es", f"E_iso normalised by solid modulus Es = {es} GPa"],
        ["G_over_Gs", f"G_iso normalised by solid shear modulus Gs = Es / 2(1+nu_s) = {gs_gpa_note:.2f} GPa (nu_s = {nu_s})"],
        ["TAI", "Tensorial anisotropy index ||C - C_iso||_F / ||C||_F (0 = isotropic)"],
        ["Zener_Z / A_U", "Zener ratio (1 = cubic-isotropic) / universal anisotropy index"],
        ["E_min/E_max/dE_GPa", "Directional Young's modulus range (8000-direction sphere sampling)"],
        ["C11/C12/C44_GPa", "Key stiffness tensor terms"],
        ["-- COMPRESSION YIELD (static compression run) --", ""],
        ["Sigma_applied_MPa", f"Macroscopic compressive stress = E_iso x strain (strain = {strain*100:.3f}%, TopDisp = InnerSize x -0.00667)"],
        ["Sigma_p99_MPa", "99th-percentile nodal von Mises from the static compression (robust peak)"],
        ["Sigma_max_raw_MPa", "Raw max nodal von Mises (mesh-sensitive at cut edges - use with caution)"],
        ["Yield_onset_MPa", f"Sigma_applied x (sigma_ys / sigma_p99); sigma_ys = {sigma_ys} MPa (editable in runs_input.xlsx 'settings' sheet)"],
        ["SCF_p99", "Stress concentration factor = sigma_p99 / sigma_applied"],
        ["-- SHEAR YIELD (static shear run - NEW) --", ""],
        ["Tau_applied_MPa", f"Macroscopic shear stress = G_iso x strain (strain = {strain*100:.3f}%, TopShearDisp = InnerSize x +0.00667 along X)"],
        ["Shear_p99_MPa", "99th-percentile nodal von Mises from the static shear (robust peak)"],
        ["Shear_max_raw_MPa", "Raw max nodal von Mises under shear (mesh-sensitive at cut edges)"],
        ["Shear_onset_MPa", f"Tau_applied x (sigma_ys / shear_p99); macroscopic shear stress at first local yield"],
        ["SCF_shear_p99", "Stress concentration factor under shear = shear_p99 / tau_applied"],
        ["", ""],
        ["Method", "Linear elastic on the SAME FE volume mesh. Bottom face fully clamped in both cases. "
                   "Compression: top Uz = -InnerSize*0.00667. Shear: top Ux = +InnerSize*0.00667, Uy=Uz=0. "
                   "Onsets assume stress scales linearly with applied load (valid pre-yield)."],
        ["Per-run data", "Results/Data/<run>/ has the C tensor CSV, compression stress point-map CSV, "
                         "shear stress point-map CSV, log, and a cached .ntop you can open in the nTop GUI."],
    ]
    for n in notes:
        ws2.append(n)
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 110
    for c in ws2[1]:
        c.font = hdr_font
        c.fill = hdr_fill

    wb.save(SUMMARY_XLSX)
    return rows


# ============================================================
# master summary (cumulative, append-only, across all batches)
# ============================================================
def load_master_rows():
    """Read Results/Results_summary.xlsx (if it exists) into a list of dicts.
    Returns [] if the file is missing or unreadable."""
    if not MASTER_SUMMARY_XLSX.exists():
        return []
    import openpyxl
    try:
        wb = openpyxl.load_workbook(MASTER_SUMMARY_XLSX, data_only=True)
    except Exception as e:
        print(f"[warn] could not read master summary ({e}); treating as empty")
        return []
    sheet = "summary" if "summary" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            vals[h] = ws.cell(r, i + 1).value
        if vals.get("Run"):
            rows.append(vals)
    return rows


def master_stem_counts(master_rows):
    """Map stem -> number of prior (Status starts with 'OK') rows in master."""
    counts = {}
    for r in master_rows:
        status = str(r.get("Status") or "")
        if not status.startswith("OK"):
            continue          # FAIL rows don't block a rerun
        stem = r.get("Run")
        if not stem:
            continue
        counts[stem] = counts.get(stem, 0) + 1
    return counts


def next_version(count_before):
    """v1 for a new stem, v2 for one prior, v3 for two prior, ..."""
    return f"v{count_before + 1}"


def prompt_reruns(input_rows, defaults, master_counts, auto_yes=False):
    """UP-FRONT prompt loop. For each input row that already has an OK entry
    in the master, ask [Y]es / [N]o / [A]ll. Blocking; runs to completion
    before any simulation starts, so the user can walk away afterwards.

    Returns:
      to_run    -- list of (input_dict, stem, version_label) to execute
      new_master_counts -- master_counts updated so each rerun's version label
                           reflects what row number in the master it will be.
    """
    counts = dict(master_counts)
    to_run = []
    all_yes = bool(auto_yes)
    if all_yes:
        print("[dedupe] --rerun-all: every duplicate will be rerun as the next "
              "version, no prompts.")

    # first, compute stems + duplicate status
    plan = []
    for iv in input_rows:
        merged = dict(defaults)
        merged.update(iv)
        stem = run_stem(merged)
        n_prior = counts.get(stem, 0)
        plan.append((iv, stem, n_prior))

    dup_count = sum(1 for _, _, n in plan if n > 0)
    if dup_count == 0:
        print(f"[dedupe] no duplicates vs master ({len(plan)} fresh runs)")
        for iv, stem, n in plan:
            counts[stem] = counts.get(stem, 0) + 1
            to_run.append((iv, stem, next_version(n)))
        return to_run, counts

    print(f"\n[dedupe] {dup_count} of {len(plan)} rows already exist in "
          f"master ({MASTER_SUMMARY_XLSX}).")
    print("        Answer for each below; simulation starts AFTER all "
          "answers.\n")

    for idx, (iv, stem, n_prior) in enumerate(plan, 1):
        if n_prior == 0:
            print(f"  [{idx}/{len(plan)}] {stem}  NEW  -> will run as v1")
            counts[stem] = 1
            to_run.append((iv, stem, "v1"))
            continue

        next_v = next_version(n_prior)
        if all_yes:
            print(f"  [{idx}/{len(plan)}] {stem}  DUP (has {n_prior} entry) "
                  f"-> rerun as {next_v} [A auto]")
            counts[stem] = n_prior + 1
            to_run.append((iv, stem, next_v))
            continue

        while True:
            try:
                ans = input(
                    f"  [{idx}/{len(plan)}] {stem}  already has {n_prior} "
                    f"entry in master.\n"
                    f"       Rerun as {next_v}? [Y] / [N] skip / "
                    f"[A] yes-to-all-remaining : "
                ).strip().lower()
            except (EOFError, KeyboardInterrupt):
                print("\n[abort] user cancelled at prompt")
                sys.exit(130)
            if ans in ("y", "yes", ""):
                counts[stem] = n_prior + 1
                to_run.append((iv, stem, next_v))
                break
            if ans in ("n", "no"):
                print(f"       -> skip")
                break
            if ans in ("a", "all"):
                print(f"       -> yes to this and all remaining duplicates")
                all_yes = True
                counts[stem] = n_prior + 1
                to_run.append((iv, stem, next_v))
                break
            print("       (please answer Y / N / A)")

    print(f"\n[dedupe] final plan: will run {len(to_run)} of {len(plan)} rows.")
    return to_run, counts


# ============================================================
# analysis-sheet rebuild — series_data, fit_curves, parameter_effects, charts
# ============================================================
# Charts group data by DESIGN PARAMETER (Thickness for scaling laws, Inner_Size
# for anisotropy/RVE), not by admin batch. Colours ramp with parameter value
# so points from any batch with the same design collapse onto one series.
_THICKNESS_COLOURS = {
    0.2: "004AAD",  # deep blue
    0.3: "0072CE",  # blue
    0.4: "007E67",  # teal
    0.5: "7AB800",  # yellow-green
    0.6: "F5B300",  # gold
    0.7: "D66A00",  # orange
    0.8: "AC1414",  # red
}
_INNER_COLOURS = {
    9:  "AC145A",   # magenta
    12: "D66A00",   # orange
    15: "500778",   # UCL purple (5x5x5 = standard RVE)
    18: "0072CE",   # blue
}
_FALLBACK_COLOURS = ["555F6D", "93272C", "9C2AA0", "222222", "8DB9CA"]


def _colour_for(param_val, mapping, fallback_idx=0):
    if param_val in mapping:
        return mapping[param_val]
    # closest-neighbour fallback for out-of-set numeric values
    if isinstance(param_val, (int, float)) and mapping:
        closest = min(mapping.keys(), key=lambda k: abs(k - param_val))
        return mapping[closest]
    return _FALLBACK_COLOURS[fallback_idx % len(_FALLBACK_COLOURS)]


def _dedup_highest(rows_iter):
    from collections import defaultdict
    g = defaultdict(list)
    for r in rows_iter:
        k = (r.get("Density"), r.get("Thickness"), r.get("Inner_Size"),
             r.get("Size_Multi"), r.get("Seed"))
        g[k].append(r)
    return [sorted(gv, key=lambda r: int(str(r.get("Version") or "v0")[1:] or 0))[-1]
            for gv in g.values()]


def _powerfit(xs, ys):
    import math
    xs, ys = np.array(xs), np.array(ys)
    lx, ly = np.log(xs), np.log(ys)
    n, c = np.polyfit(lx, ly, 1)
    yhat = np.exp(n * lx + c)
    ss_res = float(np.sum((ys - yhat) ** 2))
    ss_tot = float(np.sum((ys - ys.mean()) ** 2))
    r2 = 1 - ss_res / ss_tot if ss_tot else 0.0
    return math.exp(c), n, r2, len(xs)


def rebuild_analysis_sheets(wb, all_rows):
    """Wipe + rebuild series_data, fit_curves, parameter_effects, charts.
    Called by update_master(). Charts group data by DESIGN PARAMETER
    (Thickness for scaling laws, Inner_Size for anisotropy/RVE view).
    One series per parameter value, one colour per value."""
    import openpyxl
    from collections import defaultdict
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import ScatterChart, BarChart, Reference, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.axis import ChartLines
    from openpyxl.drawing.line import LineProperties

    for name in ("chart_data", "charts", "series_data", "fit_curves",
                 "parameter_effects"):
        if name in wb.sheetnames:
            del wb[name]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="500778")

    # ---- pivot data by DESIGN PARAMETER ----
    def pivot_by(param_key, y_key):
        """param_value -> [(rho, y), ...] sorted by rho."""
        out = defaultdict(list)
        for r in all_rows:
            rho = r.get("Rho_rel"); y = r.get(y_key)
            p = r.get(param_key)
            if rho is None or y is None or p is None or rho <= 0:
                continue
            out[p].append((rho, y))
        for v in out:
            out[v].sort()
        return dict(sorted(out.items()))

    # Scaling laws grouped by Thickness (design knob that also drives rho)
    E_p_t = pivot_by("Thickness", "E_over_Es")
    G_p_t = pivot_by("Thickness", "G_over_Gs")
    O_p_t = pivot_by("Thickness", "Yield_onset_MPa")
    # Anisotropy grouped by Inner_Size (drives RVE convergence)
    T_p_i = pivot_by("Inner_Size", "TAI")
    # Auxiliary: E/Es grouped by Inner_Size (shows RVE effect on stiffness)
    E_p_i = pivot_by("Inner_Size", "E_over_Es")

    # ---- series_data sheet ----
    ws_sd = wb.create_sheet("series_data")
    ws_sd["A1"] = ("Pivoted data feeding the charts. Series grouped by DESIGN PARAMETER "
                   "(Thickness or Inner_Size), not batch. Same-parameter points from different "
                   "batches share a series so scaling-law collapse is visible.")
    ws_sd["A1"].font = Font(italic=True, color="500778")

    def write_block(sheet, header_row, start_col, label, pivot_dict, param_short, unit):
        """Column pairs (rho, y) per param value. Header cell = 'param = value unit' for series title.
        Returns dict: param_val -> (x_col_num, y_col_num, last_data_row_num)."""
        sheet.cell(header_row - 1, start_col, f"BLOCK: {label}").font = \
            Font(bold=True, color="500778")
        refs = {}
        col = start_col
        for pval in pivot_dict:
            data = pivot_dict[pval]
            if not data:
                continue
            sheet.cell(header_row, col, f"{param_short}={pval} ρ").font = Font(bold=True, size=9)
            title = f"{param_short} = {pval}{' ' + unit if unit else ''}"
            sheet.cell(header_row, col + 1, title).font = Font(bold=True, size=9)
            for i, (x, y) in enumerate(data, 1):
                sheet.cell(header_row + i, col, x)
                sheet.cell(header_row + i, col + 1, y)
            refs[pval] = (col, col + 1, header_row + len(data))
            col += 2
        return refs, col

    E_refs_t, next1 = write_block(ws_sd, 3, 1,          "E/Es by Thickness",  E_p_t, "t",     "mm")
    G_refs_t, _     = write_block(ws_sd, 3, next1 + 1,  "G/Gs by Thickness",  G_p_t, "t",     "mm")
    O_refs_t, next2 = write_block(ws_sd, 30, 1,         "Onset by Thickness", O_p_t, "t",     "mm")
    T_refs_i, _     = write_block(ws_sd, 30, next2 + 1, "TAI by Inner_Size",  T_p_i, "Inner", "mm")
    E_refs_i, _     = write_block(ws_sd, 55, 1,         "E/Es by Inner_Size", E_p_i, "Inner", "mm")

    # Colour resolvers
    def _map_for(param_ref):
        return _THICKNESS_COLOURS if param_ref == "t" else _INNER_COLOURS

    # ---- compute fits ----
    dedup = _dedup_highest(all_rows)
    def _safe_xy(key):
        xs = []; ys = []
        for r in dedup:
            rho = r.get("Rho_rel"); y = r.get(key)
            if rho is None or y is None or rho <= 0:
                continue
            xs.append(rho); ys.append(y)
        return xs, ys

    E_xs, E_ys = _safe_xy("E_over_Es")
    G_xs, G_ys = _safe_xy("G_over_Gs")
    O_xs, O_ys = _safe_xy("Yield_onset_MPa")

    C_E = n_E = R2_E = N_E_ = 0
    C_G = n_G = R2_G = N_G_ = 0
    C_O = n_O = R2_O = N_O_ = 0
    if len(E_xs) >= 3: C_E, n_E, R2_E, N_E_ = _powerfit(E_xs, E_ys)
    if len(G_xs) >= 3: C_G, n_G, R2_G, N_G_ = _powerfit(G_xs, G_ys)
    if len(O_xs) >= 3: C_O, n_O, R2_O, N_O_ = _powerfit(O_xs, O_ys)

    # ---- fit_curves sheet ----
    ws_fc = wb.create_sheet("fit_curves")
    ws_fc.append(["ρ_rel", "E/Es fit", "G/Gs fit", "Onset fit (MPa)"])
    for c in ws_fc[1]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    rho_min = min([x for x in (E_xs + G_xs + O_xs) if x > 0], default=0.05)
    rho_max = max(E_xs + G_xs + O_xs, default=0.40)
    npts = 58
    xs_fit = np.linspace(max(rho_min * 0.9, 0.03), rho_max * 1.05, npts)
    for x in xs_fit:
        ws_fc.append([float(x),
                      C_E * (x ** n_E) if C_E else None,
                      C_G * (x ** n_G) if C_G else None,
                      C_O * (x ** n_O) if C_O else None])

    # ---- CHARTS ----
    ws_ch = wb.create_sheet("charts")
    ws_ch["A1"] = ("Master scaling-law plots. Series = one per study, colours stable across charts. "
                   "Data on 'series_data', fit line on 'fit_curves'. Rebuilt each batch.")
    ws_ch["A1"].font = Font(bold=True, color="500778", size=11)

    def _make_series(x_ref, y_ref_with_hdr, colour):
        s = Series(y_ref_with_hdr, x_ref, title_from_data=True)
        s.marker = Marker(symbol="circle", size=8)
        s.marker.graphicalProperties = GraphicalProperties(solidFill=colour)
        s.marker.graphicalProperties.line = LineProperties(solidFill=colour)
        gp = GraphicalProperties(); gp.line = LineProperties(noFill=True)
        s.graphicalProperties = gp
        return s

    def _fit_series(fit_col_num, npts_, colour="222222", label="fit"):
        r0, r1 = 2, 1 + npts_
        y_hdr_ref = Reference(ws_fc, min_col=fit_col_num, min_row=1, max_col=fit_col_num, max_row=r1)
        x_ref = Reference(ws_fc, min_col=1, min_row=r0, max_col=1, max_row=r1)
        s = Series(y_hdr_ref, x_ref, title_from_data=True)
        s.marker = Marker(symbol="none")
        gp = GraphicalProperties(); gp.line = LineProperties(solidFill=colour, w=22000)
        s.graphicalProperties = gp
        return s

    def build_chart(title, refs, colour_map, log=False, add_fit_col=None):
        ch = ScatterChart()
        ch.title = title
        ch.style = 2
        ch.x_axis.title = "ρ_rel"
        ch.legend.position = "r"
        ch.height = 12; ch.width = 22
        ch.x_axis.majorGridlines = ChartLines()
        ch.y_axis.majorGridlines = ChartLines()
        if log:
            ch.x_axis.scaling.logBase = 10
            ch.y_axis.scaling.logBase = 10
        for i, (pval, (xc, yc, r_end)) in enumerate(refs.items()):
            colour = _colour_for(pval, colour_map, fallback_idx=i)
            x_ref = Reference(ws_sd, min_col=xc, min_row=4, max_col=xc, max_row=r_end)
            y_hdr = Reference(ws_sd, min_col=yc, min_row=3, max_col=yc, max_row=r_end)
            ch.series.append(_make_series(x_ref, y_hdr, colour))
        if add_fit_col and C_E:
            ch.series.append(_fit_series(add_fit_col, npts))
        return ch

    # Charts grouped by DESIGN PARAMETER
    # Primary panel: colour = Thickness (design knob driving ρ)
    ch_E_t = build_chart(f"E/Es = {C_E:.3f}·ρ^{n_E:.2f}  (R²={R2_E:.3f}, N={N_E_})   colour = Thickness (mm)",
                          E_refs_t, _THICKNESS_COLOURS, add_fit_col=2 if C_E else None)
    ch_E_t.y_axis.title = "E_iso / E_s"
    ch_G_t = build_chart(f"G/Gs = {C_G:.3f}·ρ^{n_G:.2f}  (R²={R2_G:.3f}, N={N_G_})   colour = Thickness (mm)",
                          G_refs_t, _THICKNESS_COLOURS, add_fit_col=3 if C_G else None)
    ch_G_t.y_axis.title = "G_iso / G_s"
    ch_O_t = build_chart(f"Yield onset = {C_O:.1f}·ρ^{n_O:.2f} MPa  (R²={R2_O:.3f}, N={N_O_})   colour = Thickness (mm)",
                          O_refs_t, _THICKNESS_COLOURS, add_fit_col=4 if C_O else None)
    ch_O_t.y_axis.title = "Yield onset (MPa)"

    # log-log companion for stiffness (best view of power-law)
    ch_E_t_log = build_chart(f"E/Es vs ρ (log-log, slope = {n_E:.2f})   colour = Thickness (mm)",
                              E_refs_t, _THICKNESS_COLOURS, log=True, add_fit_col=2 if C_E else None)
    ch_E_t_log.y_axis.title = "E_iso / E_s"

    # RVE / anisotropy view: colour = Inner_Size
    ch_T_i = build_chart("Anisotropy TAI vs ρ  —  colour = Inner_Size (RVE cube, mm)",
                          T_refs_i, _INNER_COLOURS)
    ch_T_i.y_axis.title = "TAI  (0 = isotropic)"
    ch_E_i = build_chart("E/Es vs ρ  —  colour = Inner_Size (RVE convergence view)",
                          E_refs_i, _INNER_COLOURS, add_fit_col=2 if C_E else None)
    ch_E_i.y_axis.title = "E_iso / E_s"

    # Layout — main by-Thickness plots on left, RVE / log-log companions on right
    ws_ch.add_chart(ch_E_t,     "A4")
    ws_ch.add_chart(ch_E_t_log, "N4")
    ws_ch.add_chart(ch_G_t,     "A29")
    ws_ch.add_chart(ch_O_t,     "N29")
    ws_ch.add_chart(ch_E_i,     "A54")
    ws_ch.add_chart(ch_T_i,     "N54")

    # ---- parameter_effects sheet ----
    _build_parameter_effects(wb, all_rows, hdr_font, hdr_fill)


def _build_parameter_effects(wb, all_rows, hdr_font, hdr_fill):
    """Rebuild parameter_effects sheet with summary table, per-parameter
    1D slices, native scatter charts, and fold-ΔE bar chart."""
    from collections import defaultdict
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import ScatterChart, BarChart, Reference, Series
    from openpyxl.chart.marker import Marker
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.chart.axis import ChartLines
    from openpyxl.drawing.line import LineProperties

    ws = wb.create_sheet("parameter_effects")

    SLICES = {
        "Density":    {"Thickness": 0.4,   "Inner_Size": 15, "Size_Multi": 1.6, "Seed": 2},
        "Thickness":  {"Density": 20.83,   "Inner_Size": 15, "Size_Multi": 1.6, "Seed": 2},
        "Inner_Size": {"Density": 20.83,   "Thickness": 0.4, "Size_Multi": 1.6, "Seed": 2},
        "Size_Multi": {"Density": 20.83,   "Thickness": 0.4, "Inner_Size": 15,  "Seed": 2},
        "Seed":       {"Density": 20.83,   "Thickness": 0.4, "Inner_Size": 15,  "Size_Multi": 1.6},
    }
    P_COLOURS = {"Density":"500778","Thickness":"D66A00","Inner_Size":"007E67",
                 "Size_Multi":"0072CE","Seed":"AC145A"}
    Es = ES_GPA_ACTIVE; N_STIFF = N_STIFF_ACTIVE

    dedup = _dedup_highest(all_rows)

    def slice_rows(param, base):
        ms = [r for r in dedup
              if all(r.get(k) == v for k, v in base.items())
              and r.get(param) is not None
              and r.get("E_iso_GPa") is not None
              and r.get("Rho_rel") not in (None, 0)]
        ms.sort(key=lambda r: r[param])
        # keep highest version per param value
        seen, out = set(), []
        for r in ms:
            if r[param] in seen: continue
            seen.add(r[param]); out.append(r)
        return out

    # -- Summary block --
    r = 1
    ws.cell(r, 1, "PARAMETER EFFECTS SUMMARY").font = Font(bold=True, size=14, color="500778")
    r += 1
    ws.cell(r, 1, "Each row = a 1D slice from the master where 4 params are held at study baseline.").font = Font(italic=True)
    r += 2
    hdr_row = r
    labels = ["Parameter", "Baseline", "N", "Range", "ρ range",
              "E range (GPa)", "Fold ΔE", "E*_res", "Verdict"]
    for j, L in enumerate(labels, 1):
        c = ws.cell(r, j, L); c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
    r += 1

    fold_data = []
    slice_starts = {}  # param -> (data_start_row, data_end_row)
    verdicts = []
    for param, base in SLICES.items():
        ms = slice_rows(param, base)
        if not ms: continue
        base_str = ", ".join(f"{k}={v}" for k, v in base.items())
        xs = [row[param] for row in ms]
        e_ys = [row["E_iso_GPa"] for row in ms]
        rhos = [row["Rho_rel"] for row in ms]
        e_star = [(row["E_iso_GPa"]/Es) / (row["Rho_rel"] ** N_STIFF) for row in ms]
        fold = max(e_ys)/min(e_ys) if min(e_ys) else 0
        res = max(e_star)/min(e_star) if min(e_star) else 0

        if param == "Seed":
            verdict = f"INERT — bit-identical E across seeds (Spherene SDK bug)"
        elif param == "Size_Multi":
            verdict = f"WEAK — E fluctuates {res:.2f}× via cut-boundary noise"
        elif param == "Inner_Size":
            verdict = f"STRUCTURAL — E* residual {res:.2f}× (RVE convergence)"
        else:
            verdict = f"DENSITY-driving — {fold:.1f}× E, E*_res {res:.2f}×"
        verdicts.append((param, base_str, len(ms), fold, res, verdict, xs, e_ys, rhos, e_star))
        fold_data.append((param, fold))

    for row_data in verdicts:
        param, base_str, N, fold, res, verdict, *_ = row_data
        vals = [param, base_str, N, f"{min(row_data[6])}-{max(row_data[6])}",
                f"{min(row_data[8]):.3f}-{max(row_data[8]):.3f}",
                f"{min(row_data[7]):.2f}-{max(row_data[7]):.2f}",
                round(fold, 2), round(res, 3), verdict]
        for j, v in enumerate(vals, 1):
            ws.cell(r, j, v)
        ws.cell(r, 1).font = Font(bold=True)
        r += 1

    widths = [12, 45, 4, 15, 15, 15, 10, 10, 60]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, j).column_letter].width = w

    # -- Per-parameter data blocks --
    r += 2
    slice_blocks = {}
    for row_data in verdicts:
        param, base_str, N, fold, res, verdict, xs, e_ys, rhos, e_star = row_data
        ws.cell(r, 1, f"Slice: {param}").font = Font(bold=True, size=12, color="500778")
        r += 1
        hdrs = [param, "Rho_rel", "E_iso_GPa", "E* = (E/Es)/ρ^1.83"]
        for j, h in enumerate(hdrs, 1):
            c = ws.cell(r, j, h); c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
        data_start = r + 1
        r += 1
        for x, rho, e, e_s in zip(xs, rhos, e_ys, e_star):
            ws.cell(r, 1, x); ws.cell(r, 2, rho)
            ws.cell(r, 3, e); ws.cell(r, 4, round(e_s, 4))
            r += 1
        slice_blocks[param] = (data_start, r - 1)
        r += 1

    # -- Fold ΔE bar chart --
    if fold_data:
        fold_data.sort(key=lambda p: -p[1])
        r_fb = r + 1
        ws.cell(r_fb, 1, "Fold ΔE ranking (bar-chart source)").font = Font(bold=True, color="500778")
        ws.cell(r_fb + 1, 1, "Parameter").font = Font(bold=True)
        ws.cell(r_fb + 1, 2, "Fold ΔE").font = Font(bold=True)
        for i, (p, f) in enumerate(fold_data, 1):
            ws.cell(r_fb + 1 + i, 1, p)
            ws.cell(r_fb + 1 + i, 2, round(f, 3))
        bar = BarChart()
        bar.title = "Fold ΔE per parameter (effect strength)"
        bar.style = 2
        bar.type = "bar"   # horizontal bars
        bar.y_axis.title = ""
        bar.x_axis.title = "Max E / Min E"
        bar.legend = None
        bar.height = 9; bar.width = 15
        cats = Reference(ws, min_col=1, min_row=r_fb + 2, max_row=r_fb + 1 + len(fold_data))
        data = Reference(ws, min_col=2, min_row=r_fb + 1, max_row=r_fb + 1 + len(fold_data))
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        ws.add_chart(bar, "L1")
        r = r_fb + len(fold_data) + 3

    # -- Per-parameter scatter charts (E vs param, E* vs param) --
    r += 2
    ws.cell(r, 1, "Per-parameter effect plots").font = Font(bold=True, size=14, color="500778")
    r += 1
    ws.cell(r, 1, "LEFT: raw E vs param.   RIGHT: E* = (E/Es)/ρ^1.83 vs param (flat = pure density knob).").font = Font(italic=True)
    r += 2

    def scatter_single(x_range, y_range_with_hdr, title, x_lbl, y_lbl, colour):
        ch = ScatterChart()
        ch.title = title
        ch.style = 2
        ch.x_axis.title = x_lbl
        ch.y_axis.title = y_lbl
        ch.legend.position = "r"
        ch.height = 9; ch.width = 15
        ch.x_axis.majorGridlines = ChartLines()
        ch.y_axis.majorGridlines = ChartLines()
        s = Series(y_range_with_hdr, x_range, title_from_data=True)
        s.marker = Marker(symbol="circle", size=10)
        s.marker.graphicalProperties = GraphicalProperties(solidFill=colour)
        s.marker.graphicalProperties.line = LineProperties(solidFill=colour)
        gp = GraphicalProperties(); gp.line = LineProperties(noFill=True)
        s.graphicalProperties = gp
        ch.series.append(s)
        return ch

    row_anchor = r
    for i, (param, (dstart, dend)) in enumerate(slice_blocks.items()):
        colour = P_COLOURS.get(param, "500778")
        # E vs param
        x_ref = Reference(ws, min_col=1, min_row=dstart, max_col=1, max_row=dend)
        y_ref = Reference(ws, min_col=3, min_row=dstart - 1, max_col=3, max_row=dend)
        ch_E = scatter_single(x_ref, y_ref, f"E vs {param}", param, "E_iso (GPa)", colour)
        # E* vs param
        y_ref2 = Reference(ws, min_col=4, min_row=dstart - 1, max_col=4, max_row=dend)
        ch_Estar = scatter_single(x_ref, y_ref2, f"E* vs {param}", param,
                                   "E* = (E/Es)/ρ^1.83", colour)
        ws.add_chart(ch_E,     f"A{row_anchor + i * 22}")
        ws.add_chart(ch_Estar, f"K{row_anchor + i * 22}")


def update_master(batch_rows):
    """Append this batch's rows to the master summary. Never overwrite; each
    entry is stored with its own Version + Batch tags. Also rebuilds the
    analysis sheets (series_data, fit_curves, parameter_effects, charts) so
    the master's plots always reflect the current dataset."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="500778")

    # start from existing master (or empty)
    existing = load_master_rows() if MASTER_SUMMARY_XLSX.exists() else []
    combined = existing + [dict(r) for r in batch_rows]

    # --- NEVER DROP A COLUMN ---------------------------------------------
    # This function rebuilds the workbook from scratch, writing row.get(name)
    # for each name in SUMMARY_COLS. Anything the master carries under a name
    # NOT in that list would silently vanish - which is exactly what nearly
    # happened when the curvature columns were spelled _nTop here and _STL in
    # the master. Carry every extra column through, at the end, and say so.
    prior_cols = []
    if existing:
        for r in existing:
            for k in r:
                if k and k not in SUMMARY_COLS and k not in prior_cols:
                    prior_cols.append(k)
    if prior_cols:
        carried = {c: sum(1 for r in existing if r.get(c) not in (None, ""))
                   for c in prior_cols}
        print(f"[master] carrying through {len(prior_cols)} column(s) the script "
              f"does not write: " +
              ", ".join(f"{c} ({n} values)" for c, n in carried.items()))
    out_cols = list(SUMMARY_COLS) + prior_cols

    # --- NEVER DROP A SHEET ----------------------------------------------
    # openpyxl.Workbook() below starts empty, so any sheet this function does
    # not rebuild (e.g. curvature_stl, the STL-vs-nTop backfill audit) would be
    # destroyed. Capture their values first and re-append them.
    OWNED_SHEETS = {"summary", "readme", "parameter_effects",
                    "chart_config", "charts"}
    preserved = {}
    if MASTER_SUMMARY_XLSX.exists():
        try:
            _old = openpyxl.load_workbook(MASTER_SUMMARY_XLSX, data_only=True)
            for nm in _old.sheetnames:
                if nm.lower() in OWNED_SHEETS:
                    continue
                preserved[nm] = [list(r) for r in
                                 _old[nm].iter_rows(values_only=True)]
            _old.close()
        except Exception as e:
            print(f"[warn] could not read existing sheets to preserve them ({e})")
        if preserved:
            print("[master] preserving sheet(s) not rebuilt here: " +
                  ", ".join(f"{k} ({len(v)} rows)" for k, v in preserved.items()))

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(out_cols)
    for c in ws[1]:
        c.font = hdr_font
        c.fill = hdr_fill
    for row in combined:
        ws.append([row.get(c) for c in out_cols])
    for i, col in enumerate(out_cols, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = \
            max(11, len(col) + 2)
    ws.freeze_panes = "D2"   # freeze through Batch column

    # readme sheet
    ws2 = wb.create_sheet("readme")
    ws2.append(["What this workbook is", ""])
    ws2.append(["Purpose",
                "Cumulative master summary across every RUN_SIMULATION.bat "
                "invocation. Append-only: reruns of the same geometry are "
                "kept as v2/v3/... rather than replacing v1."])
    ws2.append(["Row identity",
                "One row per (Run, Version). Run = 5-parameter stem (same "
                "geometry). Version = v1 for first ever run of that stem, "
                "v2 for next rerun, etc."])
    ws2.append(["Batch column",
                "The timestamped batch folder name under Results/ that "
                "produced this row - lets you find the raw CSVs + cached "
                ".ntop for that specific run."])
    ws2.append(["Column meanings",
                "Same as each batch's Results_summary.xlsx 'method' sheet."])
    ws2.append(["Analysis sheets",
                "series_data / fit_curves / parameter_effects / charts are "
                "REBUILT from scratch every time update_master runs, so plots "
                "always match the current dataset."])
    for c in ws2[1]:
        c.font = hdr_font
        c.fill = hdr_fill
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 110

    # ---- parameter_effects sheet (verdict table + 1D slices) ----
    # Must be rebuilt here: the workbook is recreated from scratch above,
    # so any sheet not rebuilt would silently vanish (happened 2026-07-13).
    try:
        _build_parameter_effects(wb, combined, hdr_font, hdr_fill)
    except Exception as e:
        print(f"[warn] parameter_effects rebuild failed: {e}  "
              f"(summary still saved)")

    # re-append the sheets captured above, after the rebuilt ones
    for nm, rws in preserved.items():
        wsx = wb.create_sheet(nm)
        for r in rws:
            wsx.append(r)
        if rws:
            for c in wsx[1]:
                c.font = hdr_font
                c.fill = hdr_fill

    wb.save(MASTER_SUMMARY_XLSX)

    # ---- rebuild charts (matplotlib PNGs + chart_config sheet) ----
    # Uses chart_builder.py sitting next to this script. Idempotent.
    try:
        from chart_builder import rebuild_charts
        rebuild_charts(MASTER_SUMMARY_XLSX)
    except Exception as e:
        print(f"[warn] chart rebuild failed: {e}  (summary still saved)")
    print(f"[master] {MASTER_SUMMARY_XLSX}  (total rows: {len(combined)})")


# ============================================================
def main():
    global FILE_PREFIX, NTOP_FILE, TEMPLATE_IN
    dry = "--dry-run" in sys.argv
    # Force a schema rebuild even when the mtime check says it is current
    # (e.g. the .ntop was restored from a backup with an older timestamp).
    refresh_schema = "--refresh-schema" in sys.argv
    # Auto-answer [A] to every duplicate prompt. Used by the 3-type smoke test
    # so the whole check runs unattended; never changes what is simulated, only
    # who answers the rerun question.
    rerun_all = "--rerun-all" in sys.argv
    # Run a different ADMS variant without editing runs_input.xlsx. The type
    # still comes from a setting by default; this only overrides it for one
    # invocation.
    type_override = _arg_value("--type")
    ntop_override = _arg_value("--ntop")
    # Run a DIFFERENT parameter sheet without editing runs_input.xlsx, so an
    # unattended overnight batch can chain several per-type sheets in one go.
    input_override = _arg_value("--input")
    print(f"nTop batch runner - {SCRIPT_DIR}")

    global INPUT_XLSX
    if input_override:
        cand = Path(input_override)
        if not cand.is_absolute():
            cand = SCRIPT_DIR / cand
        if not cand.exists():
            die(f"--input {input_override}: file not found ({cand})")
        INPUT_XLSX = cand
        print(f"[cli] --input overrides the parameter sheet -> {INPUT_XLSX.name}")

    # Load settings FIRST so we can pick the ADMS type + notebook before schema.
    rows, settings = load_input_xlsx()
    if not rows:
        die("no parameter rows found in runs_input.xlsx (sheet 'runs')")

    # --- multi-type selection (settings sheet: adms_type + ntop_file) ---
    if type_override:
        settings["adms_type"] = type_override.strip()
        settings["ntop_file"] = ntop_override or resolve_notebook_for_type(
            settings["adms_type"])
        print(f"[cli] --type {settings['adms_type']} overrides the settings "
              f"sheet -> notebook {settings['ntop_file']}")
    elif ntop_override:
        settings["ntop_file"] = ntop_override.strip()
        print(f"[cli] --ntop overrides the settings sheet -> {settings['ntop_file']}")

    FILE_PREFIX = f"ADMS_{settings['adms_type']}"
    NTOP_FILE   = SCRIPT_DIR / settings["ntop_file"]
    TEMPLATE_IN = SCRIPT_DIR / f"input_template_{settings['adms_type']}.json"
    print(f"[type] {settings['adms_type']}  ->  notebook {NTOP_FILE.name}, "
          f"file prefix {FILE_PREFIX}")
    if not NTOP_FILE.exists():
        die(f"notebook not found next to the script: {NTOP_FILE.name}")

    ntopcl = None if dry else find_ntopcl(settings)
    if not dry:
        print(f"[ok] ntopcl: {ntopcl}")
        # regenerates the per-type template if missing OR if the .ntop is newer
        ensure_schema(ntopcl, force=refresh_schema)
    elif not TEMPLATE_IN.exists():
        die(f"{TEMPLATE_IN.name} missing (needed for --dry-run)")

    print(f"[load] {len(rows)} run(s) from runs_input.xlsx")
    print(f"[settings] sigma_ys = {settings['sigma_ys_mpa']} MPa,  "
          f"Es = {settings['es_gpa']} GPa,  nu_s = {settings.get('nu_s', DEFAULT_SETTINGS['nu_s'])},  "
          f"strain = {settings['strain']}")

    # Publish the material properties to the summary/chart builders, which run
    # outside this function's scope. Without this they would silently fall back
    # to steel and mis-scale every chart for a different base material.
    global ES_GPA_ACTIVE, N_STIFF_ACTIVE
    ES_GPA_ACTIVE = settings["es_gpa"]
    N_STIFF_ACTIVE = float(settings.get("n_stiff", N_STIFF_ACTIVE))
    try:
        import chart_builder as _cb
        _cb.ES_GPA = settings["es_gpa"]
    except Exception:
        pass   # charts are optional; never block a simulation over plotting

    # defaults from schema for anything not given in a row
    schema_defaults = {e["name"]: e.get("value") for e in load_schema()["inputs"]
                       if e.get("name") not in (NAME_OUTPATH, NAME_STRESSPATH,
                                                NAME_SHEAR_STRESSPATH, NAME_STLPATH)
                       and isinstance(e.get("value"), (int, float))}

    # --- UPFRONT: cross-check runs vs master, prompt user for reruns ---
    # (all input happens BEFORE any simulation, so user can walk away after)
    master_rows = load_master_rows()
    master_counts = master_stem_counts(master_rows)
    if master_rows:
        print(f"[master] found {len(master_rows)} existing rows in "
              f"{MASTER_SUMMARY_XLSX.name}")
    else:
        print(f"[master] no existing master summary (fresh dataset)")

    to_run, updated_counts = prompt_reruns(rows, schema_defaults, master_counts,
                                           auto_yes=rerun_all)
    if not to_run:
        print("\n[abort] nothing left to run after skip choices - exiting.")
        return
    # stem -> version label to use for this batch's summary rows
    stem_versions = {stem: ver for _, stem, ver in to_run}

    if dry:
        print("\n[dry-run] the runs above would execute in order. Exiting.")
        return

    # Create a new timestamped output folder for THIS invocation.
    # e.g. Results/20260713_17-24_RUN/{Data,Results_summary.xlsx}
    _init_batch_dir()

    # --- SIMULATE (no more prompts from here on) ---
    t0 = time.time()
    results = []
    for i, (iv, stem, ver) in enumerate(to_run, 1):
        print(f"\n[{i}/{len(to_run)}]  {stem}  ({ver})")
        results.append(run_one(ntopcl, iv, schema_defaults))

    print(f"\n{'=' * 62}\nBATCH DONE ({(time.time() - t0) / 60:.1f} min total)")
    n_ok = sum(1 for r in results if r["ok"])
    print(f"  {n_ok}/{len(results)} runs OK")

    # --- BATCH summary ---
    print("\n[post] building this batch's Results_summary.xlsx ...")
    srows = build_summary(settings, stem_versions=stem_versions)
    print(f"[ok] {SUMMARY_XLSX}  ({len(srows)} rows)")

    # --- MASTER summary (append-only across all batches) ---
    print("\n[post] updating master Results_summary.xlsx ...")
    update_master(srows)

    print()
    for r in srows:
        e  = r.get("E_iso_GPa")
        g  = r.get("G_iso_GPa")
        yc = r.get("Yield_onset_MPa")
        ys = r.get("Shear_onset_MPa")
        ver = r.get("Version") or ""
        print(f"  {r['Run']:<42} {ver:<3} {r.get('Status', ''):<8} "
              f"rho={r.get('Rho_rel') or 'n/a'}  "
              f"E={e if e is not None else 'n/a'} G={g if g is not None else 'n/a'} GPa  "
              f"comp_onset={yc if yc is not None else 'n/a'}  "
              f"shear_onset={ys if ys is not None else 'n/a'} MPa")


if __name__ == "__main__":
    main()
