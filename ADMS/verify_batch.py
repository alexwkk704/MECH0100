#!/usr/bin/env python3
"""verify_batch.py - did a batch actually collect everything the ML needs?

Run after RUN_SIMULATION.bat / SMOKE_TEST_3TYPES.bat. Checks the N newest batch
folders under Results/ and reports, per run:

  FILES   the C-tensor CSV, both von-Mises CSVs, the STL, and outputs.json
  LABELS  every column the ML actually reads, taken from the 'label_map' sheet
          of ML_settings.xlsx - NOT a list written into this file. Add a target
          in the workbook and this check picks it up automatically.
  TABLE   the curvature table came back through outputs.json (p_rel / avg_Mn_K /
          avg_Mn_H), which is what makes the run "native, not backfilled".

Exit code 0 = everything present, 1 = something missing (so a .bat can stop).

    python verify_batch.py            # 3 newest batches
    python verify_batch.py --last 1
    python verify_batch.py --batch 20260805_19-00_RUN
"""
import json
import sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
RESULTS_DIR = SCRIPT_DIR / "Results"
ML_SETTINGS = SCRIPT_DIR.parent / "ML" / "ML_settings.xlsx"


def die(msg):
    print(f"\n[ERROR] {msg}")
    sys.exit(1)


def arg(flag, default=None):
    for i, a in enumerate(sys.argv):
        if a == flag and i + 1 < len(sys.argv):
            return sys.argv[i + 1]
        if a.startswith(flag + "="):
            return a.split("=", 1)[1]
    return default


def ml_label_columns():
    """Source columns the ML reads, straight from ML_settings.xlsx 'label_map'.

    Returns (targets, crosschecks). Falls back to an empty list with a loud
    warning if the workbook cannot be read - never silently to a guess.
    """
    if not ML_SETTINGS.exists():
        print(f"[!] {ML_SETTINGS} not found - cannot check label completeness")
        return [], []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(ML_SETTINGS, data_only=True)
    except Exception as e:
        print(f"[!] could not read {ML_SETTINGS.name} ({e}) - label check skipped")
        return [], []
    if "label_map" not in wb.sheetnames:
        print("[!] ML_settings.xlsx has no 'label_map' sheet - label check skipped")
        return [], []
    tg, cc = [], []
    for r in wb["label_map"].iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        src = str(r[0]).strip()
        role = (str(r[4]).strip().lower() if len(r) > 4 and r[4] is not None
                else "target")
        (cc if role == "crosscheck" else tg).append(src)
    return tg, cc


def summary_rows(batch_dir):
    xlsx = batch_dir / "Results_summary.xlsx"
    if not xlsx.exists():
        return None, f"no Results_summary.xlsx in {batch_dir.name}"
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    name = "summary" if "summary" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[name]
    hdr = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = {h: ws.cell(r, i + 1).value for i, h in enumerate(hdr) if h}
        if d.get("Run"):
            rows.append(d)
    return rows, None


def check_run_files(rdir):
    stem = rdir.name
    want = {
        "C tensor":     rdir / f"{stem}.csv",
        "compression":  rdir / f"{stem}_stress.csv",
        "shear":        rdir / f"{stem}_shear_stress.csv",
        "STL":          rdir / f"{stem}.stl",
        "outputs.json": rdir / "outputs.json",
    }
    missing = []
    for label, p in want.items():
        floor = 1000 if label == "STL" else 20
        if not p.exists():
            missing.append(f"{label} MISSING")
        elif p.stat().st_size < floor:
            missing.append(f"{label} EMPTY ({p.stat().st_size} B)")
    return missing


def check_output_table(rdir):
    """outputs.json must carry the curvature table, else curvature is backfill."""
    p = rdir / "outputs.json"
    if not p.exists():
        return "outputs.json missing"
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        return f"outputs.json unreadable ({e})"
    for entry in data if isinstance(data, list) else []:
        if isinstance(entry, dict) and entry.get("type") == "table":
            cols = ((entry.get("value") or {}).get("table") or {}).get("columns", [])
            names = [str(c.get("name", "")) for c in cols]
            if len(names) < 3:
                return f"table has only {len(names)} column(s): {names}"
            return None
    return "no table output (curvature would have to be backfilled)"


def main():
    if not RESULTS_DIR.is_dir():
        die(f"no Results folder next to the script ({RESULTS_DIR})")

    one = arg("--batch")
    if one:
        batches = [RESULTS_DIR / one]
        if not batches[0].is_dir():
            die(f"batch folder not found: {batches[0]}")
    else:
        n = int(arg("--last", "3"))
        allb = sorted((d for d in RESULTS_DIR.iterdir()
                       if d.is_dir() and (d / "Data").is_dir()),
                      key=lambda d: d.stat().st_mtime)
        if not allb:
            die("no batch folders with a Data/ subfolder under Results/")
        batches = allb[-n:]

    targets, crosschecks = ml_label_columns()
    print(f"ML label columns read from {ML_SETTINGS.name}:")
    print(f"  targets     ({len(targets)}): {targets}")
    print(f"  crosschecks ({len(crosschecks)}): {crosschecks}")

    bad = 0
    total = 0
    for b in batches:
        print(f"\n{'=' * 68}\nBATCH {b.name}")
        rows, err = summary_rows(b)
        if err:
            print(f"  [FAIL] {err}")
            bad += 1
            continue
        by_run = {r["Run"]: r for r in rows}
        for rdir in sorted((b / "Data").iterdir()):
            if not rdir.is_dir():
                continue
            total += 1
            problems = check_run_files(rdir)
            t = check_output_table(rdir)
            if t:
                problems.append(f"TABLE: {t}")
            row = by_run.get(rdir.name)
            if row is None:
                problems.append("no row in Results_summary.xlsx")
            else:
                for c in targets + crosschecks:
                    if c not in row:
                        problems.append(f"column {c} not in summary")
                    elif row[c] in (None, ""):
                        problems.append(f"{c} EMPTY")
            if problems:
                bad += 1
                print(f"  [FAIL] {rdir.name}")
                for p in problems:
                    print(f"         - {p}")
            else:
                curv = (f"rho={row.get('Rho_rel')} "
                        f"K={row.get('GaussCurv_STL')} H={row.get('MeanCurv_STL')}")
                print(f"  [ OK ] {rdir.name}\n         {curv}")

    print(f"\n{'=' * 68}")
    if bad:
        print(f"RESULT: {bad} problem(s) across {total} run(s). "
              f"Do NOT expand the database until these are clean.")
        sys.exit(1)
    print(f"RESULT: all {total} run(s) complete - every file and every ML "
          f"column present. Safe to expand the database.")
    sys.exit(0)


if __name__ == "__main__":
    main()
