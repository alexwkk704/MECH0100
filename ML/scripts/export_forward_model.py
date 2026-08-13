"""
export_forward_model.py - JOB 10: freeze the forward model.

This is the LAST step of the forward ML pipeline (Share\\ML). Steps 1-9 build the
database and measure how well the model generalises; this step takes the same
validated recipe, trains one model on every row, and writes a self-contained
bundle that anything downstream can load without re-reading the training data.

Inverse design (Share\\Inverse design) is a CONSUMER of what this writes. It is a
separate namespace and nothing in it is imported here.

WHAT IT WRITES  ->  Final Model\\<PROFILE>\\
    model_<PROFILE>.pt        weights + targets + normalisation + prep spec
    model_card.txt            plain-English description of exactly what was trained
    gate0_reference.csv       in-sample predictions on every training row
                              (GATE 0 replays 10 of these through predict.py)

WHY THE BUNDLE CARRIES SO MUCH
    A .pt with only weights is a trap: the caller has to re-derive the target
    order, the log flag, the standardisation constants and the point-cloud
    preparation, and any one of those silently returns plausible wrong numbers.
    Everything needed to reproduce a prediction is stored IN the bundle.

SEED
    Training is stochastic (weight init, batch order, jitter, rotation draws,
    dropout). The fold assignment is NOT - cv_split greedy-balances groups with
    no RNG. So a different seed gives a different model, and the seed is recorded
    in the bundle and the card.
    ensemble_seeds > 1 trains that many models on the same data and averages
    their predictions, which REDUCES seed variance instead of reporting around
    it. Selecting the best-scoring seed after the fact would be selection on
    validation data - this script deliberately offers no way to do that.

NOTHING IS HARDCODED
    Every path and hyper-parameter comes from ML_settings.xlsx via run_paths.
    The workbook is opened READ-ONLY: unlike train_pointnet_v2.load_pn_cfg this
    script never backfills or saves, so exporting can never mutate the settings
    a run is using.

Usage:  set ML_PROFILE=COMBINED  &  python export_forward_model.py
        (or just double-click 10_EXPORT_FORWARD_MODEL.bat)
"""

from __future__ import annotations
import os, sys, re, glob, time, json, hashlib
import numpy as np
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import run_paths                     # noqa: E402
import tensor_ops                    # noqa: E402

SETTINGS = os.path.join(run_paths.BASE_DIR, "ML_settings.xlsx")


def die(msg):
    print(f"\n[E] {msg}")
    sys.exit(1)


# ---------------------------------------------------------------------------
# settings (READ-ONLY - never save the workbook from here)
# ---------------------------------------------------------------------------
def load_spec():
    import openpyxl
    if not os.path.exists(SETTINGS):
        die(f"settings workbook not found: {SETTINGS}")
    wb = openpyxl.load_workbook(SETTINGS, data_only=True)

    pn_sheet = run_paths.sheet_for(wb, "pointnet")
    if pn_sheet not in wb.sheetnames:
        die(f"sheet '{pn_sheet}' not in {os.path.basename(SETTINGS)}")
    pn = {str(k).strip(): v
          for k, v, *_ in wb[pn_sheet].iter_rows(min_row=2, values_only=True) if k}

    def lst(x):
        return [s.strip() for s in str(x or "").split(",") if s.strip()]

    # es_gpa: the 'labels' sheet sets the material constant, the pointnet sheet
    # overrides it. For COMBINED the tensors are cached already normalised, so
    # the override is 1.0 - that is what lets steel ADMS and polymer TPMS share
    # one dataset. Replicated EXACTLY from train_pointnet_v2.load_pn_cfg; getting
    # this wrong corrupts every rotation-augmented label.
    es = 200.0
    if "labels" in wb.sheetnames:
        lab = {str(k).strip(): v
               for k, v, *_ in wb["labels"].iter_rows(min_row=2, values_only=True) if k}
        try:
            es = float(lab.get("es_gpa", 200.0))
        except (TypeError, ValueError):
            es = 200.0
    try:
        if pn.get("es_gpa") not in (None, ""):
            es = float(pn["es_gpa"])
    except (TypeError, ValueError):
        pass

    # point-cloud preparation - predict.py MUST reuse these or the clouds it
    # builds are not the clouds the net was trained on.
    pc_sheet = run_paths.sheet_for(wb, "pointcloud")
    pc = {str(k).strip(): v
          for k, v, *_ in wb[pc_sheet].iter_rows(min_row=2, values_only=True) if k} \
        if pc_sheet in wb.sheetnames else {}

    # where the STLs live, per family - GATE 0 needs to find the originals.
    stl_folders = []
    for sh in wb.sheetnames:
        if sh == "settings" or sh.startswith("settings_"):
            for k, v, *_ in wb[sh].iter_rows(min_row=2, values_only=True):
                if k and str(k).strip() == "stl_folder" and v:
                    p = str(v)
                    p = p if os.path.isabs(p) else os.path.normpath(
                        os.path.join(run_paths.BASE_DIR, p))
                    if p not in stl_folders:
                        stl_folders.append(p)

    spec = dict(
        targets=lst(pn["targets"]),
        n_points=int(pn["n_points"]),
        epochs=int(pn["epochs"]),
        batch_size=int(pn["batch_size"]),
        lr=float(pn["lr"]),
        lr_decay=float(pn.get("lr_decay", 1.0) or 1.0),
        lr_step=int(pn.get("lr_step", 200) or 200),
        jitter_std=float(pn.get("jitter_std", 0.0) or 0.0),
        rotation_aug=bool(int(pn.get("rotation_aug", 0) or 0)),
        n_rot=int(pn.get("n_rot", 0) or 0),
        frame_dependent=set(lst(pn.get("frame_dependent", ""))),
        rotation_invariant=set(lst(pn.get("rotation_invariant", ""))),
        use_density=bool(int(pn.get("use_density_channel", 0) or 0)),
        use_log=bool(int(pn.get("log_targets", 0) or 0)),
        deep=bool(int(pn.get("deep_head", 0) or 0)),
        grad_clip=float(pn.get("grad_clip", 0) or 0),
        seed=int(pn.get("random_seed", 1) or 1),
        n_folds=int(pn.get("n_folds", 5) or 5),
        group_cols=lst(pn.get("group_cols", "")),
        es_gpa=es,
        pc_seed=int(pc.get("random_seed", 1) or 1),
        pc_n_points=int(pc.get("n_points", pn["n_points"]) or pn["n_points"]),
        stl_folders=stl_folders,
        pointnet_sheet=pn_sheet,
    )
    if spec["pc_n_points"] != spec["n_points"]:
        print(f"[!] pointcloud sheet n_points={spec['pc_n_points']} but pointnet "
              f"sheet n_points={spec['n_points']}. The stored clouds have the "
              f"pointcloud value; the net slices to the pointnet value.")
    return spec


# ---------------------------------------------------------------------------
# data - matched EXACTLY to train_pointnet_v2.build_xy, loose match included
# ---------------------------------------------------------------------------
def build_xy(ds_csv, pc_dir, n_points):
    df = pd.read_csv(ds_csv)
    stem_col = "Run" if "Run" in df.columns else "file"
    npz = {os.path.splitext(os.path.basename(p))[0]: p
           for p in glob.glob(os.path.join(pc_dir, "*.npz"))}
    X, rows = [], []
    for _, r in df.iterrows():
        stem = re.sub(r"\.stl$", "", str(r[stem_col]), flags=re.I)
        path = npz.get(stem)
        if path is None:                      # loose contains-match, as in training
            hit = [v for k, v in npz.items() if stem in k or k in stem]
            path = hit[0] if hit else None
        if path is None:
            continue
        pts = np.load(path)["pts"].astype(np.float32)
        if len(pts) >= n_points:
            pts = pts[:n_points]
        else:
            idx = np.random.choice(len(pts), n_points, replace=True)
            pts = pts[idx]
        X.append(pts)
        rows.append(r)
    if not X:
        die(f"no point clouds matched. dataset {os.path.basename(ds_csv)} "
            f"({len(df)} rows) vs {len(npz)} .npz in {pc_dir}")
    print(f"[i] matched {len(X)}/{len(df)} rows to clouds ({len(npz)} .npz available)")
    return np.stack(X), pd.DataFrame(rows).reset_index(drop=True), stem_col


def build_net(nn, n_out, in_dim, deep):
    """Byte-identical to train_pointnet_v2.PointNet - verified 12/08/2026 by
    diffing the two class bodies. If that file's architecture changes this must
    change with it, or torch.load will fail on a key mismatch (loudly, at least)."""
    class PointNet(nn.Module):
        def __init__(self):
            super().__init__()
            self.enc = nn.Sequential(
                nn.Conv1d(in_dim, 64, 1), nn.BatchNorm1d(64), nn.ReLU(),
                nn.Conv1d(64, 128, 1), nn.BatchNorm1d(128), nn.ReLU(),
                nn.Conv1d(128, 1024, 1), nn.BatchNorm1d(1024), nn.ReLU())
            if deep:
                self.head = nn.Sequential(
                    nn.Linear(1024, 512), nn.BatchNorm1d(512), nn.ReLU(), nn.Dropout(0.3),
                    nn.Linear(512, 256), nn.BatchNorm1d(256), nn.ReLU(), nn.Dropout(0.3),
                    nn.Linear(256, n_out))
            else:
                self.head = nn.Sequential(
                    nn.Linear(1024, 256), nn.ReLU(), nn.Dropout(0.3),
                    nn.Linear(256, 64), nn.ReLU(),
                    nn.Linear(64, n_out))

        def forward(self, x):
            x = x.transpose(1, 2)
            x = self.enc(x).max(dim=2).values
            return self.head(x)
    return PointNet()


def sha1(path, n=1 << 20):
    h = hashlib.sha1()
    with open(path, "rb") as f:
        while True:
            b = f.read(n)
            if not b:
                break
            h.update(b)
    return h.hexdigest()[:16]


# ---------------------------------------------------------------------------
def main():
    prof = run_paths.profile() or "ADMS"
    n_ens = max(1, int(os.environ.get("ML_ENSEMBLE", "1") or 1))
    s = load_spec()

    ds_csv = run_paths.dataset_path()
    pc_dir = run_paths.pointclouds_dir()
    tens_npz = run_paths.tensor_npz_path()
    out_dir = os.path.join(run_paths.BASE_DIR, "Final Model", prof)
    os.makedirs(out_dir, exist_ok=True)

    print("=" * 68)
    print(f"  EXPORT FORWARD MODEL   profile={prof}   sheet={s['pointnet_sheet']}")
    print("=" * 68)
    print(f"  dataset   {ds_csv}")
    print(f"  clouds    {pc_dir}")
    print(f"  tensors   {tens_npz}")
    print(f"  output    {out_dir}")
    print(f"  seeds     {[s['seed'] + i for i in range(n_ens)]}"
          f"{'   (ensemble - predictions averaged)' if n_ens > 1 else ''}")
    if not os.path.exists(ds_csv):
        die(f"dataset missing: {ds_csv} - run the setup/merge steps first")

    try:
        import torch
        import torch.nn as nn
    except ImportError:
        die("PyTorch not installed - run: pip install torch")
    dev = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"  device    {dev}")

    np.random.seed(s["seed"])
    X, meta, stem_col = build_xy(ds_csv, pc_dir, s["n_points"])

    tgts = [t for t in s["targets"] if t in meta.columns]
    missing = [t for t in s["targets"] if t not in meta.columns]
    if missing:
        print(f"[!] targets absent from the dataset, skipped: {missing}")
    if not tgts:
        die("none of the configured targets exist in the dataset")

    Y = meta[tgts].apply(pd.to_numeric, errors="coerce").values.astype(np.float32)
    MASK = ~np.isnan(Y)
    keep = MASK.any(axis=1)
    X, Y, MASK = X[keep], Y[keep], MASK[keep]
    meta = meta[keep].reset_index(drop=True)
    print("[i] labels per target: "
          + ", ".join(f"{t}={int(MASK[:, j].sum())}" for j, t in enumerate(tgts)))

    # ---- physical label range, stored for GATE 4 (reject out-of-range asks) ----
    label_min = {t: float(np.nanmin(Y[:, j][MASK[:, j]])) for j, t in enumerate(tgts)}
    label_max = {t: float(np.nanmax(Y[:, j][MASK[:, j]])) for j, t in enumerate(tgts)}

    # ---- density channel ----
    dens_col = None
    if s["use_density"]:
        if "Rho_rel" not in meta.columns:
            die("use_density_channel=1 but the dataset has no 'Rho_rel' column")
        dens_col = "Rho_rel"
        rho = pd.to_numeric(meta[dens_col], errors="coerce").values.astype(np.float32)
        if np.isnan(rho).any():
            die(f"{int(np.isnan(rho).sum())} rows have a blank {dens_col}")
        X = np.concatenate(
            [X, np.repeat(rho[:, None, None], X.shape[1], axis=1)], axis=2).astype(np.float32)
        print(f"[i] density channel ON from '{dens_col}' -> {X.shape}")
    in_dim = X.shape[2]

    # ---- log transform ----
    use_log = s["use_log"]
    if use_log:
        if (Y[MASK] <= 0).any():
            print("[!] non-positive target values present -> log transform DISABLED")
            use_log = False
        else:
            Y = np.where(MASK, np.log10(np.where(MASK, Y, 1.0)), np.nan).astype(np.float32)
            print("[i] targets log10-transformed")

    ymean = np.array([Y[:, j][MASK[:, j]].mean() if MASK[:, j].any() else 0.0
                      for j in range(len(tgts))], np.float32)
    ystd = np.array([Y[:, j][MASK[:, j]].std() if MASK[:, j].any() else 1.0
                     for j in range(len(tgts))], np.float32) + 1e-8
    Ys = np.where(MASK, (Y - ymean) / ystd, 0.0).astype(np.float32)
    Ms = MASK.astype(np.float32)

    # ---- rotation augmentation, exactly as fit_fold does it ----
    Ctens = None
    if s["rotation_aug"] and s["n_rot"] > 0:
        if not os.path.exists(tens_npz):
            die(f"rotation_aug=1 but the tensor cache is missing: {tens_npz}\n"
                f"    run collect_tensors.py (step 2/3 of the pipeline) first")
        z = np.load(tens_npz)
        store = {k: z[k] for k in z.files}
        Ctens, have = [], 0
        for _, r in meta.iterrows():
            stem = re.sub(r"\.stl$", "", str(r[stem_col]), flags=re.I)
            C = store.get(stem)
            if C is None:
                hit = [v for k, v in store.items() if stem in k or k in stem]
                C = hit[0] if hit else None
            Ctens.append(None if C is None else np.asarray(C, float))
            have += C is not None
        print(f"[i] rotation_aug: matched {have}/{len(meta)} tensors")
        if have == 0:
            die("rotation_aug=1 but not one tensor matched - the augmented rows "
                "would all be dropped and this export would NOT match the "
                "validated recipe. Fix collect_tensors.py first.")

    def augment_rows(seed_off):
        """Rotated copies of every row. Labels recomputed from the rotated tensor,
        cloud re-centred and re-scaled to the unit box - both exactly as fit_fold."""
        if Ctens is None:
            return None
        rng = np.random.RandomState(s["seed"] + 1000 + seed_off)
        fdep, finv = s["frame_dependent"], s["rotation_invariant"]
        aX, aY, aM = [], [], []
        for i in range(len(X)):
            C = Ctens[i]
            if C is None:
                continue
            for _ in range(s["n_rot"]):
                R = tensor_ops.random_rotation(rng)
                xc = X[i].copy()
                p = X[i][:, :3] @ R.T
                p = p - p.mean(axis=0)
                ext = float(np.ptp(p, axis=0).max())
                if ext > 0:
                    p = p / ext
                xc[:, :3] = p
                Crot = tensor_ops.rotate_C(C, R)
                yrow = np.zeros(len(tgts), np.float32)
                mrow = np.zeros(len(tgts), np.float32)
                for j, t in enumerate(tgts):
                    if t in fdep:
                        val = tensor_ops.frame_dependent_value(t, Crot, s["es_gpa"])
                        if val is None or (use_log and val <= 0):
                            continue
                        v = np.log10(val) if use_log else val
                        yrow[j] = (v - ymean[j]) / ystd[j]
                        mrow[j] = 1.0
                    elif t in finv:
                        yrow[j] = Ys[i, j]
                        mrow[j] = Ms[i, j]
                aX.append(xc.astype(np.float32))
                aY.append(yrow)
                aM.append(mrow)
        if not aX:
            return None
        return np.stack(aX), np.stack(aY), np.stack(aM)

    def masked_mse(pred, target, mask):
        d = (pred - target) * mask
        return (d ** 2).sum() / torch.clamp(mask.sum(), min=1.0)

    # ---- train one model per seed ----
    states, preds = [], []
    for k in range(n_ens):
        seed = s["seed"] + k
        print(f"\n--- training seed {seed} ({k + 1}/{n_ens}) "
              f"on ALL {len(X)} rows -----------------------")
        np.random.seed(seed)
        torch.manual_seed(seed)

        Xtr, Ytr, Mtr = X, Ys, Ms
        aug = augment_rows(k)
        if aug is not None:
            Xtr = np.concatenate([X, aug[0]], axis=0)
            Ytr = np.concatenate([Ys, aug[1]], axis=0)
            Mtr = np.concatenate([Ms, aug[2]], axis=0)
            print(f"    +{len(aug[0])} rotated copies -> {len(Xtr)} training rows")

        net = build_net(nn, len(tgts), in_dim, s["deep"]).to(dev)
        opt = torch.optim.Adam(net.parameters(), lr=s["lr"])
        sched = (torch.optim.lr_scheduler.StepLR(opt, s["lr_step"], s["lr_decay"])
                 if s["lr_decay"] != 1.0 else None)
        t0 = time.time()
        for ep in range(s["epochs"]):
            net.train()
            perm = np.random.permutation(len(Xtr))
            last = float("nan")
            for b in range(0, len(Xtr), s["batch_size"]):
                bi = perm[b:b + s["batch_size"]]
                if len(bi) < 2:          # BatchNorm needs >= 2 in train mode
                    continue
                xb = Xtr[bi]
                if s["jitter_std"] > 0:
                    xb = xb + np.random.normal(0, s["jitter_std"], xb.shape).astype(np.float32)
                xb = torch.tensor(xb, device=dev)
                yb = torch.tensor(Ytr[bi], device=dev)
                mb = torch.tensor(Mtr[bi], device=dev)
                opt.zero_grad()
                loss = masked_mse(net(xb), yb, mb)
                loss.backward()
                if s["grad_clip"] > 0:
                    nn.utils.clip_grad_norm_(net.parameters(), s["grad_clip"])
                opt.step()
                last = float(loss.item())
            if sched:
                sched.step()
            if ep == 0 or (ep + 1) % 50 == 0 or ep == s["epochs"] - 1:
                print(f"    epoch {ep + 1}/{s['epochs']}  loss={last:.5f}  "
                      f"({time.time() - t0:.0f}s)")

        net.eval()
        with torch.no_grad():
            p = []
            for b in range(0, len(X), 32):
                p.append(net(torch.tensor(X[b:b + 32], device=dev)).cpu().numpy())
            p = np.concatenate(p, axis=0)
        preds.append(p)
        states.append({k2: v.detach().cpu().clone() for k2, v in net.state_dict().items()})

    # ---- in-sample reference predictions (GATE 0 replays these) ----
    P = np.mean(np.stack(preds), axis=0) * ystd + ymean
    if use_log:
        P = np.power(10.0, P)

    bundle = dict(
        state_dicts=states,                      # list; length 1 unless ensembling
        state_dict=states[0],                    # convenience for single-model use
        n_models=len(states),
        arch="pointnet_v2",
        targets=tgts, in_dim=in_dim, deep=bool(s["deep"]),
        ymean=ymean, ystd=ystd,
        use_log=bool(use_log), use_density=bool(s["use_density"]),
        density_column=dens_col,
        n_points=s["n_points"],
        pc_seed=s["pc_seed"],
        prep=dict(sampler="trimesh.sample.sample_surface",
                  centre="subtract mean of sampled points",
                  scale="divide by max peak-to-peak extent (unit box)",
                  source="pointcloud_prep.process - imported, not reimplemented"),
        n_train=int(len(X)), seed=s["seed"],
        seeds=[s["seed"] + i for i in range(n_ens)],
        profile=prof,
        label_min=label_min, label_max=label_max,
        vf_min=float(pd.to_numeric(meta["VF"], errors="coerce").min()) if "VF" in meta else None,
        vf_max=float(pd.to_numeric(meta["VF"], errors="coerce").max()) if "VF" in meta else None,
        es_gpa=s["es_gpa"],
        epochs=s["epochs"], batch_size=s["batch_size"], lr=s["lr"],
        jitter_std=s["jitter_std"], rotation_aug=s["rotation_aug"], n_rot=s["n_rot"],
        dataset_name=os.path.basename(ds_csv), dataset_sha1=sha1(ds_csv),
        stl_folders=s["stl_folders"],
        built_utc=time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()),
    )
    model_path = os.path.join(out_dir, f"model_{prof}.pt")
    torch.save(bundle, model_path)
    print(f"\n[OK] {model_path}")

    ref = pd.DataFrame({stem_col: meta[stem_col].values})
    for c in ("family", "topology", "VF", "Rho_rel"):
        if c in meta.columns:
            ref[c] = meta[c].values
    for j, t in enumerate(tgts):
        ref[f"true_{t}"] = meta[t].values
        ref[f"pred_{t}"] = P[:, j]
    ref_path = os.path.join(out_dir, "gate0_reference.csv")
    ref.to_csv(ref_path, index=False)
    print(f"[OK] {ref_path}")

    with open(os.path.join(out_dir, "model_card.txt"), "w") as f:
        f.write(f"FORWARD MODEL - {prof}\n")
        f.write(f"built (UTC)     : {bundle['built_utc']}\n")
        f.write(f"trained on      : ALL {len(X)} rows of {bundle['dataset_name']} "
                f"(sha1 {bundle['dataset_sha1']})\n")
        if "family" in meta.columns:
            f.write("                  " + ", ".join(
                f"{k} {v}" for k, v in meta['family'].value_counts().items()) + "\n")
        f.write(f"targets         : {', '.join(tgts)}\n")
        f.write(f"seeds           : {bundle['seeds']}"
                f"{'  (ensemble, predictions averaged)' if n_ens > 1 else ''}\n")
        f.write(f"n_points        : {s['n_points']}   density channel: {s['use_density']}"
                f"{f' (from {dens_col})' if dens_col else ''}\n")
        f.write(f"log10 targets   : {use_log}    deep head: {s['deep']}\n")
        f.write(f"rotation_aug    : {s['rotation_aug']} (n_rot={s['n_rot']}, "
                f"es_gpa={s['es_gpa']})\n")
        f.write(f"epochs          : {s['epochs']}  batch {s['batch_size']}  lr {s['lr']}\n")
        f.write("\nNO EARLY STOPPING. The epoch count is the CV-validated budget from\n")
        f.write("the settings sheet, used as-is. The cross-validated runs picked a best\n")
        f.write("epoch against a validation slice; a model trained on 100% of the rows\n")
        f.write("has no such slice, so the budget is fixed in advance instead. State\n")
        f.write("this in the thesis rather than implying early stopping was used.\n")
        f.write("\nTRAINING LABEL RANGE (GATE 4 rejects asks outside this):\n")
        for t in tgts:
            f.write(f"  {t:<12} {label_min[t]:.6g} .. {label_max[t]:.6g}\n")
        f.write("\nDENSITY CHANNEL SOURCE - READ BEFORE PREDICTING ON NEW GEOMETRY\n")
        f.write("  Training used the dataset's Rho_rel column.\n")
        f.write("  TPMS rows: Rho_rel == VF measured from the STL, exactly (0.000%).\n")
        f.write("  ADMS rows: Rho_rel is nTop's p_rel and differs from STL VF by a\n")
        f.write("  median 0.33% and up to 6.59%. For a new ADMS candidate feed nTop's\n")
        f.write("  p_rel, not a measured VF.\n")
        f.write("\nspec source     : ML_settings.xlsx sheet "
                f"'{s['pointnet_sheet']}', opened read-only\n")
    print(f"[OK] {os.path.join(out_dir, 'model_card.txt')}")

    with open(os.path.join(out_dir, "bundle_keys.json"), "w") as f:
        json.dump({k: (str(type(v).__name__) if k.startswith("state") else v)
                   for k, v in bundle.items()
                   if k not in ("state_dict", "state_dicts")}, f, indent=2, default=str)

    print("\nnext: GATE 0 replays 10 of these rows from their ORIGINAL STL through "
          "predict.py.\n")


if __name__ == "__main__":
    main()
