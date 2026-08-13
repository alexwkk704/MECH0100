"""
validate_dataset.py — prove the extracted features describe the real geometry,
BEFORE any model is trained on them.

Runs automatically as part of RUN_ALL_ML.bat, between the join and the training.
Model accuracy is meaningless if the inputs are wrong, so this gate comes first.

The checks
----------
1. DENSITY CROSS-CHECK (the important one)
   VF is measured from the STL surface. Rho_rel is measured from the FEA volume
   mesh. They are produced by completely separate tools that never see each
   other. If they agree, the STL reading is sound. If they diverge, something
   upstream is wrong and no model result should be trusted.

2. WALL THICKNESS
   Measured wall thickness vs the t encoded in the filename. Confirms the
   thickness_convention setting (total vs double) is right for this dataset —
   a wrong setting here silently halves or doubles a feature.

3. DATA INTEGRITY
   NaNs, non-positive volume fractions, degenerate meshes.

4. FEATURE DISTRIBUTIONS
   Flags heavy skew, which hurts GPR (it assumes roughly Gaussian inputs).
   A warning, not a failure — RF is unaffected.

Exit code 0 = safe to train. 1 = do not trust anything downstream.

    python validate_dataset.py            # uses ML_settings.xlsx
    python validate_dataset.py --strict   # skew warnings also fail
"""

import os
import sys

import numpy as np
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))

def _find_settings(base, name="ML_settings.xlsx"):
    """Locate the settings workbook.

    Scripts live in ML/scripts/ but the workbook sits one level up in ML/ so
    the user sees it immediately. Checking the parent too means the layout can
    change (or a user can keep the workbook beside the scripts) without edits.
    """
    import os as _os
    here = _os.path.join(base, name)
    if _os.path.exists(here):
        return here
    up = _os.path.join(_os.path.dirname(base), name)
    return up if _os.path.exists(up) else here

SETTINGS_XLSX = _find_settings(HERE)

# Relative paths in the settings workbook resolve against the folder that HOLDS
# the workbook, NOT against this file. That way the scripts can be moved into a
# subfolder (ML/scripts/) without touching a single setting.
BASE_DIR = os.path.dirname(SETTINGS_XLSX)
import sys; sys.path.insert(0, HERE); import run_paths   # clean filing: data/ + runs/


# Pass/fail thresholds live in the 'validate' sheet of ML_settings.xlsx, not
# here. They are judgement calls, not physics — a different material, mesher or
# lattice family may justify different tolerances, and nobody should have to
# edit a .py to change one.
DEFAULTS = [
    ("tol_density", 0.05,
     "Max |VF/Rho_rel - 1| for a row to count as agreeing"),
    ("min_agree", 0.90,
     "Share of rows that must agree before the dataset passes"),
    ("tol_thickness", 0.05,
     "Max |measured wall / filename t - 1| for a row to count as agreeing"),
    ("max_skew", 3.0,
     "Feature skew above this is flagged (hurts GPR, not RF)"),
    ("min_triangles", 1000,
     "Meshes below this triangle count are flagged as suspiciously coarse"),
]

fails, warns = [], []


def load_thresholds():
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX)
    if "validate" not in wb.sheetnames:
        ws = wb.create_sheet("validate")
        ws.append(("setting", "value", "explanation"))
        for r in DEFAULTS:
            ws.append(r)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 70
        wb.save(SETTINGS_XLSX)
        print("[i] created 'validate' sheet in ML_settings.xlsx with defaults")
    cfg = {k: v for k, v, *_ in DEFAULTS}
    for k, v, *_ in wb["validate"].iter_rows(min_row=2, values_only=True):
        if k and v is not None:
            cfg[str(k).strip()] = float(v)
    return cfg


def load_dataset_path():
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    cfg = {str(k).strip(): v for k, v, *_ in
           wb["train"].iter_rows(min_row=2, values_only=True) if k}
    return run_paths.data_path(cfg["dataset_csv"])   # dataset lives in ML/data/


def check_density(d, T):
    print("\n1. DENSITY CROSS-CHECK  (STL surface vs FEA mesh — independent)")
    if "VF" not in d or "Rho_rel" not in d:
        warns.append("VF or Rho_rel missing — cross-check skipped")
        print("   [WARN] columns missing, skipped")
        return
    r = (d["VF"] / d["Rho_rel"]).replace([np.inf, -np.inf], np.nan).dropna()
    if r.empty:
        fails.append("no comparable VF/Rho_rel rows")
        print("   [FAIL] nothing to compare")
        return
    agree = float((abs(r - 1) < T["tol_density"]).mean())
    print(f"   mean ratio {r.mean():.4f}   median {r.median():.4f}   "
          f"range {r.min():.4f}-{r.max():.4f}")
    print(f"   within {T['tol_density']:.0%}: {int(agree * len(r))}/{len(r)}  ({agree:.1%})")
    if agree < T["min_agree"]:
        fails.append(f"only {agree:.0%} of rows agree on density "
                     f"(need {T['min_agree']:.0%}) — STL/FEA mismatch")
        print("   [FAIL] the two measurements disagree — do NOT train on this")
        worst = d.assign(_r=d["VF"] / d["Rho_rel"])
        worst["_off"] = (worst["_r"] - 1).abs()
        for _, w in worst.nlargest(5, "_off").iterrows():
            print(f"          {w.get('Run', '?')}  VF={w['VF']:.4f}  "
                  f"Rho={w['Rho_rel']:.4f}  ratio={w['_r']:.3f}")
    else:
        print("   [ OK ] independent measurements agree — STL extraction is sound")


def check_thickness(d, T):
    print("\n2. WALL THICKNESS  (measured vs filename t)")
    if "thickness_ratio" not in d:
        warns.append("thickness_ratio missing")
        print("   [WARN] column missing, skipped")
        return
    t = d["thickness_ratio"].dropna()
    if t.empty:
        warns.append("no thickness_ratio values")
        print("   [WARN] no values")
        return
    agree = float((abs(t - 1) < T["tol_thickness"]).mean())
    print(f"   mean {t.mean():.4f}   range {t.min():.4f}-{t.max():.4f}   "
          f"within {T['tol_thickness']:.0%}: {int(agree * len(t))}/{len(t)}")
    if agree < T["min_agree"]:
        fails.append(f"thickness mismatch ({agree:.0%} agree) — check "
                     f"thickness_convention (total vs double) in ML_settings.xlsx")
        print("   [FAIL] measured wall does not match the filename t")
    else:
        print("   [ OK ] thickness_convention setting is correct for this dataset")


def check_integrity(d, feats, T):
    print("\n3. DATA INTEGRITY")
    bad = False
    for f in feats:
        if f not in d:
            fails.append(f"feature column '{f}' missing from dataset")
            print(f"   [FAIL] {f}: MISSING"); bad = True; continue
        s = d[f]
        n_nan = int(s.isna().sum())
        if n_nan:
            fails.append(f"{f}: {n_nan} NaN")
            print(f"   [FAIL] {f}: {n_nan} NaN"); bad = True
    if "VF" in d:
        nvf = int(((d["VF"] <= 0) | (d["VF"] >= 1)).sum())
        if nvf:
            fails.append(f"{nvf} row(s) with impossible VF (<=0 or >=1)")
            print(f"   [FAIL] {nvf} impossible volume fraction(s)"); bad = True
    if "n_tris" in d:
        ntri = int((d["n_tris"] < T["min_triangles"]).sum())
        if ntri:
            warns.append(f"{ntri} suspiciously coarse mesh(es) (<{T['min_triangles']:.0f} triangles)")
            print(f"   [WARN] {ntri} mesh(es) under {T['min_triangles']:.0f} triangles")
    # Watertightness matters ONLY because mesh.volume (hence VF) is undefined on
    # an open mesh. But that is a PROXY: when FEA labels are present we can test
    # VF directly against Rho_rel, which is the thing we actually care about.
    #
    # Observed 2026-07-20: 3 of 116 ADMS meshes report non-watertight, yet their
    # VF/Rho_rel is 0.9965-0.9970 — better than the dataset median (0.9964), with
    # consistent winding and normal Euler numbers. A handful of unmerged edges in
    # 4-6M triangles. Failing the whole run on that would discard good data.
    #
    # So: non-watertight + VF agrees  -> warn (leak is cosmetic)
    #     non-watertight + VF disagrees or unavailable -> fail (VF untrustworthy)
    if "watertight" in d:
        leaky = d[pd.to_numeric(d["watertight"], errors="coerce") == 0]
        if not len(leaky):
            print(f"   [ OK ] all {len(d)} meshes watertight — VF is valid")
        else:
            can_check = "Rho_rel" in d and "VF" in d
            if can_check:
                ratio = (leaky["VF"] / leaky["Rho_rel"]).replace(
                    [np.inf, -np.inf], np.nan)
                agrees = (ratio - 1).abs() < T["tol_density"]
                good, bad_n = int(agrees.sum()), int((~agrees).sum())
                if good:
                    warns.append(f"{good} non-watertight mesh(es), but their VF "
                                 f"agrees with the FEA mesh — leak is cosmetic")
                    print(f"   [WARN] {good} non-watertight mesh(es), VF still "
                          f"agrees within {T['tol_density']:.0%} — accepted:")
                    for r, v in zip(leaky["Run"][agrees.values].head(5),
                                    ratio[agrees.values].head(5)):
                        print(f"          {r}  VF/Rho_rel={v:.4f}")
                if bad_n:
                    fails.append(f"{bad_n} non-watertight mesh(es) whose VF also "
                                 f"disagrees with the FEA mesh — VF invalid")
                    print(f"   [FAIL] {bad_n} non-watertight AND VF disagrees:")
                    for r, v in zip(leaky["Run"][~agrees.values].head(5),
                                    ratio[~agrees.values].head(5)):
                        print(f"          {r}  VF/Rho_rel={v:.4f}")
            else:
                fails.append(f"{len(leaky)} mesh(es) NOT watertight and no Rho_rel "
                             f"to cross-check against — VF cannot be trusted")
                print(f"   [FAIL] {len(leaky)} non-watertight, no FEA density to "
                      f"verify VF against:")
                for r in leaky["Run"].head(5):
                    print(f"          {r}")
    if "winding_ok" in d:
        badw = int((pd.to_numeric(d["winding_ok"], errors="coerce") == 0).sum())
        if badw:
            fails.append(f"{badw} mesh(es) with inconsistent winding — volume sign unreliable")
            print(f"   [FAIL] {badw} mesh(es) with inconsistent face winding")
    if not bad:
        print("   [ OK ] no NaNs, no impossible values")


def check_targets(d, es_gpa, sigma_ys_mpa):
    """Catch a wrong material / normalization constant.

    The density and thickness checks pass even when es_gpa/sigma_ys_mpa are wrong,
    because those constants only scale the TARGETS, not VF or thickness. A wrong
    material therefore produces plausible-looking-but-wrong labels with no error
    (labels_join's own docstring warns of exactly this). These bands are physical
    limits, so a normalized target outside them means the constant — or the join —
    is wrong. Generous on purpose: only gross errors trip, real lattices never do.
    """
    print("\n5. TARGET SANITY  (normalized labels within physical limits)")
    print(f"   normalization constants in use:  es_gpa={es_gpa}   sigma_ys_mpa={sigma_ys_mpa}")
    print("   (these are NOT validated by anything else — confirm they are YOUR material)")
    # MATERIAL-INDEPENDENT BY DESIGN. Every target here is dimensionless (already divided
    # by the solid Es or sigma_ys), so steel and polymer must fall in the SAME ranges — the
    # whole point of normalization. To avoid smuggling a material assumption back in:
    #   HARD  = a bound that is physically impossible for ANY material -> FAIL (blocks training)
    #   SOFT  = an upper bound that depends on Poisson ratio (e.g. C11/E rises with nu; a
    #           near-incompressible polymer legitimately exceeds a steel-calibrated cutoff)
    #           -> WARN only, so a collaborator's valid data is never falsely blocked.
    # (col, low, high, hard?, reason)
    bands = [
        ("E_over_Es",     0.0, 1.05, True,  "effective modulus cannot exceed the solid (E/Es<=1) — any material"),
        ("G_over_Gs",     0.0, 1.05, True,  "effective shear modulus cannot exceed the solid (G/Gs<=1) — any material"),
        ("nu_iso",       -1.00, 0.50, True, "isotropic Poisson ratio must lie in (-1, 0.5) — any material"),
        ("TAI",           0.0, None, True,  "TAI is a Frobenius norm ratio, cannot be negative"),
        ("C11_n",         0.0, None, True,  "stiffness cannot be negative"),
        ("C44_n",         0.0, None, True,  "stiffness cannot be negative"),
        ("onset_n",       0.0, None, True,  "yield onset cannot be negative"),
        ("shear_onset_n", 0.0, None, True,  "shear onset cannot be negative"),
        # soft upper bounds (Poisson-dependent / unusual) — warn, do not block
        ("C11_n",         None, 4.00, False, "C11/Es unusually high (dense + high-Poisson, or wrong es_gpa) — inspect"),
        ("C44_n",         None, 2.00, False, "C44/Es unusually high (or wrong es_gpa) — inspect"),
        ("C12_n",        -0.10, 3.00, False, "C12/Es out of usual range — inspect"),
        ("TAI",           None, 2.00, False, "TAI unusually large (>2) — inspect"),
        ("onset_n",       None, 3.00, False, "onset/sigma_ys unusually high (or wrong sigma_ys_mpa) — inspect"),
        ("shear_onset_n", None, 3.00, False, "shear onset/sigma_ys unusually high — inspect"),
    ]
    checked = set()
    for col, lo, hi, hard, why in bands:
        if col not in d.columns:
            continue
        s = pd.to_numeric(d[col], errors="coerce").dropna()
        if s.empty:
            continue
        checked.add(col)
        mask = pd.Series(False, index=s.index)
        if lo is not None:
            mask |= s < lo
        if hi is not None:
            mask |= s > hi
        n = int(mask.sum())
        rng = (f"[{lo if lo is not None else '-inf'}, {hi if hi is not None else '+inf'}]")
        if n:
            msg = f"{col}: {n} value(s) outside {rng} — {why}"
            (fails if hard else warns).append(msg)
            tag = "FAIL" if hard else "WARN"
            print(f"   [{tag}] {col:14} {n} outside {rng}  (min {s.min():.3f}, max {s.max():.3f}) — {why}")
    for col in sorted(checked):
        s = pd.to_numeric(d[col], errors="coerce").dropna()
        print(f"   [info] {col:14} observed range {s.min():.3f} .. {s.max():.3f}")
    if not checked:
        warns.append("no normalized target columns present to sanity-check")
        print("   [WARN] no target columns found to check")


def check_distributions(d, feats, strict, T):
    print(f"\n4. FEATURE DISTRIBUTIONS  (skew > {T['max_skew']} hurts GPR, not RF)")
    skewed = []
    for f in feats:
        if f not in d:
            continue
        s = d[f].dropna()
        if len(s) < 8:
            continue
        sk = float(s.skew())
        if abs(sk) > T["max_skew"]:
            ls = float(np.log10(s.abs() + 1e-12).skew())
            skewed.append((f, sk, ls))
            print(f"   [WARN] {f:24} skew={sk:6.2f}   (log10 would give {ls:6.2f})")
    if skewed:
        msg = f"{len(skewed)} feature(s) heavily skewed — consider log transform"
        (fails if strict else warns).append(msg)
    else:
        print("   [ OK ] no heavy skew")


def main():
    strict = "--strict" in sys.argv
    path = load_dataset_path()
    print("=" * 68)
    print("  DATASET VALIDATION — run BEFORE trusting any model result")
    print(f"  {path}")
    print("=" * 68)
    if not os.path.exists(path):
        print(f"[FAIL] dataset not found: {path}")
        return 1
    d = pd.read_csv(path)
    print(f"\nrows: {len(d)}   columns: {len(d.columns)}")
    if "adms_type" in d:
        print("by variant: " + ", ".join(f"{k}={v}" for k, v in
                                         d["adms_type"].value_counts().items()))

    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    cfg = {str(k).strip(): v for k, v, *_ in
           wb["train"].iter_rows(min_row=2, values_only=True) if k}
    feats = [s.strip() for s in str(cfg["features"]).split(",") if s.strip()]
    # material constants used to normalize the labels (labels sheet) — echoed and
    # sanity-checked by check_targets so a wrong material cannot pass silently.
    lab = {str(k).strip(): v for k, v, *_ in
           wb["labels"].iter_rows(min_row=2, values_only=True) if k} if "labels" in wb.sheetnames else {}
    es_gpa = lab.get("es_gpa", "?")
    sigma_ys_mpa = lab.get("sigma_ys_mpa", "?")

    T = load_thresholds()
    check_density(d, T)
    check_thickness(d, T)
    check_integrity(d, feats, T)
    check_targets(d, es_gpa, sigma_ys_mpa)
    check_distributions(d, feats, strict, T)

    print("\n" + "=" * 68)
    if fails:
        print(f"  VALIDATION FAILED — {len(fails)} problem(s). "
              f"Model results would be meaningless:")
        for f in fails:
            print(f"    - {f}")
    else:
        print("  VALIDATION PASSED — features describe the real geometry.")
    if warns:
        print(f"\n  {len(warns)} warning(s):")
        for w in warns:
            print(f"    - {w}")
    print("=" * 68)
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(main())
