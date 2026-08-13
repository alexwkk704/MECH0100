"""
labels_join.py — join STL geometric features with FEA labels from the master.

Pipeline position: feature_extraction.py -> features_<tag>.csv -> [THIS] -> dataset_<tag>.csv
Run AFTER the full feature extraction finishes. Safe to re-run (rebuilds output from scratch).

FULLY SPREADSHEET-DRIVEN — nothing here is specific to steel, to ADMS, or to any unit.
A new user (different material, topology, naming, or unit system) edits ML_settings.xlsx
ONLY and never touches this file. The three things that used to be hardcoded are now settings:

  * HOW to match a geometry to its label   -> 'labels' sheet: join_on = params | file
  * WHICH label columns map to which target -> 'label_map' sheet (source col | target | norm | unit)
  * WHAT unit the labels are in             -> the 'unit' column of 'label_map' (GPa/MPa/kPa/Pa/-)

What it does
  1. Load features CSV (from feature_extraction.py).
  2. Load the label source (Results_summary.xlsx 'summary' sheet, OR a .csv); keep the
     LATEST Version per run when a Version column exists.
  3. Join features <-> labels:
       join_on=params  ADMS stem parser (exact stem, fallback = normalised 5-param key)
       join_on=file    exact match of the STL stem to the label source's run column
                       (works for ANY naming — no parser, no ADMS assumptions)
  4. (params only) Dedupe seed-twins: seed variants are byte-identical geometry ->
     keep ONE row per unique geometry key; preferred seed order from settings.
  5. Normalise targets per the 'label_map' sheet: each value is converted from its stated
     unit to the divisor's unit, then divided by Es (stiffness) or sigma_ys (strength), or
     passed through (already-dimensionless columns like E_over_Es, nu_iso, TAI). Es and
     sigma_ys come from the 'labels' sheet — set them to YOUR material.
  6. Cross-check ratio columns, defined in the 'crosscheck' sheet (our column /
     reference column / ratio column). Nothing is named in this file.
  7. Write dataset_<tag>.csv + print join/dedupe report.

Usage:  python labels_join.py            (uses ML_settings.xlsx)
"""

import csv, os, re, sys
from collections import defaultdict

HERE = os.path.dirname(os.path.abspath(__file__))

def _find_settings(base, name="ML_settings.xlsx"):
    """Locate the settings workbook (beside the scripts, or one level up in ML/)."""
    import os as _os
    here = _os.path.join(base, name)
    if _os.path.exists(here):
        return here
    up = _os.path.join(_os.path.dirname(base), name)
    return up if _os.path.exists(up) else here

SETTINGS_XLSX = _find_settings(HERE)
BASE_DIR = os.path.dirname(SETTINGS_XLSX)
import sys; sys.path.insert(0, HERE); import run_paths   # clean filing: data/ + runs/


DEFAULTS = [
    ("features_csv", "outputs/features_ADMS.csv",
     "Input: feature table from feature_extraction.py"),
    ("master_xlsx", "../ADMS/Results/Results_summary.xlsx",
     "Input: label source. An .xlsx (uses master_sheet) OR a .csv. Column names are "
     "declared in the 'label_map' sheet, so any schema works."),
    ("master_sheet", "summary",
     "Sheet name inside master_xlsx that holds the labels (ignored for a .csv source)"),
    ("output_csv", "outputs/dataset_ADMS.csv",
     "Output: joined features + labels, ready for train.py"),
    ("join_on", "params",
     "How to match a geometry to its label. 'params' = ADMS stem parser (i/d/t/s/m tokens, "
     "with seed-twin dedup). 'file' = exact match of the STL stem to the run column below "
     "(topology-agnostic; use this for TPMS or any non-ADMS naming)."),
    ("master_run_col", "Run",
     "The label-source column whose value equals the STL file stem (used by join_on=file, "
     "and as the run id everywhere). ADMS master = 'Run'."),
    ("variant_col", "",
     "join_on=file only: label-source column to copy into 'adms_type' (a grouping label). "
     "Blank -> use variant_default."),
    ("variant_default", "",
     "join_on=file only: constant written into 'adms_type' when variant_col is blank/absent "
     "(e.g. a topology tag like TPMS). Blank -> left empty."),
    ("es_gpa", 200.0,
     "Solid Young's modulus in GPa; divides the stiffness targets (norm=Es). "
     "Steel=200, Ti-6Al-4V=113.8, polymer~1.8. NOTE E_over_Es/G_over_Gs may arrive already "
     "normalised (norm=none) — then this does not touch them."),
    ("sigma_ys_mpa", 250.0,
     "Solid yield strength in MPa; divides the strength targets (norm=sigma_ys)"),
    ("seed_pref", "2,3,5,1",
     "join_on=params only: seed priority when deduping identical seed-twin geometries"),
]

# Default label map = the ADMS/nTop master schema. Reproduces the historic normalisation
# exactly (C/es_gpa in GPa, onset/sigma_ys in MPa, the rest already dimensionless).
# Columns: source_column | output_target | normalize_by | source_unit
#   normalize_by : none | Es | sigma_ys
#   source_unit  : GPa | MPa | kPa | Pa | -   (converted to the divisor's unit before dividing)
LABEL_MAP_DEFAULTS = [
    ("source_column", "output_target", "normalize_by", "source_unit"),
    ("E_over_Es",       "E_over_Es",     "none",     "-"),
    ("G_over_Gs",       "G_over_Gs",     "none",     "-"),
    ("nu_iso",          "nu_iso",        "none",     "-"),
    ("TAI",             "TAI",           "none",     "-"),
    ("Zener_Z",         "Zener_Z",       "none",     "-"),
    ("C11_GPa",         "C11_n",         "Es",       "GPa"),
    ("C12_GPa",         "C12_n",         "Es",       "GPa"),
    ("C44_GPa",         "C44_n",         "Es",       "GPa"),
    ("Yield_onset_MPa", "onset_n",       "sigma_ys", "MPa"),
    ("Shear_onset_MPa", "shear_onset_n", "sigma_ys", "MPa"),
    ("Rho_rel",         "Rho_rel",       "none",     "-"),
]

# unit -> multiplier that converts INTO the divisor's unit
_TO_GPA = {"gpa": 1.0, "mpa": 1e-3, "kpa": 1e-6, "pa": 1e-9, "-": 1.0, "": 1.0, "none": 1.0}
_TO_MPA = {"mpa": 1.0, "gpa": 1e3, "kpa": 1e-3, "pa": 1e-6, "-": 1.0, "": 1.0, "none": 1.0}


def load_settings():
    """Read the 'labels' sheet, creating it with defaults if absent."""
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX)
    changed = False
    if "labels" not in wb.sheetnames:
        ws = wb.create_sheet("labels")
        ws.append(("setting", "value", "explanation"))
        for r in DEFAULTS:
            ws.append(r)
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 80
        changed = True
        print("[i] created 'labels' sheet in ML_settings.xlsx with defaults")
    else:
        # add any missing keys (so an OLD workbook gains the new join settings, no manual edit)
        ws = wb["labels"]
        have = {str(c[0].value).strip() for c in ws.iter_rows(min_row=2) if c[0].value}
        for k, v, expl in DEFAULTS:
            if k not in have:
                ws.append((k, v, expl)); changed = True
                print(f"[i] added missing 'labels' setting '{k}' (default: {v!r})")
    if changed:
        wb.save(SETTINGS_XLSX)
    cfg = {k: v for k, v, *_ in DEFAULTS}
    for k, v, *_ in wb["labels"].iter_rows(min_row=2, values_only=True):
        if k and v is not None:
            cfg[str(k).strip()] = v
    # master is an external INPUT -> resolve against the ML folder
    p = str(cfg["master_xlsx"])
    cfg["master_xlsx"] = os.path.normpath(p if os.path.isabs(p) else os.path.join(BASE_DIR, p))
    # features (in) and dataset (out) are DERIVED INPUTS -> ML/data/
    cfg["features_csv"] = run_paths.data_path(cfg["features_csv"])
    cfg["output_csv"] = run_paths.data_path(cfg["output_csv"])
    for k in ("es_gpa", "sigma_ys_mpa"):
        cfg[k] = float(cfg[k])
    cfg["seed_pref"] = [int(s) for s in str(cfg["seed_pref"]).split(",") if str(s).strip()]
    cfg["join_on"] = str(cfg["join_on"]).strip().lower()
    for k in ("master_sheet", "master_run_col", "variant_col", "variant_default"):
        cfg[k] = str(cfg[k]).strip() if cfg[k] is not None else ""
    return cfg


def load_label_map():
    """Read the 'label_map' sheet, creating it with the ADMS default if absent.

    Returns a list of (source_column, output_target, normalize_by, source_unit).
    """
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX)
    if "label_map" not in wb.sheetnames:
        ws = wb.create_sheet("label_map")
        for r in LABEL_MAP_DEFAULTS:
            ws.append(r)
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 12
        wb.save(SETTINGS_XLSX)
        print("[i] created 'label_map' sheet in ML_settings.xlsx with the ADMS default map")
    rows = []
    for r in wb["label_map"].iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None:
            continue
        src = str(r[0]).strip()
        out = str(r[1]).strip() if len(r) > 1 and r[1] is not None else src
        nby = (str(r[2]).strip().lower() if len(r) > 2 and r[2] is not None else "none")
        unit = (str(r[3]).strip().lower() if len(r) > 3 and r[3] is not None else "-")
        # role: 'target' (a thing the model predicts) or 'crosscheck' (a reference
        # value carried into the dataset only so it can be compared). Data, not code.
        role = (str(r[4]).strip().lower() if len(r) > 4 and r[4] is not None else "target")
        rows.append((src, out, nby, unit, role))
    return rows


def load_crosschecks():
    """Rows of the 'crosscheck' sheet: our_column / reference_column / ratio_column.
    Every cross-check ratio the dataset carries is defined THERE, never in this file."""
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX)
    if "crosscheck" not in wb.sheetnames:
        return []
    out = []
    for r in wb["crosscheck"].iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None or r[1] is None or r[2] is None:
            continue
        lo = float(r[3]) if len(r) > 3 and r[3] is not None else 0.9
        hi = float(r[4]) if len(r) > 4 and r[4] is not None else 1.1
        out.append((str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip(), lo, hi))
    return out


def _normalize(value, norm_by, unit, es_gpa, sigma_ys_mpa):
    """Convert a raw label into its dimensionless / material-normalised form."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    nb = norm_by.lower()
    if nb in ("es", "es_gpa", "e_s", "modulus"):
        return v * _TO_GPA.get(unit, 1.0) / es_gpa
    if nb in ("sigma_ys", "sigma", "sys", "strength", "yield"):
        return v * _TO_MPA.get(unit, 1.0) / sigma_ys_mpa
    return v  # none -> passthrough (already dimensionless)


def _read_table(path, sheet):
    """Read the label source into (header_list, list_of_row_tuples). CSV or XLSX."""
    if path.lower().endswith(".csv"):
        with open(path, newline="") as f:
            rows = list(csv.reader(f))
        if not rows:
            return [], []
        return list(rows[0]), [tuple(r) for r in rows[1:]]
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    return list(rows[0]), rows[1:]


def load_master(cfg, label_map):
    """Load + normalise labels, keyed by the run column. Latest Version per run kept."""
    hdr, rows = _read_table(cfg["master_xlsx"], cfg["master_sheet"])
    idx = {h: i for i, h in enumerate(hdr)}
    run_col = cfg["master_run_col"]
    if run_col not in idx:
        sys.exit(f"[E] master_run_col '{run_col}' not found in the label source. "
                 f"Available columns: {list(idx)}. Set master_run_col in the 'labels' sheet.")
    es, sy = cfg["es_gpa"], cfg["sigma_ys_mpa"]
    variant_col = cfg["variant_col"]
    have_ver = "Version" in idx
    best = {}
    for r in rows:
        run = r[idx[run_col]] if idx[run_col] < len(r) else None
        if run in (None, ""):
            continue
        run = str(run).strip()
        run = re.sub(r"\.stl$", "", run, flags=re.I)   # tolerate names stored with .stl
        ver = 0
        if have_ver and idx["Version"] < len(r):
            try:
                ver = int(re.sub(r"\D", "", str(r[idx["Version"]])) or 0)
            except Exception:
                ver = 0
        if run not in best or ver >= best[run][0]:
            best[run] = (ver, r)
    label = {}
    for run, (_v, r) in best.items():
        row = {}
        for src, out, nby, unit, role in label_map:
            v = r[idx[src]] if src in idx and idx[src] < len(r) else None
            row[out] = _normalize(v, nby, unit, es, sy) if v is not None else None
        row["Status"] = r[idx["Status"]] if "Status" in idx and idx["Status"] < len(r) else "OK"
        row["Version"] = _v
        if variant_col and variant_col in idx and idx[variant_col] < len(r):
            row["_variant"] = r[idx[variant_col]]
        label[run] = row
    return label


# ---------------- ADMS stem parser (join_on=params only) ----------------
_num = r"(\d+p?\d*)"
def _f(tok):
    return float(tok.replace("p", ".")) if tok else None

def parse_stem(stem):
    """ADMS naming: dict(adms_type, inner, density, t, seed, size_multi) from a stem."""
    s = stem
    d = {}
    m = re.search(r"ADMS[_ ]?(DF|raw|flow)", s, re.I)
    d["adms_type"] = m.group(1).lower() if m else "df"
    m = re.search(r"_i" + _num, s);  d["inner"]      = _f(m.group(1)) if m else None
    m = re.search(r"_d" + _num, s);  d["density"]    = _f(m.group(1)) if m else None
    m = re.search(r"_t" + _num, s);  d["t"]          = _f(m.group(1)) if m else None
    m = re.search(r"_s(\d+)\b", s);  d["seed"]       = int(m.group(1)) if m else None
    m = re.search(r"_m" + _num, s);  d["size_multi"] = _f(m.group(1)) if m else None
    if "5cube" in s:
        d["inner"]      = d["inner"] if d["inner"] is not None else 15.0
        d["size_multi"] = d["size_multi"] if d["size_multi"] is not None else 1.6
        d["t"]          = d["t"] if d["t"] is not None else 0.4
        d["seed"]       = d["seed"] if d["seed"] is not None else 2
    return d

def key5(p):
    return (p["adms_type"], p["inner"], p["density"], p["t"], p["seed"], p["size_multi"])

def geo_key(p):
    return (p["adms_type"], p["inner"], p["density"], p["t"], p["size_multi"])


def main():
    cfg = load_settings()
    label_map = load_label_map()
    FEATURES_CSV = cfg["features_csv"]
    OUT_CSV = cfg["output_csv"]
    join_on = cfg["join_on"]

    if not os.path.exists(FEATURES_CSV):
        sys.exit(f"[E] features CSV not found: {FEATURES_CSV}")
    if not os.path.exists(cfg["master_xlsx"]):
        sys.exit(f"[E] label source not found: {cfg['master_xlsx']}")
    with open(FEATURES_CSV, newline="") as f:
        feats = [r for r in csv.DictReader(f)]
    feats = [r for r in feats if r.get("status", "OK").startswith("OK")]
    print(f"[i] features rows OK: {len(feats)}")
    print(f"[i] join_on = {join_on}   label source = {os.path.basename(cfg['master_xlsx'])}"
          f"   targets mapped = {len(label_map)}")

    labels = load_master(cfg, label_map)
    print(f"[i] label rows (latest version): {len(labels)}")

    if join_on == "params":
        lab_bykey = {key5(parse_stem(run)): run for run in labels}

    # join
    joined, miss = [], []
    for r in feats:
        stem = re.sub(r"\.stl$", "", r["file"], flags=re.I)
        if join_on == "file":
            run = stem if stem in labels else None
        else:  # params
            run = stem if stem in labels else lab_bykey.get(key5(parse_stem(stem)))
        if run is None:
            miss.append(stem); continue
        row = dict(r)
        row["Run"] = run
        if join_on == "file":
            v = labels[run].get("_variant")
            row["adms_type"] = (str(v).strip() if v not in (None, "")
                                else cfg["variant_default"])
        else:
            row["adms_type"] = parse_stem(run)["adms_type"]
        lab = {k: v for k, v in labels[run].items() if k != "_variant"}
        row.update(lab)
        joined.append(row)
    print(f"[i] joined: {len(joined)}   unmatched: {len(miss)}")
    for s in miss:
        print(f"    [!] no label: {s}")
    # A large unmatched fraction means a systematic naming/key drift (wrong join_on, wrong
    # master_run_col, or a different dataset). Fail loudly so the .bat stops here instead of
    # training on an accidentally-truncated dataset.
    if feats and len(miss) / len(feats) > 0.20:
        hint = ("check master_run_col matches the STL stems"
                if join_on == "file" else
                "the stems do not parse as ADMS names — set join_on=file in the 'labels' sheet")
        sys.exit(f"[E] {len(miss)}/{len(feats)} feature rows ({len(miss)/len(feats):.0%}) got "
                 f"NO label — too many to be pending sims. Usually the feature stems and the "
                 f"label source disagree ({hint}). Nothing was written.")

    # drop rows the FEA could NOT label (e.g. Status=FAIL — homogenization failed
    # because the geometry disconnected, so every target is None). They carry no
    # training signal and just inflate the count with all-NaN lines. They stay
    # recorded in the master; they are simply excluded from the ML dataset.
    # NOTE: cross-check columns (nTop curvature) are mapped through label_map so they
    # reach the dataset, but they are NOT training targets. A geometry whose FEA failed
    # still has curvature, so counting them here would keep an all-NaN row alive.
    target_names = [out for (_s, out, _n, _u, role) in label_map if role == "target"]
    def _labelless(row):
        return all(row.get(t) in (None, "") for t in target_names)
    failed = [r for r in joined if _labelless(r)]
    joined = [r for r in joined if not _labelless(r)]
    if failed:
        print(f"[i] dropped {len(failed)} row(s) with NO valid labels "
              f"(FEA FAIL / homogenization failed — kept in the master, not the dataset):")
        for r in failed:
            print(f"    FAIL: {r['Run']}  (status: {r.get('Status')})")

    # dedupe seed-twins — ADMS-specific, params mode only
    if join_on == "params":
        pref = {s: i for i, s in enumerate(cfg["seed_pref"])}
        groups = defaultdict(list)
        for row in joined:
            groups[geo_key(parse_stem(row["Run"]))].append(row)
        kept, dropped = [], 0
        for k, rows in groups.items():
            rows.sort(key=lambda r: pref.get(parse_stem(r["Run"]).get("seed"), 99))
            kept.append(rows[0]); dropped += len(rows) - 1
        print(f"[i] unique geometries kept: {len(kept)}   seed-twins dropped: {dropped}")
    else:
        kept = joined
        print(f"[i] rows kept: {len(kept)}  (no seed-twin dedup in file mode)")

    # ---- cross-checks, ALL defined in the 'crosscheck' sheet ------------------
    # our_column / reference_column / ratio_column. Adding a new check is a row in
    # the workbook, not an edit here.
    checks = load_crosschecks()
    for our, ref, ratio, lo, hi in checks:
        n_ok = 0
        for row in kept:
            try:
                row[ratio] = round(float(row[our]) / float(row[ref]), 4)
                n_ok += 1
            except (TypeError, ValueError, KeyError, ZeroDivisionError):
                row[ratio] = ""
        vals = [row[ratio] for row in kept if row.get(ratio) not in (None, "")]
        if not vals:
            print(f"[!] cross-check '{ratio}': no rows - is '{our}' or '{ref}' missing?")
            continue
        outside = [v for v in vals if not lo < v < hi]
        mean = sum(vals) / len(vals)
        flag = "" if not outside else f"   [!] {len(outside)} outside {lo}-{hi}"
        print(f"[i] cross-check {ratio:<22} n={n_ok:<4} mean={mean:.4f}{flag}")

    if not kept:
        sys.exit("[E] no rows survived the join — 0 geometries matched a label. Check that the "
                 "label source and the STL/feature naming are the same dataset. Nothing written.")
    kept.sort(key=lambda r: float(r["VF"]) if r.get("VF") not in (None, "") else 0.0)
    # union of keys so a missing optional column never drops a field
    fieldnames = list(dict.fromkeys(k for r in kept for k in r.keys()))
    OUT_DIR = os.path.dirname(OUT_CSV)
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader(); w.writerows(kept)
    n_onset = sum(1 for r in kept if r.get("onset_n"))
    n_shear = sum(1 for r in kept if r.get("shear_onset_n"))
    print(f"[OK] wrote {OUT_CSV}: {len(kept)} rows, {len(fieldnames)} cols "
          f"(onset labels: {n_onset}, shear labels: {n_shear})")

if __name__ == "__main__":
    main()
