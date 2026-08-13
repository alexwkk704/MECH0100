"""
train_pointnet.py — Farooq pipeline steps 3-5: PointNet regression on point clouds.

Step 3  Architecture : PointNet (shared MLP -> max-pool global feature),
                       classification head STRIPPED, replaced with a linear
                       Regression Head that outputs continuous property values.
Step 4  Training      : augmentation = coordinate jitter (+ OPTIONAL rotation,
                       OFF by default — see warning), MSE loss, Adam optimiser.
Step 5  Validation    : 80/20 split; train on 80%, predict the hidden 20%,
                       compare predictions vs FEA (R2 + MAPE); parity PNG.

Reads settings from ML_settings.xlsx sheet "pointnet" (auto-created with
defaults on first run). Point clouds come from pointcloud_prep.py (.npz),
labels from the same dataset CSV used by the feature-based models.

  ROTATION AUGMENTATION (Farooq's full-tensor method, rotation_aug=1)
    Plain rotation with FROZEN labels is wrong: C11/C12/C44/TAI are direction-
    dependent, so rotating the shape changes them. The correct method rotates the
    cloud AND the stiffness tensor together, then recomputes the direction-dependent
    labels from the rotated tensor (tensor_ops.py). Requires the full 6x6 per
    geometry -> run collect_tensors.py first (data/tensors_ADMS.npz).
      * frame_dependent targets (C11_n/C12_n/C44_n/TAI) -> recomputed per rotation
      * rotation_invariant targets (E/Es, G/Gs, nu_iso) -> kept (VRH invariants)
      * anything else (onset_n, shear_onset_n — plastic, not tensor-derivable)
        -> MASKED on rotated copies, so they are never taught false physics.
    Which target is which is settings-driven (frame_dependent / rotation_invariant).
    rotation_aug defaults to FALSE; enable it only for a measured A/B vs the baseline.

Requires: pip install torch numpy pandas openpyxl matplotlib
Usage:    python train_pointnet.py
"""

import os, sys, glob, re
import numpy as np
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)          # so `import cv_split` works when run from ML/
import cv_split                   # shared, deterministic grouped CV folds
import run_paths                  # clean filing: data/ inputs, runs/<stamp>/ results
import tensor_ops                 # rotation math for Farooq's full-tensor augmentation

def _find_settings(base, name="ML_settings.xlsx"):
    """Locate the settings workbook.

    Scripts live in ML/scripts/ but the workbook sits one level up in ML/ so
    the user sees it immediately. Checking the parent too means the layout can
    change (or a user can keep the workbook beside the scripts) without edits.
    """
    import os as _os
    here = _os.path.join(base, name)
    if _os.path.exists(here):
        return here
    up = _os.path.join(_os.path.dirname(base), name)
    return up if _os.path.exists(up) else here

SETTINGS_XLSX = _find_settings(HERE)

# Relative paths in the settings workbook resolve against the folder that HOLDS
# the workbook, NOT against this file. That way the scripts can be moved into a
# subfolder (ML/scripts/) without touching a single setting.
BASE_DIR = os.path.dirname(SETTINGS_XLSX)


PN_DEFAULTS = [
    ("pc_folder", os.path.join(HERE, "pointclouds"), "Folder of <stem>.npz point clouds"),
    ("dataset_csv", os.path.join(HERE, "dataset_ADMS.csv"), "Joined features+labels CSV (for targets + Run stems)"),
    ("charts_dir", os.path.join(HERE, "_charts"), "Where parity PNGs are written"),
    ("targets", "E_over_Es, G_over_Gs, onset_n", "Property columns to regress (comma-separated)"),
    ("n_points", 2048, "Points per cloud (must match pointcloud_prep)"),
    ("epochs", 300, "Training epochs"),
    ("batch_size", 16, "Mini-batch size"),
    ("lr", 0.001, "Adam learning rate"),
    ("test_frac", 0.2, "Hidden validation fraction (step 5)"),
    ("jitter_std", 0.01, "Gaussian coord jitter std (unit-box scale)"),
    ("rotation_aug", 0,
     "1 = Farooq full-tensor rotation aug (rotate cloud + tensor, recompute "
     "direction-dependent labels). Needs collect_tensors.py first. 0 = off."),
    ("n_rot", 8,
     "rotation_aug only: number of extra rotated copies per TRAIN geometry "
     "(train-only; val/test are never augmented)."),
    ("frame_dependent", "C11_n, C12_n, C44_n",
     "rotation_aug only: GENUINELY direction-dependent targets, RECOMPUTED from the "
     "rotated tensor. NOTE: TAI is rotation-INVARIANT (proven — 0.0%% drift in Mandel "
     "notation vs 28%% in Voigt), so it belongs in rotation_invariant, NOT here. "
     "Recomputing TAI in Voigt notation injected artifact noise and collapsed it."),
    ("rotation_invariant", "E_over_Es, G_over_Gs, nu_iso, TAI",
     "rotation_aug only: targets UNCHANGED by rotation (VRH invariants E/G/nu + the "
     "anisotropy index TAI) -> kept as-is on rotated copies. Any target in NEITHER "
     "list is masked on rotated copies (e.g. plastic onset, not tensor-derivable)."),
    ("random_seed", 42, "Reproducibility"),
    # ---- merged from Zhezhe's train_pointnet_v3.py (2026-07-20) ----
    ("use_density_channel", 1,
     "1 = append relative density as a 4th per-point channel (xyz -> xyzd). "
     "Clouds are scaled to a unit box, which DELETES absolute scale; density "
     "is the strongest single predictor (r=0.967 with stiffness) and the net "
     "cannot otherwise see it. Zhezhe's idea. 0 = xyz only (original)."),
    ("val_frac", 0.15,
     "Fraction held out for model SELECTION. Split is train/val/test: val picks "
     "the best epoch, test is never seen. Selecting on test would leak."),
    ("lr_decay", 0.5, "Multiply lr by this every lr_step epochs. 1.0 = constant"),
    ("lr_step", 200, "Epochs between lr decays"),
    ("log_targets", 1,
     "1 = train on log10(target). Targets span ~129x; plain MSE minimises "
     "ABSOLUTE error so small values are proportionally very wrong (MAPE ~102%). "
     "Log makes MSE approximate RELATIVE error. 0 = original behaviour."),
    ("deep_head", 1,
     "1 = 1024-512-256-out regression head with BatchNorm (Zhezhe). "
     "0 = 1024-256-64-out (original)."),
    ("grad_clip", 1.0, "Max gradient norm; 0 = off. Prevents the dead-ReLU collapse"),
    # ---- grouped cross-validation (2026-07-20, replaces single random split) ----
    ("n_folds", 5,
     "Cross-validation folds. Every geometry is predicted once by a model that "
     "never trained on it (out-of-fold). Set 1 to fall back to a single split."),
    ("stratify_col", "",
     "pointnet_split_8020 only: column defining one geometry type for the "
     "stratified 80/20 split. MUST be a column that survives whatever built the "
     "dataset - adms_type is ADMS-only and merge_datasets.py drops it, so a "
     "COMBINED dataset needs the unified 'topology'. Blank = auto-pick the first "
     "of adms_type / topology / family that is present."),
    ("group_cols", "adms_type, density_param",
     "Columns whose unique combos define a CV GROUP. Near-identical geometries "
     "(same group) are kept in the SAME fold so they cannot leak across "
     "train/test. Must match train.py. Blank/absent -> per-row split (no "
     "leakage protection). Shared with collaborators: set to their own columns."),
]


def load_pn_cfg():
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX)
    if run_paths.sheet_for(wb, "pointnet") not in wb.sheetnames:
        ws = wb.create_sheet("pointnet")
        ws.append(("setting", "value", "explanation"))
        for r in PN_DEFAULTS:
            ws.append(r)
        wb.save(SETTINGS_XLSX)
        print("[i] created 'pointnet' sheet in ML_settings.xlsx with defaults")
    ws = wb[run_paths.sheet_for(wb, "pointnet")]
    have = {str(r[0]).strip() for r in ws.iter_rows(min_row=2, values_only=True) if r[0]}
    # Backfill any setting the sheet predates. Without this, a sheet created by
    # an older version silently returns None for every new key and the feature
    # is skipped without warning. (Caught 2026-07-20 by testing, not reading.)
    added = [(k, v, note) for k, v, note in PN_DEFAULTS if k not in have]
    if added:
        for row in added:
            ws.append(row)
        wb.save(SETTINGS_XLSX)
        print(f"[i] added {len(added)} new setting(s) to the 'pointnet' sheet: "
              + ", ".join(k for k, _, _ in added))
    cfg = {}
    for k, v, *_ in ws.iter_rows(min_row=2, values_only=True):
        if k:
            cfg[str(k).strip()] = v
    cfg["targets"] = [s.strip() for s in str(cfg["targets"]).split(",") if s.strip()]
    cfg["group_cols"] = [s.strip() for s in str(cfg.get("group_cols", "")).split(",") if s.strip()]
    cfg["frame_dependent"] = [s.strip() for s in str(cfg.get("frame_dependent", "")).split(",") if s.strip()]
    cfg["rotation_invariant"] = [s.strip() for s in str(cfg.get("rotation_invariant", "")).split(",") if s.strip()]
    for k in ("n_points", "epochs", "batch_size", "random_seed", "n_folds"):
        cfg[k] = int(cfg[k])
    cfg["n_rot"] = int(cfg.get("n_rot", 8) or 0)
    for k in ("lr", "test_frac", "jitter_std"):
        cfg[k] = float(cfg[k])
    cfg["rotation_aug"] = bool(int(cfg["rotation_aug"]))
    # es_gpa (to normalize recomputed stiffness entries) comes from the 'labels'
    # sheet — the same material constant labels_join used. No hardcoding.
    es = 200.0
    if "labels" in wb.sheetnames:
        lab = {str(k).strip(): v for k, v, *_ in wb["labels"].iter_rows(min_row=2, values_only=True) if k}
        try:
            es = float(lab.get("es_gpa", 200.0))
        except (TypeError, ValueError):
            es = 200.0
    # The pointnet sheet overrides it. collect_tensors.py caches C/Es already
    # normalised, so no further division is needed (es_gpa = 1.0). That is what
    # lets one combined dataset hold steel ADMS and polymer TPMS rows together.
    try:
        if cfg.get("es_gpa") not in (None, ""):
            es = float(cfg["es_gpa"])
    except (TypeError, ValueError):
        pass
    cfg["es_gpa"] = es
    # clean filing: clouds + dataset are shared derived INPUTS (ML/data/);
    # charts are per-run OUTPUTS (ML/runs/<stamp>/charts).
    cfg["pc_folder"] = run_paths.pointclouds_dir()
    cfg["dataset_csv"] = run_paths.data_path(cfg["dataset_csv"])
    cfg["charts_dir"] = run_paths.run_path("charts")
    return cfg


# --------------------------------------------------------------------------
def build_xy(cfg):
    """Match .npz point clouds to dataset rows by Run stem; return clouds + labels."""
    df = pd.read_csv(cfg["dataset_csv"])
    stem_col = "Run" if "Run" in df.columns else "file"
    npz = {os.path.splitext(os.path.basename(p))[0]: p
           for p in glob.glob(os.path.join(cfg["pc_folder"], "*.npz"))}
    X, rows = [], []
    for _, r in df.iterrows():
        stem = re.sub(r"\.stl$", "", str(r[stem_col]), flags=re.I)
        path = npz.get(stem)
        if path is None:  # try loose contains-match (seed twins / naming)
            hit = [v for k, v in npz.items() if stem in k or k in stem]
            path = hit[0] if hit else None
        if path is None:
            continue
        pts = np.load(path)["pts"].astype(np.float32)
        if len(pts) >= cfg["n_points"]:
            pts = pts[:cfg["n_points"]]
        else:
            idx = np.random.choice(len(pts), cfg["n_points"], replace=True)
            pts = pts[idx]
        X.append(pts); rows.append(r)
    print(f"[i] matched {len(X)} clouds to labels "
          f"(of {len(df)} dataset rows, {len(npz)} npz files)")
    return np.stack(X) if X else np.empty((0, cfg["n_points"], 3)), pd.DataFrame(rows)


def augment(batch, cfg):
    """Per-epoch coordinate jitter ONLY. Rotation is NOT done here — it is applied
    once as precomputed rotated copies of the training rows (with correctly
    recomputed labels) in fit_fold, so a rotated cloud always has physically valid
    labels. Jitter affects xyz (and, harmlessly, the density channel)."""
    out = batch.copy()
    if cfg["jitter_std"] > 0:
        out += np.random.normal(0, cfg["jitter_std"], out.shape).astype(np.float32)
    return out


def load_tensors(meta, stem_col, cfg):
    """Return a list aligned to meta: each entry is the geometry's 6x6 stiffness
    (GPa) or None if no tensor was cached. Used only when rotation_aug is on."""
    # filename comes from the settings sheet, not from this file
    path = run_paths.data_path(str(cfg.get("tensor_npz", "tensors_ADMS.npz")))
    if not os.path.exists(path):
        sys.exit(f"[E] rotation_aug=1 but data/{os.path.basename(path)} is missing — "
                 "run collect_tensors.py first.")
    z = np.load(path)
    store = {k: z[k] for k in z.files}
    out, have = [], 0
    for _, r in meta.iterrows():
        stem = re.sub(r"\.stl$", "", str(r[stem_col]), flags=re.I)
        C = store.get(stem)
        if C is None:
            hit = [v for k, v in store.items() if stem in k or k in stem]
            C = hit[0] if hit else None
        out.append(None if C is None else np.asarray(C, float))
        have += C is not None
    print(f"[i] rotation_aug: matched {have}/{len(meta)} tensors "
          f"(geometries without a tensor are not rotated)")
    return out


def run():
    cfg = load_pn_cfg()
    np.random.seed(cfg["random_seed"])
    try:
        import torch
        import torch.nn as nn
    except ImportError:
        sys.exit("[E] PyTorch not installed — run: pip install torch")
    torch.manual_seed(cfg["random_seed"])

    X, meta = build_xy(cfg)
    if len(X) < 10:
        sys.exit(f"[E] only {len(X)} clouds matched — run pointcloud_prep.py first "
                 "and check dataset_csv stems.")

    class PointNet(nn.Module):
        """PointNet: shared MLP per point -> global max-pool -> regression head.

        in_dim is 3 (xyz) or 4 (xyzd) depending on use_density_channel.
        Deep head (1024-512-256) follows Zhezhe's v3; the shallow head is the
        original and is kept so the two can be compared.
        """
        def __init__(self, n_out, in_dim=3, deep=True):
            super().__init__()
            self.enc = nn.Sequential(
                nn.Conv1d(in_dim, 64, 1), nn.BatchNorm1d(64), nn.ReLU(),
                nn.Conv1d(64, 128, 1), nn.BatchNorm1d(128), nn.ReLU(),
                nn.Conv1d(128, 1024, 1), nn.BatchNorm1d(1024), nn.ReLU())
            if deep:
                self.head = nn.Sequential(
                    nn.Linear(1024, 512), nn.BatchNorm1d(512), nn.ReLU(), nn.Dropout(0.3),
                    nn.Linear(512, 256), nn.BatchNorm1d(256), nn.ReLU(), nn.Dropout(0.3),
                    nn.Linear(256, n_out))
            else:
                self.head = nn.Sequential(
                    nn.Linear(1024, 256), nn.ReLU(), nn.Dropout(0.3),
                    nn.Linear(256, 64), nn.ReLU(),
                    nn.Linear(64, n_out))
        def forward(self, x):                        # x: (B, N, C)
            x = x.transpose(1, 2)                    # -> (B, C, N)
            x = self.enc(x).max(dim=2).values        # global feature (B,1024)
            return self.head(x)

    tgts = [t for t in cfg["targets"] if t in meta.columns]
    Y = meta[tgts].apply(pd.to_numeric, errors="coerce").values.astype(np.float32)

    # MASKED multi-target training.
    # Targets are labelled to different depths (2026-07-20: C11/C12/C44/TAI/E/G
    # 115 rows, onset_n 99, shear_onset_n 72). Dropping every row that is
    # missing ANY target would cut 116 -> 70, a 39% loss, just to include one
    # sparse target. Instead keep every row with at least one label and mask
    # the loss, so a row with 8 of 9 labels still trains those 8.
    MASK = ~np.isnan(Y)
    keep = MASK.any(axis=1)
    X, Y, MASK, meta = X[keep], Y[keep], MASK[keep], meta[keep].reset_index(drop=True)
    print("[i] labels per target: " +
          ", ".join(f"{t}={int(MASK[:, j].sum())}" for j, t in enumerate(tgts)))
    print(f"[i] training on {len(X)} samples, targets={tgts}, "
          f"rotation_aug={cfg['rotation_aug']}")

    # ---- density as a 4th per-point channel (Zhezhe) ----------------------
    # The clouds were scaled to a unit box by pointcloud_prep, which removes
    # absolute scale. Density is re-injected explicitly, broadcast to every
    # point, so the encoder can use it.
    if int(cfg.get("use_density_channel", 0)):
        if "Rho_rel" not in meta.columns:
            sys.exit("[E] use_density_channel=1 but the dataset has no 'Rho_rel' column")
        rho = pd.to_numeric(meta["Rho_rel"], errors="coerce").values.astype(np.float32)
        if np.isnan(rho).any():
            sys.exit(f"[E] {int(np.isnan(rho).sum())} row(s) have no Rho_rel — "
                     "cannot build the density channel")
        d_chan = np.repeat(rho[:, None, None], X.shape[1], axis=1)   # (N,P,1)
        X = np.concatenate([X, d_chan], axis=2).astype(np.float32)   # (N,P,4)
        print(f"[i] density channel ON  -> input is xyzd, shape {X.shape}")
    in_dim = X.shape[2]

    # ---- log targets ------------------------------------------------------
    # Targets span ~129x. MSE minimises absolute error, so the net fits the
    # large values and is proportionally far off on the small ones (MAPE ~102%
    # at R2=0.43). Training on log10 makes MSE approximate relative error.
    use_log = int(cfg.get("log_targets", 0))
    if use_log:
        # only inspect LABELLED entries; NaNs are absent labels, not bad values
        if (Y[MASK] <= 0).any():
            bad = [t for j, t in enumerate(tgts) if (Y[MASK[:, j], j] <= 0).any()]
            print(f"[!] non-positive values in {bad} — log_targets disabled "
                  f"(log is undefined there)")
            use_log = 0
        else:
            Y = np.where(MASK, np.log10(np.where(MASK, Y, 1.0)), np.nan).astype(np.float32)
            print("[i] targets log10-transformed (un-logged before scoring)")

    # ---- grouped cross-validation --------------------------------------
    # REPLACES the old single random 80/20 split. That split let near-identical
    # geometries (same group) sit in train AND test, inflating the score. Now
    # every geometry is predicted out-of-fold by a model that never saw its
    # group. Inside each fold we still keep a THREE-way split: the fold's test
    # rows are never touched, and a val slice (whole groups) picks the epoch.
    dev = "cuda" if torch.cuda.is_available() else "cpu"
    clip = float(cfg.get("grad_clip", 0) or 0)

    def masked_mse(pred, target, mask):
        """MSE over labelled entries only. Unlabelled entries contribute nothing
        to the loss/gradient, so a row missing one target still trains the rest."""
        diff = (pred - target) * mask
        return (diff ** 2).sum() / torch.clamp(mask.sum(), min=1.0)

    def fit_fold(tr, va, te, fold):
        """Train one fresh net on tr (val=va picks the best epoch), predict te.
        Returns predictions for te in Y-space (log if use_log), destandardised.
        Standardisation stats come from tr's labelled rows only."""
        ymean = np.array([Y[tr, j][MASK[tr, j]].mean() if MASK[tr, j].any() else 0.0
                          for j in range(len(tgts))], dtype=np.float32)
        ystd = np.array([Y[tr, j][MASK[tr, j]].std() if MASK[tr, j].any() else 1.0
                         for j in range(len(tgts))], dtype=np.float32) + 1e-8
        net = PointNet(len(tgts), in_dim=in_dim,
                       deep=bool(int(cfg.get("deep_head", 0)))).to(dev)
        opt = torch.optim.Adam(net.parameters(), lr=cfg["lr"])
        decay = float(cfg.get("lr_decay", 1.0))
        sched = (torch.optim.lr_scheduler.StepLR(
                     opt, step_size=int(cfg.get("lr_step", 200)), gamma=decay)
                 if decay and decay != 1.0 else None)
        # NaNs -> 0 AFTER standardising; the mask makes them ignored.
        Ytr_s = np.where(MASK[tr], (Y[tr] - ymean) / ystd, 0.0).astype(np.float32)
        Yva_s = np.where(MASK[va], (Y[va] - ymean) / ystd, 0.0).astype(np.float32)
        Mtr, Mva = MASK[tr].astype(np.float32), MASK[va].astype(np.float32)
        Xtr, Xva = X[tr], X[va]

        # ---- Farooq full-tensor rotation augmentation (TRAIN rows only) ----
        # For each training geometry with a cached tensor, add n_rot rotated copies:
        # rotate the cloud xyz AND the stiffness tensor by the same R, then recompute
        # the direction-dependent labels from the rotated tensor. Invariant targets
        # reuse the parent's value; non-tensor targets (onset) stay masked. va/te are
        # NEVER augmented, so scoring is untouched.
        if cfg["rotation_aug"] and cfg["n_rot"] > 0:
            rng = np.random.RandomState(cfg["random_seed"] + 1000 + fold)
            es = cfg["es_gpa"]
            fdep, finv = set(cfg["frame_dependent"]), set(cfg["rotation_invariant"])
            aX, aY, aM = [], [], []
            for pos, gi in enumerate(tr):
                C = Ctens[gi]
                if C is None:
                    continue
                for _ in range(cfg["n_rot"]):
                    R = tensor_ops.random_rotation(rng)
                    xc = X[gi].copy()
                    # rotate xyz, THEN re-center + re-scale to the unit box exactly like
                    # pointcloud_prep does — a rotated box's extent grows to ~1.7, so
                    # without this the rotated copies sit at a different scale than every
                    # real (unit-box) cloud the net trained on. Density channel untouched.
                    p = X[gi][:, :3] @ R.T
                    p = p - p.mean(axis=0)
                    ext = float(np.ptp(p, axis=0).max())
                    if ext > 0:
                        p = p / ext
                    xc[:, :3] = p
                    Crot = tensor_ops.rotate_C(C, R)
                    yrow = np.zeros(len(tgts), np.float32)
                    mrow = np.zeros(len(tgts), np.float32)
                    for j, t in enumerate(tgts):
                        if t in fdep:
                            val = tensor_ops.frame_dependent_value(t, Crot, es)
                            if val is None or (use_log and val <= 0):
                                continue                     # can't use -> stays masked
                            v = np.log10(val) if use_log else val
                            yrow[j] = (v - ymean[j]) / ystd[j]; mrow[j] = 1.0
                        elif t in finv:                      # invariant -> reuse parent
                            yrow[j] = Ytr_s[pos, j]; mrow[j] = Mtr[pos, j]
                        # else: not tensor-derivable (onset) -> stays masked
                    aX.append(xc.astype(np.float32)); aY.append(yrow); aM.append(mrow)
            if aX:
                Xtr = np.concatenate([Xtr, np.stack(aX)], axis=0)
                Ytr_s = np.concatenate([Ytr_s, np.stack(aY)], axis=0)
                Mtr = np.concatenate([Mtr, np.stack(aM)], axis=0)
                print(f"    fold {fold}: +{len(aX)} rotated copies "
                      f"(train {len(tr)} -> {len(Xtr)} rows)")

        best_val, best_state = float("inf"), None
        for ep in range(cfg["epochs"]):
            net.train(); perm = np.random.permutation(len(Xtr))
            for b in range(0, len(Xtr), cfg["batch_size"]):
                bi = perm[b:b + cfg["batch_size"]]
                if len(bi) < 2:   # BatchNorm needs >=2 samples in train mode;
                    continue      # skip a lone trailing sample (re-shuffled next epoch)
                xb = torch.tensor(augment(Xtr[bi], cfg), device=dev)
                yb = torch.tensor(Ytr_s[bi], device=dev)
                mb = torch.tensor(Mtr[bi], device=dev)
                opt.zero_grad(); out = net(xb); loss = masked_mse(out, yb, mb)
                loss.backward()
                if clip > 0:
                    nn.utils.clip_grad_norm_(net.parameters(), clip)
                opt.step()
            if sched:
                sched.step()
            net.eval()
            with torch.no_grad():
                vloss = float(masked_mse(net(torch.tensor(Xva, device=dev)),
                                         torch.tensor(Yva_s, device=dev),
                                         torch.tensor(Mva, device=dev)).item())
            if vloss < best_val:
                best_val = vloss
                best_state = {k: v.detach().clone() for k, v in net.state_dict().items()}
        if best_state is not None:
            net.load_state_dict(best_state)
        net.eval()
        with torch.no_grad():
            return net(torch.tensor(X[te], device=dev)).cpu().numpy() * ystd + ymean

    # fold assignment: SAME group map as train.py (read from the full dataset
    # so the mapping is identical regardless of which rows matched clouds).
    full_df = pd.read_csv(cfg["dataset_csv"])
    fold_of, grouped, stem_col = cv_split.build_fold_map(
        full_df, cfg["group_cols"], cfg["n_folds"], seed=cfg["random_seed"])
    meta_fold = cv_split.folds_for(meta, cfg["group_cols"], fold_of, stem_col)
    gkey = (meta[cfg["group_cols"]].fillna("NA").astype(str).agg("|".join, axis=1).to_numpy()
            if grouped else meta[stem_col].astype(str).to_numpy())
    split_name = f"grouped{cfg['n_folds']}foldCV" if grouped else f"{cfg['n_folds']}foldCV"
    print(f"\n[i] {split_name}" + (f" on {cfg['group_cols']}" if grouped
          else " (no group_cols -> per-row split, near-duplicates NOT protected)"))

    def pick_val(trv):
        """Hold out whole groups from trv as the val set (~val_frac of its rows),
        so val is group-disjoint from train too. Deterministic given the seed."""
        groups = sorted(set(gkey[trv]))
        rng = np.random.RandomState(cfg["random_seed"])
        rng.shuffle(groups)
        want = max(2, int(round(float(cfg.get("val_frac", 0.15)) * len(trv))))
        chosen, cnt = set(), 0
        for g in groups:
            if cnt >= want or len(groups) - len(chosen) <= 1:
                break
            chosen.add(g); cnt += int(np.sum(gkey[trv] == g))
        va = trv[np.isin(gkey[trv], list(chosen))]
        tr = trv[~np.isin(gkey[trv], list(chosen))]
        return tr, va

    # per-geometry stiffness tensors (for rotation augmentation), aligned to meta.
    Ctens = load_tensors(meta, stem_col, cfg) if cfg["rotation_aug"] else [None] * len(meta)

    # out-of-fold predictions: every row filled exactly once, by its test fold
    oof = np.full((len(X), len(tgts)), np.nan, dtype=np.float32)
    # holdouts_only: skip the grouped-CV pass entirely. The CV pass is the
    # expensive half (n_folds fresh models); when the question is purely "does it
    # reach an unseen topology / density range", the folds add nothing.
    # env var wins so 6_VALIDATE_MODEL.bat can toggle it without editing the sheet
    _ho_only = int(os.environ.get("ML_HOLDOUTS_ONLY",
                                  cfg.get("holdouts_only", 0) or 0) or 0)
    if _ho_only:
        print("[i] holdouts_only=1 -> skipping the grouped-CV pass "
              "(no pointnet_v2_results.csv / predictions / cv_folds this run)")
    for f in ([] if _ho_only else range(cfg["n_folds"])):
        te = np.where(meta_fold == f)[0]
        trv = np.where(meta_fold != f)[0]
        if len(te) == 0 or len(trv) < 10:
            print(f"  fold {f}: test={len(te)} trainval={len(trv)} — skipped")
            continue
        tr, va = pick_val(trv)
        if len(va) == 0:                     # tiny data: fall back to random 15%
            rng = np.random.RandomState(cfg["random_seed"])
            va = tr[rng.permutation(len(tr))[:max(2, len(tr) // 7)]]
            tr = np.setdiff1d(tr, va)
        print(f"  fold {f}: train={len(tr)} val={len(va)} test={len(te)}")
        oof[te] = fit_fold(tr, va, te, f)

    # score out-of-fold predictions over ALL rows (same protocol as RF/GPR now)
    pred, ytrue = oof.copy(), Y.copy()
    if use_log:                              # un-log before scoring (labelled only)
        pred = np.where(~np.isnan(pred), np.power(10.0, pred), np.nan)
        ytrue = np.where(MASK, np.power(10.0, ytrue), np.nan)
    if not _ho_only:
        print(f"\n=== PointNet v2 — {split_name} out-of-fold "
          f"(N={len(X)}, density_ch={int(cfg.get('use_density_channel',0))}, "
          f"log={use_log}, deep={int(cfg.get('deep_head',0))}) ===")
    results = []
    for j, t in enumerate([] if _ho_only else tgts):
        mj = MASK[:, j] & ~np.isnan(oof[:, j])     # labelled AND predicted
        if mj.sum() < 3:
            print(f"  {t:14s} only {int(mj.sum())} scored row(s) — skipped")
            continue
        yt, pp = ytrue[mj, j], pred[mj, j]
        ss = np.sum((yt - yt.mean()) ** 2)
        r2 = 1 - np.sum((yt - pp) ** 2) / ss if ss > 0 else float("nan")
        m = np.abs(yt) > 1e-9
        mape = np.mean(np.abs((pp[m] - yt[m]) / yt[m])) * 100
        print(f"  {t:14s} R2={r2:.3f}  MAPE={mape:.1f}%  (N={int(mj.sum())})")
        results.append(dict(target=t, model="PointNet_v2", split=split_name,
                            N_test=int(mj.sum()), R2=round(float(r2), 4),
                            MAPE=round(float(mape), 2)))

    # per-row out-of-fold predictions + the fold map, so compare_models scores
    # RF/GPR on the IDENTICAL folds. (Replaces the old single-split file.)
    rec = []
    for j, t in enumerate(tgts):
        mj = MASK[:, j] & ~np.isnan(oof[:, j])
        for i in np.flatnonzero(mj):
            rec.append(dict(Run=meta.iloc[i][stem_col], target=t,
                            y_true=float(ytrue[i, j]), y_pred=float(pred[i, j])))
    if not _ho_only:
        # never write these in holdouts_only mode - they would be EMPTY and would
        # overwrite a real CV run's outputs if pointed at the same folder
        pred_csv = run_paths.run_path("pointnet_v2_predictions.csv")
        pd.DataFrame(rec).to_csv(pred_csv, index=False)
        folds_csv = run_paths.run_path("cv_folds.csv")
        pd.DataFrame({stem_col: meta[stem_col].values, "fold": meta_fold}).to_csv(folds_csv, index=False)
        print(f"[OK] wrote {pred_csv}")
        print(f"[OK] wrote {folds_csv}  ({cfg['n_folds']} folds, for identical-split comparison)")
    te = np.flatnonzero(~np.isnan(oof).all(axis=1))   # all scored rows (for the chart below)
    yte = ytrue

    if not _ho_only:
        out_csv = run_paths.run_path("pointnet_v2_results.csv")
        pd.DataFrame(results).to_csv(out_csv, index=False)
        print(f"[OK] wrote {out_csv}")

    # ---- HOLDOUTS: unseen topology / unseen family / extrapolation ------------
    # Grouped CV does NOT test topology generalisation: group_key is
    # family|topology|density-band, so every topology appears in EVERY fold
    # (measured 04/08: 18/18 TPMS topologies in all 5 folds). RF/GPR are already
    # scored on these holdouts by train.py; PointNet was not, so its ability to
    # reach an unseen topology was simply unknown. Same specs, same sheet.
    # Written to a SEPARATE file so compare_cv_vs_8020's merge on `target`
    # (which assumes one row per target) is unaffected.
    if int(cfg.get("run_holdouts", 1) or 0):
        try:
            from train import holdout_mask
        except Exception as e:
            print(f"[!] holdouts skipped, cannot import holdout_mask: {e}")
            holdout_mask = None
        specs = []
        for k in sorted(cfg):
            if str(k).startswith("holdout_") and cfg[k]:
                parts = [q.strip() for q in str(cfg[k]).split("|")]
                if len(parts) == 3:
                    specs.append(tuple(parts))
                else:
                    print(f"[!] bad holdout spec ignored: {cfg[k]} (need name|column|value)")
        if holdout_mask is not None and specs:
            print(f"\n=== PointNet v2 — {len(specs)} holdout(s), one fresh model each ===")
            ho_res, ho_rec = [], []
            for hidx, (hname, col, val) in enumerate(specs):
                m = holdout_mask(meta, col, val)
                if m is None:
                    print(f"  [!] {hname}: column '{col}' missing — skipped"); continue
                m = np.asarray(m, bool)
                te_h, trv_h = np.flatnonzero(m), np.flatnonzero(~m)
                if len(te_h) < 3 or len(trv_h) < 20:
                    print(f"  [!] {hname}: {len(te_h)} test / {len(trv_h)} train row(s) — skipped")
                    continue
                tr_h, va_h = pick_val(trv_h)
                if len(va_h) == 0:
                    rng = np.random.RandomState(cfg["random_seed"])
                    va_h = tr_h[rng.permutation(len(tr_h))[:max(2, len(tr_h) // 7)]]
                    tr_h = np.setdiff1d(tr_h, va_h)
                print(f"  {hname}: train={len(tr_h)} val={len(va_h)} test={len(te_h)}")
                # fit_fold uses `fold` only as a deterministic RNG offset -> must be an int
                ph = fit_fold(tr_h, va_h, te_h, 100 + hidx)
                yh = Y[te_h].copy()
                if use_log:
                    ph = np.power(10.0, ph)
                    yh = np.where(MASK[te_h], np.power(10.0, yh), np.nan)
                for j, t in enumerate(tgts):
                    ok = MASK[te_h, j] & ~np.isnan(ph[:, j])
                    if ok.sum() < 3:
                        continue
                    yt, pp = yh[ok, j], ph[ok, j]
                    ss = np.sum((yt - yt.mean()) ** 2)
                    r2 = 1 - np.sum((yt - pp) ** 2) / ss if ss > 0 else float("nan")
                    mm = np.abs(yt) > 1e-9
                    mape = (np.mean(np.abs((pp[mm] - yt[mm]) / yt[mm])) * 100
                            if mm.any() else float("nan"))
                    print(f"      {t:14s} R2={r2:+.3f}  MAPE={mape:.1f}%  (N={int(ok.sum())})")
                    ho_res.append(dict(target=t, model="PointNet_v2", split=hname,
                                       N_test=int(ok.sum()), R2=round(float(r2), 4),
                                       MAPE=round(float(mape), 2)))
                    # yt/pp are ALREADY masked (length = ok.sum()), so walk them
                    # positionally and map back into te_h for the Run name.
                    for k, i in enumerate(np.flatnonzero(ok)):
                        ho_rec.append(dict(Run=meta.iloc[te_h[i]][stem_col], split=hname,
                                           target=t, y_true=float(yt[k]), y_pred=float(pp[k])))
            if ho_res:
                hc = run_paths.run_path("pointnet_holdouts.csv")
                pd.DataFrame(ho_res).to_csv(hc, index=False)
                pd.DataFrame(ho_rec).to_csv(
                    run_paths.run_path("pointnet_holdout_predictions.csv"), index=False)
                print(f"[OK] wrote {hc}")
                print("  NOTE: a NEGATIVE R2 here means worse than predicting the mean, i.e. the")
                print("        model does not reach that unseen topology/range at all.")

    # parity chart
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        os.makedirs(cfg["charts_dir"], exist_ok=True)
        k = len(tgts)
        fig, axes = plt.subplots(1, k, figsize=(4.4 * k, 4.4), squeeze=False)
        for j, t in enumerate(tgts):
            ax = axes[0][j]
            mj = MASK[:, j] & ~np.isnan(oof[:, j])       # scored rows only
            yt, pp = yte[mj, j], pred[mj, j]
            if len(yt) == 0:
                ax.set_title(f"{t} (no data)", fontsize=10); continue
            lim = [min(yt.min(), pp.min()) * 0.9, max(yt.max(), pp.max()) * 1.05]
            ax.plot(lim, lim, "k--", lw=1, alpha=0.6)
            ax.scatter(yt, pp, s=32, c="#500778", alpha=0.8, edgecolors="white", lw=0.4)
            ax.set_title(t, fontsize=10); ax.set_xlabel("FEA (actual)")
            ax.grid(alpha=0.25)
        axes[0][0].set_ylabel("PointNet prediction")
        fig.suptitle(f"PointNet — {split_name} out-of-fold (N={len(X)})",
                     color="#500778", fontweight="bold")
        fig.tight_layout(rect=[0, 0, 1, 0.93])
        fp = os.path.join(cfg["charts_dir"], "parity_pointnet.png")
        fig.savefig(fp, dpi=160); plt.close(fig); print("[OK]", fp)
    except Exception as ex:
        print("[!] chart skipped:", ex)


if __name__ == "__main__":
    run()
