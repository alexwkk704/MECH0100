"""
run_paths.py — ONE place that decides where every output goes.

Clean filing (2026-07-21). Two kinds of output, kept apart:

  ML/data/                 DERIVED INPUTS — features, joined dataset, point clouds.
                           Regenerated only when the STLs or the master labels
                           change. Shared by every run; NOT per-run.

  ML/runs/<stamp>/         RESULTS OF ONE RUN — metric CSVs, charts, MANIFEST.
                           One folder per RUN_ALL_ML invocation. Nothing here is
                           ever overwritten in place, so a chart can never be
                           mistaken for a different run (this is what fixed the
                           "charts not always updated" problem).

RUN_ALL_ML.bat computes the stamp ONCE and exports ML_RUN_DIR, so all 7 steps of
a single pipeline run write into the SAME run folder. Running one script by hand
with no env var falls back to ML/runs/scratch, so standalone use still works.

No hardcoding: this file only decides the folder layout; filenames still come
from ML_settings.xlsx. Shared with Zhezhe/Farooq — their data uses the same
scheme with their own settings.
"""
import os

def _find_base():
    here = os.path.dirname(os.path.abspath(__file__))            # .../ML/scripts
    for cand in (here, os.path.dirname(here)):
        if os.path.exists(os.path.join(cand, "ML_settings.xlsx")):
            return cand
    return os.path.dirname(here)                                 # scripts/ -> ML/

BASE_DIR = _find_base()
DATA_DIR = os.path.join(BASE_DIR, "data")


def data_dir():
    os.makedirs(DATA_DIR, exist_ok=True)
    return DATA_DIR


def data_path(name):
    """A derived-input file in ML/data/ (keeps only the filename from settings)."""
    return os.path.join(data_dir(), os.path.basename(str(name)))


def pointclouds_dir():
    p = os.path.join(data_dir(), "pointclouds")
    os.makedirs(p, exist_ok=True)
    return p


def run_dir():
    """The current run's result folder. Set by RUN_ALL_ML via ML_RUN_DIR;
    falls back to ML/runs/scratch for standalone script runs."""
    rd = os.environ.get("ML_RUN_DIR") or os.path.join(BASE_DIR, "runs", "scratch")
    os.makedirs(rd, exist_ok=True)
    return rd


def run_path(rel):
    """A result file/subpath inside the current run folder."""
    p = os.path.join(run_dir(), rel)
    d = os.path.dirname(p)
    if d:
        os.makedirs(d, exist_ok=True)
    return p


# ---------------------------------------------------------------------------
# PROFILES  (added 2026-07-31)
# One workbook, three models. A .bat sets ML_PROFILE=ADMS | TPMS | COMBINED and
# every script then reads `train_<PROFILE>` instead of `train` (same for
# `pointnet_<PROFILE>`). If that sheet does not exist it falls back to the plain
# name, so running a script by hand with no env var behaves exactly as before.
# Results go to runs/<PROFILE>/<stamp>/ so the three models never mix.
# ---------------------------------------------------------------------------

def profile():
    """Current profile name, or '' when running unprofiled."""
    return str(os.environ.get("ML_PROFILE", "")).strip()


def sheet_for(wb, base):
    """Sheet to read for `base` under the current profile, with fallback."""
    p = profile()
    if p:
        cand = f"{base}_{p}"
        if cand in wb.sheetnames:
            return cand
    return base


def _settings_path():
    return os.path.join(BASE_DIR, "ML_settings.xlsx")


def sheet_value(base_sheet, key, default=""):
    """One value from `<base_sheet>_<PROFILE>` (falling back to `<base_sheet>`).

    2026-08-10: the .bat files used to hardcode "dataset_TPMS.csv" in their
    existence guards and run snapshots, which silently defeated the whole
    "filenames live in ML_settings.xlsx" design - change a sheet and the guard
    still checked the old file. Everything now asks here instead.
    """
    import openpyxl
    xls = _settings_path()
    if not os.path.exists(xls):
        return default
    wb = openpyxl.load_workbook(xls, data_only=True)
    sh = sheet_for(wb, base_sheet)
    if sh not in wb.sheetnames:
        return default
    for row in wb[sh].iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None and str(row[0]).strip() == key:
            return "" if row[1] is None else str(row[1]).strip()
    return default


def dataset_path():
    """Absolute path of the dataset the CURRENT profile trains on."""
    return data_path(sheet_value("train", "dataset_csv", "dataset.csv"))


def tensor_npz_path():
    """Absolute path of the stiffness-tensor cache for the CURRENT profile."""
    prof = profile() or "ADMS"
    return data_path(sheet_value("pointnet", "tensor_npz", f"tensors_{prof}.npz"))


def runs_root():
    """runs/  or  runs/<PROFILE>/  when a profile is set."""
    p = profile()
    root = os.path.join(BASE_DIR, "runs", p) if p else os.path.join(BASE_DIR, "runs")
    os.makedirs(root, exist_ok=True)
    return root


# ---------------------------------------------------------------------------
# CLI so .bat files can ask for a path instead of hardcoding a filename:
#     for /f "usebackq delims=" %%p in (`python scripts\run_paths.py dataset`) do set "DS=%%p"
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import sys
    what = (sys.argv[1] if len(sys.argv) > 1 else "dataset").lower()
    if what == "dataset":
        print(dataset_path())
    elif what in ("tensors", "tensor", "npz"):
        print(tensor_npz_path())
    elif what == "datadir":
        print(data_dir())
    elif what == "name":                     # bare filename, for --out style args
        print(os.path.basename(dataset_path()))
    else:
        sys.exit(f"unknown request: {what}")
