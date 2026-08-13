#!/usr/bin/env python3
"""set_pointnet_key.py — set one key on a pointnet_<PROFILE> settings sheet.

Why this exists
---------------
train_pointnet_v2.py takes every parameter from ML_settings.xlsx and has no
command-line override. To measure how much of a cross-validation score is
training noise rather than data, the same run must be repeated at several seeds,
so the seed has to be written into the sheet between runs.

It changes only the keys you name, and prints old -> new for each so the change
is visible in the batch log. It will not create a key that does not already
exist — that would silently introduce a setting nobody chose.

Usage:
    python set_pointnet_key.py --profile ADMS random_seed=2
    python set_pointnet_key.py --profile ADMS random_seed=2 run_holdouts=0
"""
import os, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))


def find_settings(start, name="ML_settings.xlsx"):
    d = start
    for _ in range(4):
        p = os.path.join(d, name)
        if os.path.exists(p):
            return p
        d = os.path.dirname(d)
    raise SystemExit(f"[E] {name} not found above {start}")


def main():
    args = sys.argv[1:]
    if not args:
        raise SystemExit(__doc__)

    profile = "ADMS"
    if "--profile" in args:
        i = args.index("--profile")
        profile = args[i + 1]
        del args[i:i + 2]

    pairs = []
    for a in args:
        if "=" not in a:
            raise SystemExit(f"[E] expected key=value, got {a!r}")
        k, v = a.split("=", 1)
        pairs.append((k.strip(), v.strip()))
    if not pairs:
        raise SystemExit(__doc__)

    sheet = f"pointnet_{profile}"
    xlsx = find_settings(HERE)
    wb = openpyxl.load_workbook(xlsx)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"[E] sheet '{sheet}' not in {xlsx}\n    sheets: {wb.sheetnames}")
    ws = wb[sheet]

    index = {}
    for row in ws.iter_rows():
        if row[0].value is not None:
            index[str(row[0].value).strip()] = row[1]

    for k, v in pairs:
        if k not in index:
            raise SystemExit(
                f"[E] '{k}' is not an existing key on '{sheet}' — refusing to create it.\n"
                f"    keys present: {sorted(index)}")
        cell = index[k]
        old = cell.value
        try:
            cell.value = int(v)
        except ValueError:
            try:
                cell.value = float(v)
            except ValueError:
                cell.value = v
        print(f"[set] {sheet}!{k}  {old}  ->  {cell.value}")

    wb.save(xlsx)


if __name__ == "__main__":
    main()
