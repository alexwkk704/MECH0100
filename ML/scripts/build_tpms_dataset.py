"""
build_tpms_dataset.py - turn the partner's RAW STL + RAW FEA tensors into a
dataset that is directly mergeable with dataset_ADMS.csv.

TRUST BOUNDARY (Alex, 2026-07-31)
---------------------------------
The ONLY partner inputs used are:
    * his STL files            -> features come from OUR feature_extraction.py
    * his raw 6x6 FEA tensors  -> labels come from OUR tensor_ops formulas
Nothing else of his is read. In particular his own feature CSVs are IGNORED: they
were produced by an older copy of our extractor (plain-mean curvature, std instead
of IQR, no fabric tensor), and mixing two definitions of the same column name is
the single easiest way to make a merged model quietly meaningless.

WHY THE LABELS ARE RECOMPUTED RATHER THAN COPIED
------------------------------------------------
His solid material is a polymer (E_s = 1800 MPa, nu = 0.3) and ours is steel, so
RAW C11 is not comparable between the datasets - a model trained on both would
learn "polymer or steel", not "what shape is this". Every label here is normalised
by the solid modulus (C11_n, E_over_Es, G_over_Gs) or is dimensionless by
construction (TAI, nu_iso, Zener_Z), exactly as labels_join.py does for ADMS.

INPUTS
    features_TPMS.csv   from: python feature_extraction.py --stl-folder <his stl>
                                   --output-csv features_TPMS.csv --dataset-tag TPMS
                                   --cell-size <mm> --thickness-convention double
    matrix_data/<Topology>/<name>_density<d>.csv   his raw 6x6, one per geometry

OUTPUT
    data/dataset_TPMS.csv   features + normalised labels + topology, joined on Run

USAGE
    python build_tpms_dataset.py
    python build_tpms_dataset.py --features features_TPMS.csv --tensor-root <path>
                                 --tensor-unit kPa --es 1800 --es-unit MPa --nu 0.3
"""
import argparse, csv, glob, os, re, sys
from pathlib import Path

import numpy as np

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import run_paths, tensor_ops

ML_DIR = SCRIPT_DIR.parent
SHARE = ML_DIR.parent
ROOT = SHARE.parent

TO_GPA = {"pa": 1e-9, "kpa": 1e-6, "mpa": 1e-3, "gpa": 1.0}
LABELS = ["C11_n", "C12_n", "C44_n", "TAI", "E_over_Es", "G_over_Gs", "nu_iso", "Zener_Z"]


def to_gpa(u):
    k = str(u).strip().lower()
    if k not in TO_GPA:
        sys.exit(f"[E] unknown unit '{u}' - use Pa | kPa | MPa | GPa")
    return TO_GPA[k]


def read_C6(path):
    """Read a 6x6 stiffness csv -> symmetric 6x6 in the file's own units."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.reader(fh):
            v = []
            for x in r:
                try:
                    v.append(float(x))
                except (TypeError, ValueError):
                    pass
            if len(v) >= 6:
                rows.append(v[:6])
    if len(rows) < 6:
        return None
    C = np.asarray(rows[:6], float)
    return 0.5 * (C + C.T)          # symmetrise: FEA output can differ in the last digit


def labels_from_C6(C6_raw, unit, es_raw, es_unit, nu_solid):
    """IDENTICAL formulas to Validate Zhezhe/extract_partner.compute_labels()."""
    C = np.asarray(C6_raw, float) * to_gpa(unit)
    es = float(es_raw) * to_gpa(es_unit)
    gs = es / (2.0 * (1.0 + float(nu_solid)))
    K, G = tensor_ops.voigt_reuss_hill(C)
    den = 3 * K + G
    E_eff = 9 * K * G / den if den else float("nan")
    nu_eff = (3 * K - 2 * G) / (2 * den) if den else float("nan")
    # Zener anisotropy: 2*C44 / (C11 - C12). Cubic-symmetry measure, dimensionless;
    # reported alongside TAI because they answer slightly different questions and
    # ADMS carries both.
    dz = C[0, 0] - C[0, 1]
    Z = (2.0 * C[3, 3] / dz) if abs(dz) > 1e-12 else float("nan")
    return {"C11_n": C[0, 0] / es, "C12_n": C[0, 1] / es, "C44_n": C[3, 3] / es,
            "TAI": tensor_ops.tai(C), "E_over_Es": E_eff / es, "G_over_Gs": G / gs,
            "nu_iso": nu_eff, "Zener_Z": Z}


LAST_NUM = re.compile(r"([0-9]*\.?[0-9]+)(?!.*[0-9])")


def last_number(s):
    m = LAST_NUM.search(str(s))
    return float(m.group(1)) if m else None


# 2026-08-10: topologies present in the partner's tensor folder but ABSENT from
# his own dataset_TPMS_quad.csv. Their tensors differ from the linear snapshot by
# 1-2%, which is the same size as confirmed-quadratic topologies (D is 1.03%), so
# they are probably genuine quadratic runs - but there is no reference tensor to
# prove it and he excluded them himself. Joining them would put ~20 possibly-linear
# rows into a quadratic training set. Empty this list once he confirms.
TOPO_EXCLUDE = {"±Y", "+-Y", "±Y Generator", "+-Y Generator",
                "C(±Y)", "C(+-Y)", "C(±Y) Generator", "C(+-Y) Generator"}


def _topo_aliases(topo):
    """Spellings of one topology folder that should all resolve to the same key.

    The partner's QUADRATIC export (Zhezhe/week7/modulus) names its folders "D'",
    "C(I2-Y)", "G'2" - it dropped the " Generator" suffix the earlier linear
    export (week7/matrix_data) used. Our `topology` column comes from HIS STL
    folder names, which still carry " Generator", so a literal folder-name match
    joins ZERO rows and this script exits with "nothing joined". Aliasing keeps
    the join working with either export layout and changes nothing when the
    folder name already matches.
    """
    names = {topo}
    if not topo.endswith(" Generator"):
        names.add(f"{topo} Generator")
    return names


def index_tensors(root):
    """(topology_folder, density rounded to 2dp) -> csv path. Mirrors extract_partner."""
    out = {}
    found, skipped = [], set()
    for p in glob.glob(os.path.join(root, "*", "*.csv")) + glob.glob(os.path.join(root, "*.csv")):
        topo = os.path.basename(os.path.dirname(p))
        if topo in TOPO_EXCLUDE:
            skipped.add(topo)
            continue
        d = last_number(os.path.splitext(os.path.basename(p))[0])
        if d is not None:
            found.append((topo, round(d, 2), p))
    # pass 1: exact folder names win outright
    for topo, d, p in found:
        out[(topo, d)] = p
    # pass 2: aliases fill gaps only - never displace a real folder of that name
    for topo, d, p in found:
        for alias in _topo_aliases(topo):
            out.setdefault((alias, d), p)
    if skipped:
        print(f"[i] TOPO_EXCLUDE: skipped {sorted(skipped)} - element order unconfirmed")
    return out


TPMS_LABEL_DEFAULTS = [
    ("features_csv", "features_TPMS.csv"),
    ("tensor_root",  "../../Zhezhe/week7/matrix_data"),
    ("tensor_unit",  "kPa"),
    ("es",           "1800"),
    ("es_unit",      "MPa"),
    ("nu",           "0.3"),
    ("element_order", "linear"),
    ("element_order_ref", ""),
    ("output_csv",   "dataset_TPMS.csv"),
    ("vf_vs_name_min", "0.85"),
    ("vf_vs_name_max", "1.15"),
]


def load_tpms_settings():
    """Read the 'tpms_labels' sheet of ML_settings.xlsx, creating it if absent.
    Paths are resolved relative to the ML/ folder, never to this script."""
    import openpyxl
    xls = ML_DIR / "ML_settings.xlsx"
    if not xls.exists():
        return dict(TPMS_LABEL_DEFAULTS)
    wb = openpyxl.load_workbook(xls)
    if "tpms_labels" not in wb.sheetnames:
        ws = wb.create_sheet("tpms_labels")
        ws.append(["setting", "value", "explanation"])
        for k, v in TPMS_LABEL_DEFAULTS:
            ws.append([k, v, ""])
        wb.save(xls)
        print("[i] created 'tpms_labels' sheet in ML_settings.xlsx with defaults")
        return dict(TPMS_LABEL_DEFAULTS)
    out = dict(TPMS_LABEL_DEFAULTS)
    for r in wb["tpms_labels"].iter_rows(min_row=2, values_only=True):
        if r and r[0] is not None and r[1] is not None:
            out[str(r[0]).strip()] = str(r[1]).strip()
    tr = out["tensor_root"]
    if not os.path.isabs(tr):
        out["tensor_root"] = os.path.normpath(os.path.join(str(ML_DIR), tr))
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    # DEFAULTS COME FROM THE WORKBOOK, sheet 'tpms_labels'. Nothing about the
    # partner's material, units or folder layout is written into this file - a
    # different partner is a different row in Excel, not a code edit.
    S = load_tpms_settings()
    ap.add_argument("--features", default=S["features_csv"],
                    help="feature csv from OUR feature_extraction.py (in ML/data/)")
    ap.add_argument("--tensor-root", default=S["tensor_root"])
    ap.add_argument("--tensor-unit", default=S["tensor_unit"],
                    help="unit of the partner's raw 6x6 values. Sanity check: "
                         "E_over_Es should land near 0.05 at VF 0.15 - if it comes "
                         "out ~50 or ~0.00005 the unit is wrong by 1000x.")
    ap.add_argument("--es", type=float, default=float(S["es"]), help="partner solid modulus")
    ap.add_argument("--es-unit", default=S["es_unit"])
    ap.add_argument("--nu", type=float, default=float(S["nu"]), help="partner solid Poisson ratio")
    ap.add_argument("--element-order", default=S.get("element_order", ""),
                    help="FE element order behind the partner's tensors (linear|quadratic). "
                         "Stamped onto every row so a combined dataset cannot silently mix orders.")
    ap.add_argument("--element-order-ref", default=S.get("element_order_ref", ""),
                    help="a tensor of KNOWN element order; used to detect a silent data swap")
    ap.add_argument("--out", default=S["output_csv"])
    ap.add_argument("--density-bin", type=float, default=0.02,
                    help="width of the density band inside group_key (default 0.02). "
                         "MUST match merge_datasets.py --density-bin so the TPMS-only "
                         "and combined models group the same way.")
    args = ap.parse_args()

    feats = Path(run_paths.data_path(args.features))
    if not feats.exists():
        sys.exit(f"[E] {feats} not found.\n"
                 f"    Run feature_extraction.py on the partner STLs first:\n"
                 f"      python feature_extraction.py --stl-folder <his stl model> \\\n"
                 f"          --output-csv {args.features} --dataset-tag TPMS \\\n"
                 f"          --cell-size <mm> --thickness-convention double")
    rows = list(csv.DictReader(feats.open(newline="", encoding="utf-8")))
    print(f"[ok] {len(rows)} feature rows from {feats.name}")

    tens = index_tensors(args.tensor_root)
    print(f"[ok] {len(tens)} raw tensors indexed under {args.tensor_root}")
    if not tens:
        sys.exit("[E] no tensor csv files found - check --tensor-root")

    out_rows, miss, mismatch = [], [], []
    VF_LO = float(S.get('vf_vs_name_min', 0.85))
    VF_HI = float(S.get('vf_vs_name_max', 1.15))
    for r in rows:
        if str(r.get("status", "OK")).upper() != "OK":
            continue
        stem = str(r.get("file", "")).rsplit(".", 1)[0]
        topo = str(r.get("topology", "")).strip()
        dens = last_number(stem)
        key = (topo, round(dens, 2)) if dens is not None else None
        path = tens.get(key)
        if path is None:
            miss.append(f"{topo}/{stem}")
            continue
        C6 = read_C6(path)
        if C6 is None:
            miss.append(f"{topo}/{stem} (unreadable 6x6)")
            continue
        # GATE: the density in the filename is what picked the tensor above, so if the
        # mesh does not match its own name the labels belong to a different geometry.
        # Found 2026-07-31: two C(+-Y) files had been regenerated without renaming, so
        # dense geometry was about to be paired with sparse-density FEA labels.
        try:
            _vf = float(r["VF"]); _ratio = _vf / dens if dens else None
        except (TypeError, ValueError, KeyError, ZeroDivisionError):
            _ratio = None
        if _ratio is not None and not (VF_LO < _ratio < VF_HI):
            mismatch.append((topo, stem, dens, _vf, _ratio))
            continue

        row = dict(r)
        row["Run"] = stem
        row["topology"] = topo
        row.update({k: round(v, 6) for k, v in
                    labels_from_C6(C6, args.tensor_unit, args.es, args.es_unit, args.nu).items()})
        # Rho_rel: use OUR VF measured from HIS STL, never his own density number.
        try:
            row["Rho_rel"] = float(r["VF"])
        except (TypeError, ValueError, KeyError):
            row["Rho_rel"] = ""
        out_rows.append(row)

    if not out_rows:
        sys.exit("[E] nothing joined - check the topology folder names match")

    # ---- group_key: the CV grouping key for the TPMS-only model ----------------
    # Same construction merge_datasets.py uses for the combined set, so the two
    # models group identically: family | topology | density band. Without this the
    # TPMS sheet had to group on density_param, which this dataset does not carry
    # (it is an ADMS filename parameter) - so grouping silently fell back to
    # topology alone and near-identical densities could straddle a fold boundary.
    band_w = float(args.density_bin)
    for r in out_rows:
        try:
            d = float(r.get("Rho_rel") or r.get("VF") or 0.0)
        except (TypeError, ValueError):
            d = 0.0
        band = int(d / band_w) if band_w > 0 else 0
        r["group_key"] = f"TPMS|{r['topology']}|{band}"
        r["element_order"] = args.element_order

    # ---- element-order provenance -------------------------------------------
    # The partner's raw tensors carry no marker for FE element order, and linear
    # vs quadratic moves E/Es by ~6-8% and TAI by 13-66% (measured 04/08 from his
    # own _linear/_quad pairs). If he re-runs and drops new files in, the numbers
    # change but nothing else does. This compares one production tensor against a
    # tensor of KNOWN order, so a swap is caught instead of silently averaged in.
    ref = str(args.element_order_ref or "").strip()
    if ref:
        rp = ref if os.path.isabs(ref) else os.path.normpath(os.path.join(str(ML_DIR), ref))
        m = LAST_NUM.search(os.path.basename(rp))
        rd = round(float(m.group(1)), 2) if m else None
        # Match the reference file to a topology by its FULL name, not a token.
        # "P Generator".split()[0] == "P", and "p" is a substring of "iwp", so a
        # first-token test silently matched the wrong topology - and because the
        # candidates came from a set, WHICH wrong one varied between runs.
        # His filenames are "<Topology>_density<rho>_<order>.csv", so take the part
        # before "_density"; fall back to the longest topology name contained in it.
        base = os.path.basename(rp).lower()
        topos = sorted({r["topology"] for r in out_rows})
        head = base.split("_density")[0].strip()
        rtopo = next((t for t in topos if t.lower() == head), None)
        if rtopo is None:
            cand = [t for t in topos if t.lower() in base]
            rtopo = max(cand, key=len) if cand else None
        prod = tens.get((rtopo, rd)) if rtopo is not None else None
        if not os.path.exists(rp):
            print(f"[!] element_order_ref not found, swap check SKIPPED: {rp}")
        elif prod is None:
            print(f"[!] no production tensor matches the ref ({rtopo}, {rd}) - swap check SKIPPED")
        else:
            A, B = read_C6(prod), read_C6(rp)
            if A is None or B is None:
                print("[!] could not read both tensors - swap check SKIPPED")
            else:
                A = np.asarray(A, float); B = np.asarray(B, float)
                d = float(np.max(np.abs(A - B) / np.maximum(np.abs(B), 1e-9)) * 100)
                dc = float(abs(A[0, 0] - B[0, 0]) / abs(B[0, 0]) * 100)
                if dc < 1.0:
                    print(f"[ok] element order CONFIRMED '{args.element_order}': the production "
                          f"{rtopo} rho={rd} tensor matches the reference (C11 differs {dc:.3f}%)")
                else:
                    print("\n" + "=" * 74)
                    print("[!!] ELEMENT ORDER MAY BE STALE. The production tensor no longer matches")
                    print(f"     the reference: C11 differs {dc:.2f}% (max term {d:.1f}%).")
                    print(f"     The sheet still declares element_order = '{args.element_order}'.")
                    print("     If the partner re-ran his FEA, set tpms_labels!element_order to the")
                    print("     new order and point element_order_ref at a matching reference,")
                    print("     otherwise every TPMS label is mislabelled in the dataset.")
                    print("=" * 74 + "\n")
    print(f"[i] labels stamped element_order = '{args.element_order}'")

    cols = list(out_rows[0].keys())
    dest = Path(run_paths.data_path(args.out))
    with dest.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in out_rows:
            w.writerow({k: r.get(k, "") for k in cols})

    e = [float(r["E_over_Es"]) for r in out_rows]
    v = [float(r["VF"]) for r in out_rows]
    print(f"\n[ok] {len(out_rows)} rows joined   ({len(miss)} features had no tensor)")
    print(f"     topologies: {len(set(r['topology'] for r in out_rows))}")
    print(f"     group_key : {len(set(r['group_key'] for r in out_rows))} groups "
          f"(family|topology|density band of {band_w})")
    print(f"     VF         {min(v):.4f} .. {max(v):.4f}")
    print(f"     E_over_Es  {min(e):.5f} .. {max(e):.5f}"
          f"   <- expect ~0.02-0.15; if it is ~1000x off, --tensor-unit is wrong")
    if mismatch:
        print("\n  *** REJECTED - the mesh does not match the density in its own filename ***")
        print("      The filename density selects the FEA tensor, so these would have been")
        print("      given labels from a DIFFERENT geometry. Tolerance %.2f-%.2f (crosscheck sheet)."%(VF_LO,VF_HI))
        for t, stm, dn, vf, rt in mismatch:
            print("      %-22s %-32s name=%.2f  measured VF=%.5f  ratio=%.3f"%(t, stm, dn, vf, rt))
        print("      -> tell the partner: these STLs were likely regenerated without renaming.")

    if miss:
        import collections
        by_topo = collections.Counter(m.split("/")[0] for m in miss)
        print("\n  EXCLUDED - no raw 6x6 tensor on disk for these geometries:")
        for t, c in by_topo.most_common():
            print(f"     {t:<28} {c} rows")
        print("""
  These are NOT approximated. His dataset_TPMS.csv does carry labels for them, and
  its C11_n / C44_n / E_over_Es / G_over_Gs agree with ours to 0.005% - but its TAI
  does NOT (median 70% off, max 768%), so it was computed with a different formula.
  TAI cannot be rebuilt from C11/C12/C44 either: assuming cubic symmetry gives a
  median error of 0.1% but a 90th percentile of 3% and a worst case of 91%, i.e. the
  tensors are only approximately cubic. Since TAI is a headline target, guessing it
  is worse than dropping the rows.
  TO RECOVER THEM: ask Zhezhe for the raw 6x6 FEA tensors for these topologies and
  drop them into the matrix_data tree. This script will pick them up automatically.""")
    print(f"\n  -> {dest}")


if __name__ == "__main__":
    main()
