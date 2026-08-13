"""
collect_tensors.py — cache the FULL 6x6 stiffness tensor per geometry.

Needed by the PointNet rotation augmentation (tensor_ops / Farooq's full-tensor
method): rotating a shape requires rotating its stiffness tensor, and the dataset
only stores the cubic-reduced C11/C12/C44. The full 6x6 lives in each run's
per-geometry CSV under the FEA results tree.

PROFILE AWARE (2026-08-03). ML_PROFILE picks BOTH the source and the output:
  ADMS      <Results>/*/Data/<stem>/<stem>.csv         our nTop FEA tree
  TPMS      <tensor_root>/<topology>/*_density<d>.csv  the partner's raw 6x6
  COMBINED  both of the above

TENSORS ARE CACHED NORMALISED (C / Es, dimensionless), not in GPa. The two
families have different solids - steel 200 GPa vs polymer 1.8 GPa - so a single
es_gpa cannot serve a combined dataset. Rotation commutes with dividing by a
scalar, so rotating C/Es is identical to normalising the rotated C. The pointnet
sheet therefore sets es_gpa = 1.0 for every profile.
Resume-safe: geometries already in the npz are kept; only missing ones are read.

Paths come from ML_settings.xlsx: the 'labels' sheet master_xlsx points at
Results_summary.xlsx, whose folder IS the results tree. No hardcoding.

Usage:  python collect_tensors.py
"""

import os, sys, re, csv, glob
import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import run_paths


def _find_settings(base, name="ML_settings.xlsx"):
    here = os.path.join(base, name)
    if os.path.exists(here):
        return here
    up = os.path.join(os.path.dirname(base), name)
    return up if os.path.exists(up) else here


SETTINGS_XLSX = _find_settings(HERE)
BASE_DIR = os.path.dirname(SETTINGS_XLSX)


def _out_name(prof):
    """Output filename for this profile - read from the pointnet sheet, never hardcoded."""
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    sh = run_paths.sheet_for(wb, "pointnet")
    d = {str(k).strip(): v for k, v, *_ in wb[sh].iter_rows(min_row=2, values_only=True) if k}
    return str(d.get("tensor_npz", f"tensors_{prof}.npz"))


def _es_gpa_adms():
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    lab = {str(k).strip(): v for k, v, *_ in wb["labels"].iter_rows(min_row=2, values_only=True) if k}
    return float(lab.get("es_gpa", 200.0))


def _cfg():
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    lab = {str(k).strip(): v for k, v, *_ in wb["labels"].iter_rows(min_row=2, values_only=True) if k}
    tr = {str(k).strip(): v for k, v, *_ in wb[run_paths.sheet_for(wb, "train")].iter_rows(min_row=2, values_only=True) if k}
    master = str(lab["master_xlsx"])
    master = os.path.normpath(master if os.path.isabs(master) else os.path.join(BASE_DIR, master))
    results_dir = os.path.dirname(master)
    dataset_csv = run_paths.data_path(str(tr["dataset_csv"]))
    return results_dir, dataset_csv


def read_c6_csv(path):
    """Parse a per-run C-tensor CSV -> symmetric 6x6 in GPa (MPa/1000)."""
    rows = []
    for line in open(path):
        nums = re.findall(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?", line)
        if len(nums) >= 6:
            rows.append([float(x) for x in nums[:6]])
    if len(rows) < 6:
        return None
    C = np.array(rows[-6:])                 # last 6x6 block
    return 0.5 * (C + C.T) / 1000.0         # symmetrise, MPa -> GPa


LAST_NUM = re.compile(r"([0-9]*\.?[0-9]+)(?!.*[0-9])")


def _tpms_cfg():
    """tensor_root / unit / es for the partner's raw 6x6 - from the tpms_labels sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    if "tpms_labels" not in wb.sheetnames:
        return None
    d = {str(k).strip(): v for k, v, *_ in wb["tpms_labels"].iter_rows(min_row=2, values_only=True) if k}
    root = str(d.get("tensor_root", ""))
    if root and not os.path.isabs(root):
        root = os.path.normpath(os.path.join(BASE_DIR, root))
    unit = str(d.get("tensor_unit", "kPa")).strip().lower()
    to_gpa = {"pa": 1e-9, "kpa": 1e-6, "mpa": 1e-3, "gpa": 1.0}.get(unit, 1e-6)
    es = float(d.get("es", 1800.0))
    es_unit = str(d.get("es_unit", "MPa")).strip().lower()
    es_gpa = es * {"pa": 1e-9, "kpa": 1e-6, "mpa": 1e-3, "gpa": 1.0}.get(es_unit, 1e-3)
    return root, to_gpa, es_gpa


def _read_raw_6x6(path):
    """The partner's headerless 6x6 csv -> symmetric 6x6 in its own units."""
    rows = []
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for r in csv.reader(fh):
            v = []
            for x in r:
                try:
                    v.append(float(x))
                except (TypeError, ValueError):
                    pass
            if len(v) >= 6:
                rows.append(v[:6])
    if len(rows) < 6:
        return None
    C = np.asarray(rows[:6], float)
    return 0.5 * (C + C.T)


def _index_tpms(root):
    """(topology, density rounded 2dp) -> csv path. Mirrors build_tpms_dataset.

    2026-08-10: the partner's QUADRATIC export (Zhezhe/week7/modulus) names its
    folders "D'", "C(I2-Y)" while the dataset's `topology` comes from his STL
    folders, which are "D' Generator". A literal folder-name key therefore
    matched NOTHING and every geometry came back "no tensor CSV", silently
    disabling rotation augmentation. Same two-pass alias as build_tpms_dataset:
    exact folder names win, aliases only fill gaps.
    """
    found = []
    for pth in glob.glob(os.path.join(root, "*", "*.csv")):
        topo = os.path.basename(os.path.dirname(pth))
        m = LAST_NUM.search(os.path.splitext(os.path.basename(pth))[0])
        if m:
            found.append((topo, round(float(m.group(1)), 2), pth))
    out = {}
    for topo, d, pth in found:
        out[(topo, d)] = pth
    for topo, d, pth in found:
        if not topo.endswith(" Generator"):
            out.setdefault((f"{topo} Generator", d), pth)
    return out


def main():
    results_dir, dataset_csv = _cfg()
    if not os.path.exists(dataset_csv):
        sys.exit(f"[E] dataset not found: {dataset_csv} — run labels_join.py first")
    rows = [r for r in csv.DictReader(open(dataset_csv)) if r.get("Run")]
    stems = [r["Run"] for r in rows]
    prof = run_paths.profile() or "ADMS"
    out = run_paths.data_path(_out_name(prof))
    es_adms = _es_gpa_adms()
    _tp_pre = _tpms_cfg()
    print(f"[i] profile={prof}   {len(stems)} geometries   -> {os.path.basename(out)}")
    print("[i] caching NORMALISED tensors (C/Es) so the two solids can share one file: "
          f"ADMS Es={es_adms:g} GPa"
          + (f",  TPMS Es={_tp_pre[2]:g} GPa" if _tp_pre else ""))

    have = {}
    if os.path.exists(out):
        z = np.load(out)
        have = {k: z[k] for k in z.files}
        print(f"[i] existing cache: {len(have)} tensors")

    added = miss = 0
    missing = []
    tp = _tpms_cfg()
    tp_index = _index_tpms(tp[0]) if tp and os.path.isdir(tp[0]) else {}
    n_adms = n_tpms = 0
    for r in rows:
        s = r["Run"]
        if s in have:
            continue
        fam = str(r.get("family", "")).strip().upper()
        C = None
        # our own FEA tree first
        hits = glob.glob(os.path.join(results_dir, "*", "Data", s, s + ".csv"))
        if hits and fam != "TPMS":
            C = read_c6_csv(hits[0])            # GPa
            if C is not None:
                C = C / es_adms                 # -> dimensionless
                n_adms += 1
        # otherwise the partner's raw 6x6, matched on (topology, density)
        if C is None and tp_index:
            topo = str(r.get("topology", "")).strip()
            m = LAST_NUM.search(s)
            dens = round(float(m.group(1)), 2) if m else None
            pth = tp_index.get((topo, dens))
            if pth:
                raw = _read_raw_6x6(pth)
                if raw is not None:
                    C = raw * tp[1] / tp[2]     # own units -> GPa -> /Es
                    n_tpms += 1
        if C is None:
            miss += 1; missing.append(s); continue
        have[s] = np.asarray(C, np.float32); added += 1

    np.savez_compressed(out, **have)
    covered = sum(1 for s in stems if s in have)
    print(f"[OK] wrote {out}")
    print(f"[i] tensors: {covered}/{len(stems)} geometries covered "
          f"(+{added} new: {n_adms} from our FEA tree, {n_tpms} from the partner's raw 6x6; "
          f"{miss} with no tensor)")
    for m in missing[:10]:
        print(f"    [!] no tensor CSV: {m}")
    if covered == 0:
        # Non-fatal: an ML-only user (no FEA results tree) simply cannot use rotation
        # augmentation. Do not break the pipeline — rotation_aug=1 will report clearly.
        print("[!] no per-run tensor CSVs found (no results tree?). Rotation augmentation "
              "will be unavailable; the rest of the pipeline is unaffected.")


if __name__ == "__main__":
    main()
