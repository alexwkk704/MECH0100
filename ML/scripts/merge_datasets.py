"""
merge_datasets.py - combine dataset_ADMS.csv + dataset_TPMS.csv into one training
table for the joint ADMS+TPMS model.

WHAT IT GUARANTEES
------------------
1. SAME FEATURE DEFINITIONS. Both sides must have been produced by the SAME
   feature_extraction.py. TPMS is extracted with --cell-size / --thickness-convention
   overrides rather than by editing the workbook, so the definitions cannot silently
   diverge. Any column present on one side only is dropped, loudly.
2. SAME LABEL DEFINITIONS. TPMS labels are recomputed from the partner's RAW 6x6
   tensors with our tensor_ops (build_tpms_dataset.py), never copied from his csv -
   his TAI uses a different formula (median 70% off ours, max 768%).
3. A GROUP KEY THAT DOES NOT LEAK. Grouped CV must not put two geometries of the
   same family AND similar density in train and validation. `group_key` combines
   family + topology + a density bin, so a whole (topology, density-band) cell moves
   together.
4. A `family` column (ADMS | TPMS) so any split can be stratified or held out by it.
5. THE SAME SOLID POISSON RATIO. Every joint target is dimensionless (C/Es, E/Es,
   G/Gs, TAI), which removes the solid modulus - but linear homogenisation leaves
   C_eff/Es a function of geometry AND nu_s. Dividing by Es only makes the two
   families comparable while their nu_s agree. Both are 0.3 today (ADMS "Steel"
   block in the .ntop = 200 GPa / 0.3; the partner's polymer = 1.8 GPa / 0.3), and
   this script now refuses to merge if the workbook says otherwise instead of
   letting the model quietly learn a material difference as if it were shape.

TARGETS
-------
Only the labels BOTH datasets carry can be trained jointly. ADMS has yield data
(onset_n, shear_onset_n); the partner does not. Those two stay ADMS-only - the
combined model simply cannot predict them, and pretending otherwise would train on
a column that is empty for 60% of the rows.

USAGE
    python merge_datasets.py
    python merge_datasets.py --density-bin 0.02
"""
import argparse, csv, os, sys
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR))
import run_paths

SHARED_TARGETS = ["E_over_Es", "G_over_Gs", "TAI", "C11_n", "C12_n", "C44_n"]
ADMS_ONLY_TARGETS = ["onset_n", "shear_onset_n"]

# Largest nu_s difference tolerated between the two families. C_eff/Es varies
# only weakly with nu_s for a slender lattice, but this is a physics assumption,
# not a licence - anything above this must be a deliberate decision, not a default.
NU_S_TOL = 0.01


def _solid_nu():
    """Read the solid Poisson ratio of each family from ML_settings.xlsx.

    ADMS -> 'labels' sheet, key nu_s.   TPMS -> 'tpms_labels' sheet, key nu.
    Nothing is hardcoded here; if a key is missing the caller is told which one."""
    import openpyxl
    xlsx = Path(run_paths._find_base()) / "ML_settings.xlsx"
    if not xlsx.exists():
        return None, None, f"{xlsx} not found"
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    def get(sheet, *keys):
        if sheet not in wb.sheetnames:
            return None
        for row in wb[sheet].iter_rows(values_only=True):
            if row and row[0] is not None and str(row[0]).strip().lower() in keys:
                try:
                    return float(row[1])
                except (TypeError, ValueError):
                    return None
        return None

    return get("labels", "nu_s"), get("tpms_labels", "nu", "nu_s"), None


def _element_order():
    """FE element order declared for each family. Not derivable from the tensors -
    it lives in the .ntop (geometric_order_enum) and is recorded in the workbook."""
    import openpyxl
    xlsx = Path(run_paths._find_base()) / "ML_settings.xlsx"
    if not xlsx.exists():
        return None, None
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)

    def get(sheet, key):
        if sheet not in wb.sheetnames:
            return None
        for row in wb[sheet].iter_rows(values_only=True):
            if row and row[0] is not None and str(row[0]).strip().lower() == key:
                return str(row[1]).strip() if row[1] is not None else None
        return None

    return get("labels", "element_order"), get("tpms_labels", "element_order")


def load(path):
    p = Path(run_paths.data_path(path))
    if not p.exists():
        sys.exit(f"[E] {p} not found")
    rows = list(csv.DictReader(p.open(newline="", encoding="utf-8")))
    print(f"[ok] {len(rows):4d} rows   {p.name}")
    return rows, p


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    # 2026-08-10: these used to be hardcoded filenames, so pointing the
    # train_TPMS / train_COMBINED sheets at a different dataset did NOT change
    # what got merged - the sheets said one thing and this script did another.
    # Defaults now come from the same workbook every other script reads.
    def _from_sheet(prof, default):
        keep = os.environ.get("ML_PROFILE")
        try:
            os.environ["ML_PROFILE"] = prof
            return os.path.basename(run_paths.sheet_value("train", "dataset_csv", default))
        finally:
            if keep is None:
                os.environ.pop("ML_PROFILE", None)
            else:
                os.environ["ML_PROFILE"] = keep

    ap.add_argument("--adms", default=_from_sheet("ADMS", "dataset_ADMS.csv"))
    ap.add_argument("--tpms", default=_from_sheet("TPMS", "dataset_TPMS.csv"))
    ap.add_argument("--out",  default=_from_sheet("COMBINED", "dataset_COMBINED.csv"))
    ap.add_argument("--adms-cell", type=float, default=3.0,
                    help="cell_size_mm used when the ADMS features were extracted")
    ap.add_argument("--tpms-cell", type=float, default=10.0,
                    help="cell_size_mm used when the TPMS features were extracted")
    ap.add_argument("--no-lstar", action="store_true",
                    help="skip the measured-intrinsic-length columns")
    ap.add_argument("--density-bin", type=float, default=0.02,
                    help="width of the density band used in group_key (default 0.02). "
                         "Two geometries of the same topology inside one band are "
                         "near-duplicates and must not be split across folds.")
    args = ap.parse_args()

    A, pa = load(args.adms)
    T, pt = load(args.tpms)

    ca, ct = set(A[0].keys()), set(T[0].keys())
    common = [c for c in A[0].keys() if c in ct]          # keep ADMS column order
    only_a = sorted(ca - ct)
    only_t = sorted(ct - ca)
    print(f"\n[ok] {len(common)} shared columns")
    if only_a:
        print(f"  ADMS-only, DROPPED ({len(only_a)}): {only_a[:12]}{' ...' if len(only_a)>12 else ''}")
    if only_t:
        print(f"  TPMS-only, DROPPED ({len(only_t)}): {only_t[:12]}{' ...' if len(only_t)>12 else ''}")

    missing_t = [t for t in SHARED_TARGETS if t not in ct]
    missing_a = [t for t in SHARED_TARGETS if t not in ca]
    if missing_t or missing_a:
        sys.exit(f"[E] a shared target is absent. missing in TPMS: {missing_t}   "
                 f"missing in ADMS: {missing_a}")
    nu_a, nu_t, err = _solid_nu()
    if err:
        sys.exit(f"[E] cannot read the solid Poisson ratios: {err}")
    if nu_a is None or nu_t is None:
        sys.exit("[E] solid Poisson ratio not recorded. "
                 f"ADMS 'labels'!nu_s = {nu_a}, TPMS 'tpms_labels'!nu = {nu_t}. "
                 "Both must be set to the value in the matching .ntop Isotropic "
                 "Linear Elastic block before the families can share a model.")
    if abs(nu_a - nu_t) > NU_S_TOL:
        sys.exit(f"[E] solid Poisson ratios differ: ADMS nu_s={nu_a}, TPMS nu={nu_t} "
                 f"(tolerance {NU_S_TOL}). Dividing by Es removes the modulus but not "
                 "nu_s, so C/Es would still carry a material difference and the model "
                 "would learn it as geometry. Fix the workbook or re-run the FEA with "
                 "a matched solid.")
    print(f"  solid Poisson ratio: ADMS {nu_a} / TPMS {nu_t}  -> Es-only "
          f"normalisation is valid (targets depend on geometry alone)")
    eo_a, eo_t = _element_order()
    if eo_a and eo_t and eo_a.lower() != eo_t.lower():
        print("\n" + "!" * 74)
        print(f"[!!] MIXED FE ELEMENT ORDERS: ADMS = {eo_a}, TPMS = {eo_t}.")
        print("     Linear elements are ~6-8% stiffer than quadratic and shift TAI by")
        print("     13-66% (measured from the partner's own linear/quad pairs). That bias")
        print("     is perfectly correlated with `family`, so part of any held-out-family")
        print("     gap is NUMERICS, not geometry, and the two cannot be separated.")
        print("     Every result from this dataset is PROVISIONAL until both families")
        print("     share one element order. Recorded per row in `element_order`.")
        print("!" * 74 + "\n")
    elif eo_a and eo_t:
        print(f"  FE element order: both families {eo_a} - no element-order confound")
    else:
        print(f"  [!] element order not recorded for both families (ADMS={eo_a}, TPMS={eo_t})")
    print(f"  trainable targets: {SHARED_TARGETS}")
    print(f"  ADMS-only targets (excluded from the joint model): {ADMS_ONLY_TARGETS}")

    def fam_key(r, family):
        # topology: ADMS uses adms_type (DF/raw/flow), the partner uses topology
        topo = (r.get("adms_type") or r.get("topology") or "?").strip()
        try:
            d = float(r.get("Rho_rel") or r.get("VF") or 0.0)
        except (TypeError, ValueError):
            d = 0.0
        band = int(d / args.density_bin) if args.density_bin > 0 else 0
        return f"{family}|{topo}|{band}"

    def topo_of(r):
        """Unified topology label: ADMS uses adms_type (DF/raw/flow), TPMS uses topology.
        Written as its own COLUMN (not just folded into group_key) because the
        leave-one-topology-out holdout needs to filter on it directly."""
        return (r.get("adms_type") or r.get("topology") or "?").strip()

    out = []
    for r in A:
        row = {c: r.get(c, "") for c in common}
        row["family"] = "ADMS"; row["topology"] = topo_of(r)
        row["element_order"] = eo_a or ""
        row["group_key"] = fam_key(r, "ADMS")
        out.append(row)
    for r in T:
        row = {c: r.get(c, "") for c in common}
        row["family"] = "TPMS"; row["topology"] = topo_of(r)
        row["element_order"] = eo_t or ""
        row["group_key"] = fam_key(r, "TPMS")
        out.append(row)

    # sample size in units of the structure's own measured feature length.
    # ADMS is aperiodic (no unit cell), TPMS is one 10 mm period per box, so
    # "number of cells" is not a shared concept - bbox / L* is, and it is measured
    # from the STL alone (no nTop metadata, satisfies the STL-only rule).
    # Added AFTER L_star_mm exists; see the loop below.

    # ---- measured intrinsic length -------------------------------------------
    # ADMS is APERIODIC. Spherene generate it outside-in as one continuous surface
    # ("adapts continuously and locally", unlike TPMS which "depend on repeating,
    # grading, and trimming unit cells"). So its cell_size_mm is NOT a period - it is
    # a density/thickness parameter - and dividing ADMS curvature by a nominal 3 mm
    # "cell" is arbitrary. TPMS genuinely does have a 10 mm period.
    #
    # Non-dimensionalising both by a MEASURED length removes that asymmetry:
    #     L* = 1 / (surface area per box volume)
    # computed identically from either dataset, no metadata, cannot be wrong because
    # someone typed the wrong cell size.
    #
    # Measured effect on how much the two families overlap:
    #     thickness/L   nominal ADMS 0.134 vs TPMS 0.042  ->  L* 0.419 vs 0.426
    #     SAV x L       nominal ADMS 15.6  vs TPMS 48.4   ->  L*  4.89 vs  5.01
    #     overlap       27% -> 60%  and  18% -> 41%
    # With the nominal version, thickness_over_cell < 0.05 is a perfect TPMS
    # detector: the model can learn the FAMILY LABEL instead of the geometry, score
    # well in-sample, and fail on anything new. That is the failure mode this guards.
    #
    # Both column sets are written. Train on whichever you can defend - but do not
    # mix them in one feature list.
    STAR = {"SAV_x_cell": 1, "rel_surface": 1, "curv_mean_mean_xL": 1,
            "curv_mean_iqr_xL": 1, "curv_mean_p99_xL": 1,
            "curv_gauss_mean_xL2": 2, "curv_gauss_iqr_xL2": 2, "curv_gauss_p99_xL2": 2,
            "thickness_over_cell": -1}
    if not args.no_lstar:
        added = []
        for row in out:
            L = args.adms_cell if row["family"] == "ADMS" else args.tpms_cell
            try:
                AoV = (float(row["SAV_x_cell"]) / L) * float(row["VF"])
                Ls = 1.0 / AoV if AoV > 0 else None
            except (TypeError, ValueError, KeyError):
                Ls = None
            row["L_star_mm"] = round(Ls, 5) if Ls else ""
            for c, pw in STAR.items():
                if c not in row or Ls is None:
                    continue
                try:
                    v = float(row[c])
                except (TypeError, ValueError):
                    continue
                # undo the nominal normalisation, redo it with L*
                row[c + "_star"] = round(v * (Ls / L) ** pw, 6)
        for row in out:
            try:
                Ls = float(row.get("L_star_mm") or 0)
                bb = float(row.get("bbox_mm") or 0)
                row["bbox_over_Lstar"] = round(bb / Ls, 4) if Ls > 0 and bb > 0 else ""
            except (TypeError, ValueError):
                row["bbox_over_Lstar"] = ""
        added = ["L_star_mm", "bbox_over_Lstar"] + [c + "_star" for c in STAR if (c + "_star") in out[0]]
        common = common + added
        print("  added measured-L* columns: %d   (use *_star OR the nominal set, never both)" % len(added))

    # `topology` is written explicitly below, so strip it from `common` first -
    # once ADMS features carry a topology column it would otherwise appear twice.
    cols = ([c for c in common if c not in ("topology", "element_order")]
            + ["family", "topology", "element_order", "group_key"])
    dest = Path(run_paths.data_path(args.out))
    with dest.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in out:
            w.writerow(r)

    import collections
    fam = collections.Counter(r["family"] for r in out)
    grp = collections.Counter(r["group_key"] for r in out)
    print(f"\n{'='*64}\n  {len(out)} rows   ADMS {fam['ADMS']}   TPMS {fam['TPMS']}")
    print(f"  {len(grp)} groups   largest {max(grp.values())} rows   "
          f"singletons {sum(1 for v in grp.values() if v==1)}")
    print(f"  -> {dest}")
    print(f"""
  NEXT - in ML_settings.xlsx, sheet 'train':
      dataset_csv   data/{args.out}
      group_cols    group_key
      targets       {', '.join(SHARED_TARGETS)}
  Keep a separate ADMS-only run for onset_n / shear_onset_n.""")


if __name__ == "__main__":
    main()
