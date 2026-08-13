"""
train.py — generic forward-model trainer: geometric features -> normalised properties.

ALL settings live in ML_settings.xlsx, sheet "train" (Excel-editable, no code changes):
  dataset_csv / results_csv / features / targets / vf_column / n_folds /
  rf_trees / random_seed / min_rows_per_target / holdout_1..N / parity_targets

Works for any dataset (ADMS, TPMS, ...) whose dataset CSV contains the listed
feature + target columns. Missing targets are skipped, not fatal.

Models per target:
  reference_density : NOT a model. A density-only power law y = a*VF^n, included
                      purely as a CONTROL. The ML models are RF, GPR and PointNet;
                      this line exists so a claim like "the geometry features
                      contribute" can be tested rather than assumed. Report it as a
                      reference line, never as a competing method.
  RF           : RandomForest(rf_trees, random_seed)
  GPR          : Gaussian Process (Matern 2.5 + white noise, standardised X)

Validation:
  k-fold CV (out-of-fold predictions) + each holdout_i = rows where column==value
  held out entirely (unseen-config test). Metrics: R2 + MAPE.

Requires: pip install scikit-learn pandas numpy openpyxl
Usage:    python train.py          (reads ML_settings.xlsx next to this file)
"""

import os, sys, warnings
import numpy as np
import pandas as pd
warnings.filterwarnings("ignore")

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)          # so `import cv_split` works when run from ML/
import cv_split                   # shared, deterministic grouped CV folds
import run_paths                  # clean filing: data/ inputs, runs/<stamp>/ results

def _find_settings(base, name="ML_settings.xlsx"):
    """Locate the settings workbook.

    Scripts live in ML/scripts/ but the workbook sits one level up in ML/ so
    the user sees it immediately. Checking the parent too means the layout can
    change (or a user can keep the workbook beside the scripts) without edits.
    """
    import os as _os
    here = _os.path.join(base, name)
    if _os.path.exists(here):
        return here
    up = _os.path.join(_os.path.dirname(base), name)
    return up if _os.path.exists(up) else here

SETTINGS_XLSX = _find_settings(HERE)

# Relative paths in the settings workbook resolve against the folder that HOLDS
# the workbook, NOT against this file. That way the scripts can be moved into a
# subfolder (ML/scripts/) without touching a single setting.
BASE_DIR = os.path.dirname(SETTINGS_XLSX)


# Fewest training rows a holdout may leave behind before it is skipped.
MIN_TRAIN_ROWS = 10

from sklearn.ensemble import RandomForestRegressor
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import Matern, WhiteKernel, ConstantKernel
from sklearn.preprocessing import StandardScaler


def load_cfg(sheet="train"):
    # profile-aware: ML_PROFILE=COMBINED reads sheet `train_COMBINED` if present
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX, data_only=True)
    sheet = run_paths.sheet_for(wb, sheet)
    if sheet not in wb.sheetnames:
        sys.exit(f"[E] sheet '{sheet}' missing from {SETTINGS_XLSX}")
    cfg = {}
    for k, v, *_ in wb[sheet].iter_rows(min_row=2, values_only=True):
        if k:
            cfg[str(k).strip()] = v
    cfg["features"] = [s.strip() for s in str(cfg["features"]).split(",") if s.strip()]
    cfg["targets"] = [s.strip() for s in str(cfg["targets"]).split(",") if s.strip()]
    cfg["parity_targets"] = [s.strip() for s in str(cfg.get("parity_targets", "")).split(",") if s.strip()]
    cfg["group_cols"] = [s.strip() for s in str(cfg.get("group_cols", "")).split(",") if s.strip()]
    cfg["holdouts"] = []
    for k in sorted(cfg):
        if k.startswith("holdout_") and cfg[k]:
            try:
                name, col, val = [p.strip() for p in str(cfg[k]).split("|")]
                cfg["holdouts"].append((name, col, val))
            except ValueError:
                print(f"[!] bad holdout spec ignored: {cfg[k]} (need name|column|value)")
    for k in ("n_folds", "rf_trees", "random_seed", "min_rows_per_target"):
        cfg[k] = int(cfg[k])
    # clean filing: dataset is a shared derived INPUT (ML/data/); results + charts
    # are per-run OUTPUTS (ML/runs/<stamp>/).
    cfg["dataset_csv"] = run_paths.data_path(cfg["dataset_csv"])
    cfg["results_csv"] = run_paths.run_path(os.path.basename(str(cfg["results_csv"])))
    cfg["charts_dir"] = run_paths.run_path("charts")
    return cfg


def mape(y, p):
    y, p = np.asarray(y, float), np.asarray(p, float)
    m = np.abs(y) > 1e-9
    return float(np.mean(np.abs((p[m] - y[m]) / y[m])) * 100)


def r2(y, p):
    y, p = np.asarray(y, float), np.asarray(p, float)
    ss = np.sum((y - y.mean()) ** 2)
    return float(1 - np.sum((y - p) ** 2) / ss) if ss > 0 else float("nan")


def make_rf(cfg):
    return RandomForestRegressor(n_estimators=cfg["rf_trees"],
                                 random_state=cfg["random_seed"], n_jobs=-1)


def make_gpr(cfg):
    k = ConstantKernel(1.0) * Matern(length_scale=1.0, nu=2.5) + WhiteKernel(1e-4)
    return GaussianProcessRegressor(kernel=k, normalize_y=True,
                                    random_state=cfg["random_seed"],
                                    n_restarts_optimizer=3)


def baseline_fit(vf_tr, y_tr):
    m = np.asarray(y_tr) > 0
    if m.sum() < 3:
        raise ValueError("too few positive labels for power-law fit")
    n, c = np.polyfit(np.log(np.asarray(vf_tr)[m]), np.log(np.asarray(y_tr)[m]), 1)
    return lambda vf: np.exp(c) * np.asarray(vf) ** n


def holdout_mask(d, col, val):
    """Rows a holdout hides. `val` is an exact match by default, but may also be a
    RANGE or COMPARISON so a holdout can test EXTRAPOLATION rather than just an
    unseen category:

        unseen_gyroid   | topology | Gyroid Generator   exact  (category)
        high_density    | VF       | >=0.25             comparison
        low_density     | VF       | <0.12              comparison
        mid_band        | VF       | 0.15..0.20         range, inclusive

    Comparison/range forms need a numeric column; a non-numeric one is reported
    rather than silently matching nothing."""
    if col not in d.columns:
        return None
    v = str(val).strip()
    num = pd.to_numeric(d[col], errors="coerce")
    ops = (">=", "<=", ">", "<")
    op = next((o for o in ops if v.startswith(o)), None)
    if op or ".." in v:
        if num.notna().sum() == 0:
            print(f"  [!] holdout on '{col}' uses '{v}' but the column is not numeric")
            return None
        if op:
            try:
                x = float(v[len(op):])
            except ValueError:
                print(f"  [!] bad holdout bound: {v}")
                return None
            return {">=": num >= x, "<=": num <= x, ">": num > x, "<": num < x}[op].to_numpy()
        try:
            lo, hi = (float(t) for t in v.split("..", 1))
        except ValueError:
            print(f"  [!] bad holdout range: {v}")
            return None
        return ((num >= lo) & (num <= hi)).to_numpy()
    try:
        return np.isclose(num, float(v))
    except (TypeError, ValueError):
        return d[col].astype(str).str.strip() == v


def main():
    cfg = load_cfg()
    df = pd.read_csv(cfg["dataset_csv"])
    print(f"[i] dataset: {cfg['dataset_csv']}  ({len(df)} rows)")
    missing_f = [f for f in cfg["features"] if f not in df.columns]
    if missing_f:
        sys.exit(f"[E] feature columns missing from dataset: {missing_f}")

    # ---- assign CV folds ONCE, on the whole dataset, by group ----
    # Every model (RF/GPR here, PointNet, compare_models) reads the same fold map,
    # so a geometry is always in the same fold everywhere and near-duplicates
    # (same group) can never straddle train/test. See cv_split.py.
    fold_ids, grouped, _stem, _fmap = cv_split.assign(
        df, cfg["group_cols"], cfg["n_folds"], seed=cfg["random_seed"])
    df["_fold"] = fold_ids
    split_name = f"grouped{cfg['n_folds']}foldCV" if grouped else f"{cfg['n_folds']}foldCV"
    if grouped:
        from collections import Counter
        print(f"[i] CV = {split_name} on {cfg['group_cols']}  "
              f"({len(set(fold_ids))} folds, rows/fold={dict(sorted(Counter(fold_ids).items()))})")
    else:
        print(f"[i] CV = {split_name}  (group_cols {cfg['group_cols'] or 'unset'} "
              f"absent -> deterministic per-row split; near-duplicates NOT protected)")

    results = []
    # Per-row out-of-fold CV predictions. Aggregate R2 over the whole
    # dataset hides where a model is strong or weak; keeping the rows lets
    # band_compare.py score any density band, which is how the effect of
    # targeted new simulations is actually measured.
    oof_rows = []
    for tgt in cfg["targets"]:
        if tgt not in df.columns:
            print(f"[!] target '{tgt}' not in dataset — skipped")
            continue
        d = df.dropna(subset=[tgt] + cfg["features"]).reset_index(drop=True)
        if len(d) < cfg["min_rows_per_target"]:
            print(f"[!] {tgt}: only {len(d)} labelled rows — skipped")
            continue
        X, y, vf = d[cfg["features"]].values, d[tgt].values, d[cfg["vf_column"]].values
        print(f"\n=== {tgt}  (N={len(d)}) ===")

        # ---- grouped out-of-fold CV predictions ----
        # Fold membership is the pre-assigned df["_fold"], carried through the
        # per-target dropna. Iterate folds manually (instead of a KFold splitter)
        # so RF/GPR use EXACTLY the same folds as PointNet, and so a target with
        # sparse labels simply skips any fold it has no test rows in.
        fold = d["_fold"].values
        preds = {"RF": np.full(len(y), np.nan), "GPR": np.full(len(y), np.nan),
                 "reference_density": np.full(len(y), np.nan)}
        for f in range(cfg["n_folds"]):
            te = np.where(fold == f)[0]
            tr = np.where(fold != f)[0]
            if len(te) == 0 or len(tr) < 3:
                continue
            rf = make_rf(cfg); rf.fit(X[tr], y[tr]); preds["RF"][te] = rf.predict(X[te])
            sc = StandardScaler().fit(X[tr])
            gp = make_gpr(cfg); gp.fit(sc.transform(X[tr]), y[tr])
            preds["GPR"][te] = gp.predict(sc.transform(X[te]))
            try:
                preds["reference_density"][te] = baseline_fit(vf[tr], y[tr])(vf[te])
            except ValueError:
                pass
        for mname in ("RF", "GPR", "reference_density"):
            p = preds[mname]
            scored = ~np.isnan(p)               # rows this model actually predicted
            if scored.sum() < 3:
                continue
            row = dict(target=tgt, split=split_name, model=mname,
                       N_test=int(scored.sum()),
                       R2=round(r2(y[scored], p[scored]), 4),
                       MAPE=round(mape(y[scored], p[scored]), 2))
            results.append(row); print(f"  {row}")
            runs = d["Run"].values if "Run" in d.columns else np.arange(len(d))
            for i in np.where(scored)[0]:
                oof_rows.append(dict(Run=runs[i], target=tgt, model=mname,
                                     VF=float(vf[i]), y_true=float(y[i]),
                                     y_pred=float(p[i])))

        # ---- holdouts ----
        for hname, col, val in cfg["holdouts"]:
            mask = holdout_mask(d, col, val)
            if mask is None:
                print(f"  [!] holdout '{hname}': column '{col}' missing — skipped"); continue
            te, tr = np.where(mask)[0], np.where(~mask)[0]
            if len(te) == 0:
                print(f"  [!] holdout '{hname}': no rows match {col}=={val} — skipped"); continue
            # A holdout that swallows (nearly) the whole dataset leaves nothing
            # to train on. Happens e.g. with unseen_variant_df on a DF-only
            # dataset. Skip rather than crash inside the regressor.
            if len(tr) < MIN_TRAIN_ROWS:
                print(f"  [!] holdout '{hname}': only {len(tr)} training row(s) left "
                      f"after holding out {col}=={val} (need {MIN_TRAIN_ROWS}) — skipped")
                continue
            rf = make_rf(cfg); rf.fit(X[tr], y[tr])
            sc = StandardScaler().fit(X[tr])
            gp = make_gpr(cfg); gp.fit(sc.transform(X[tr]), y[tr])
            model_preds = {"RF": rf.predict(X[te]),
                           "GPR": gp.predict(sc.transform(X[te]))}
            try:
                model_preds["reference_density"] = baseline_fit(vf[tr], y[tr])(vf[te])
            except ValueError:
                pass
            for mname, p in model_preds.items():
                row = dict(target=tgt, split=hname, model=mname, N_test=len(te),
                           R2=round(r2(y[te], p), 4), MAPE=round(mape(y[te], p), 2))
                results.append(row); print(f"  {row}")

        # ---- feature importances ----
        m = make_rf(cfg).fit(X, y)
        imp = sorted(zip(cfg["features"], m.feature_importances_), key=lambda t: -t[1])[:4]
        print("  top features:", ", ".join(f"{k}={v:.2f}" for k, v in imp))

    if not results:
        # Every target was skipped (too few labelled rows, or none of the
        # configured target columns exist in the dataset). Fail with a useful
        # message instead of a KeyError on an empty DataFrame.
        print("\n[!] NO RESULTS PRODUCED - nothing was written.")
        print("    Every target was skipped. Usual causes:")
        print(f"      - dataset has too few labelled rows "
              f"(min_rows_per_target = {cfg['min_rows_per_target']})")
        print(f"      - none of the target columns exist in {cfg['dataset_csv']}")
        print(f"      - targets configured: {cfg['targets']}")
        sys.exit(1)

    pd.DataFrame(results)[["target", "split", "model", "N_test", "R2", "MAPE"]] \
        .to_csv(cfg["results_csv"], index=False)
    print(f"\n[OK] wrote {cfg['results_csv']} ({len(results)} rows)")

    if oof_rows:
        oof_fp = run_paths.run_path("cv_oof_predictions.csv")
        pd.DataFrame(oof_rows)[["Run", "target", "model", "VF", "y_true", "y_pred"]] \
            .to_csv(oof_fp, index=False)
        print(f"[OK] wrote {oof_fp} ({len(oof_rows)} rows) "
              f"- per-row out-of-fold predictions for band_compare.py")


if __name__ == "__main__":
    main()
