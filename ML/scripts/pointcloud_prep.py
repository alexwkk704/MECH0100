"""
pointcloud_prep.py — Farooq pipeline step 2: STL -> normalised point cloud.

Per spec (Teams 17/07):
  * uniform surface sampling with trimesh (N_POINTS coordinates per STL)
  * centre every cloud at the origin (0,0,0)
  * scale to fit a 1x1x1 bounding box (network learns shape, not scale)

Settings: ML_settings.xlsx sheet "pointcloud" (created on first run if missing):
  stl_folder / out_folder / n_points / random_seed
Output: one <stem>.npz per STL (array 'pts', float32, shape (n_points, 3))
        + _pc_validation.csv (centre ≈ 0, max-extent ≈ 1 checks)
Resume-safe: an existing .npz is skipped ONLY if it matches BOTH the current
n_points AND the current random_seed. Change either and the stale clouds are
regenerated automatically.

⚠ 2026-08-12 - THE BUG THIS FIXES, and why it went unnoticed for three weeks.
The staleness check used to test n_points ONLY. `pointcloud!random_seed` was
changed 42 -> 1 between 23 and 31 July; every cloud that already existed kept its
seed-42 points while everything generated afterwards used seed 1, so the training
set became a silent mixture of two samplings. Nothing errored. It surfaced only
when GATE 0 replayed rows from their original STL and two of ten came back 39-48 %
different. Clouds written before this fix carry no seed record, so they are
treated as stale and regenerated once - that is intended, not a bug.

Requires: pip install trimesh numpy openpyxl
Usage:    python pointcloud_prep.py           (all files)
          python pointcloud_prep.py --test               (first 3 files only)
          python pointcloud_prep.py --stl-folder <dir>   (override the STL folder)
"""

import os, sys, csv, time
import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))

def _find_settings(base, name="ML_settings.xlsx"):
    """Locate the settings workbook.

    Scripts live in ML/scripts/ but the workbook sits one level up in ML/ so
    the user sees it immediately. Checking the parent too means the layout can
    change (or a user can keep the workbook beside the scripts) without edits.
    """
    import os as _os
    here = _os.path.join(base, name)
    if _os.path.exists(here):
        return here
    up = _os.path.join(_os.path.dirname(base), name)
    return up if _os.path.exists(up) else here

SETTINGS_XLSX = _find_settings(HERE)

# Relative paths in the settings workbook resolve against the folder that HOLDS
# the workbook, NOT against this file. That way the scripts can be moved into a
# subfolder (ML/scripts/) without touching a single setting.
BASE_DIR = os.path.dirname(SETTINGS_XLSX)
import sys; sys.path.insert(0, HERE); import run_paths   # clean filing: data/ + runs/


DEFAULTS = [
    ("stl_folder", "../ADMS/ADMS_STL",
     "Folder containing the .stl files (relative to ML_settings.xlsx folder, or absolute)"),
    ("out_folder", os.path.join(HERE, "pointclouds"),
     "Folder for <stem>.npz point clouds"),
    ("n_points", 2048, "Uniform surface samples per STL (Farooq spec ~2048)"),
    ("random_seed", 42, "Do NOT change mid-dataset (reproducibility)"),
]


def load_cfg():
    import openpyxl
    wb = openpyxl.load_workbook(SETTINGS_XLSX)
    if "pointcloud" not in wb.sheetnames:
        ws = wb.create_sheet("pointcloud")
        ws.append(("setting", "value", "explanation"))
        for r in DEFAULTS:
            ws.append(r)
        wb.save(SETTINGS_XLSX)
        print("[i] created 'pointcloud' sheet in ML_settings.xlsx with defaults")
    cfg = {}
    for k, v, *_ in wb[run_paths.sheet_for(wb, "pointcloud")].iter_rows(min_row=2, values_only=True):
        if k:
            cfg[str(k).strip()] = v
    # the STL folder for this profile lives in settings_<PROFILE>, so the clouds
    # always come from the same STLs the features did. No path in any .bat.
    _ssheet = run_paths.sheet_for(wb, "settings")
    if _ssheet != "settings" and _ssheet in wb.sheetnames:
        for k, v, *_ in wb[_ssheet].iter_rows(min_row=2, values_only=True):
            if k and str(k).strip() == "stl_folder" and v:
                cfg["stl_folder"] = v
    cfg["n_points"] = int(cfg["n_points"])
    cfg["random_seed"] = int(cfg["random_seed"])
    # CLI override so one settings file can serve several STL folders
    # (ADMS lives in one place, the partner's TPMS in another). Same pattern as
    # feature_extraction.py --stl-folder. Nothing is hardcoded here.
    if "--stl-folder" in sys.argv:
        i = sys.argv.index("--stl-folder")
        if len(sys.argv) > i + 1:
            cfg["stl_folder"] = sys.argv[i + 1]
            print(f"[override] stl_folder={cfg['stl_folder']}")
    p = str(cfg["stl_folder"])
    cfg["stl_folder"] = os.path.normpath(p if os.path.isabs(p) else os.path.join(BASE_DIR, p))
    # clouds are a DERIVED INPUT (heavy, cached) -> ML/data/pointclouds
    cfg["out_folder"] = run_paths.pointclouds_dir()
    return cfg


def process(stl_path, n_points, seed):
    import trimesh
    mesh = trimesh.load(stl_path, force="mesh")
    pts, _fid = trimesh.sample.sample_surface(mesh, n_points, seed=seed)
    pts = np.asarray(pts, dtype=np.float64)
    pts -= pts.mean(axis=0)                      # centre at origin
    extent = np.ptp(pts, axis=0).max()           # largest bbox side
    if extent > 0:
        pts /= extent                            # fit in 1x1x1 box
    return pts.astype(np.float32)


def main():
    cfg = load_cfg()
    test = "--test" in sys.argv
    # --audit: report what WOULD be regenerated and change nothing.
    audit = "--audit" in sys.argv
    # --probe N : for up to N clouds whose seed is UNRECORDED, re-sample the STL
    # at each candidate seed and report which one reproduces the stored points.
    # This is what turns "I think the seed changed" into proof, and it is bounded
    # to N files because an ADMS STL is 100-500 MB and loading one is the slow part.
    probe_n = 0
    if "--probe" in sys.argv:
        audit = True
        i = sys.argv.index("--probe")
        probe_n = int(sys.argv[i + 1]) if len(sys.argv) > i + 1 and sys.argv[i + 1].isdigit() else 3
    stale, changed, identical = [], [], []
    os.makedirs(cfg["out_folder"], exist_ok=True)
    # Recursive, to match feature_extraction.py: ADMS STLs are flat, the partner's
    # TPMS STLs sit one folder per topology. Output .npz stays FLAT (basename stem),
    # which is safe because every basename is unique across the tree.
    stls = []
    for _root, _dirs, _fs in os.walk(cfg["stl_folder"]):
        for _f in _fs:
            if _f.lower().endswith(".stl"):
                stls.append(os.path.relpath(os.path.join(_root, _f), cfg["stl_folder"]))
    stls = sorted(stls)
    _stems = [os.path.splitext(os.path.basename(x))[0] for x in stls]
    if len(set(_stems)) != len(_stems):
        sys.exit("[E] duplicate STL basenames across subfolders - flat .npz output would "
                 "overwrite. Rename the colliding files before continuing.")
    if test:
        stls = stls[:3]
    val_path = os.path.join(cfg["out_folder"], "_pc_validation.csv")
    new_file = not os.path.exists(val_path)
    done = skipped = 0
    with open(val_path, "a", newline="") as vf:
        w = csv.writer(vf)
        if new_file:
            w.writerow(["file", "n_points", "centre_max_abs", "max_extent",
                        "seconds", "status"])
        for i, f in enumerate(stls, 1):
            stem = os.path.splitext(os.path.basename(f))[0]
            out = os.path.join(cfg["out_folder"], stem + ".npz")
            if os.path.exists(out):
                # Skip ONLY if the stored cloud matches the current n_points AND
                # the current seed. A cloud with no recorded seed predates the
                # 12/08 fix and is regenerated once.
                existing_n, existing_seed = -1, None
                try:
                    z = np.load(out)
                    existing_n = int(z["pts"].shape[0])
                    if "seed" in z.files:
                        existing_seed = int(z["seed"])
                except Exception:
                    pass
                if existing_n == cfg["n_points"] and existing_seed == cfg["random_seed"]:
                    skipped += 1
                    continue
                why = []
                if existing_n != cfg["n_points"]:
                    why.append(f"n_points {existing_n} -> {cfg['n_points']}")
                if existing_seed != cfg["random_seed"]:
                    why.append("seed "
                               + (f"{existing_seed}" if existing_seed is not None else "UNRECORDED")
                               + f" -> {cfg['random_seed']}")
                stale.append((os.path.basename(f), "; ".join(why)))
                if audit:
                    continue
                print(f"[{i}/{len(stls)}] {f}: regenerating ({'; '.join(why)})")
            t0 = time.time()
            try:
                # Keep the old points so we can say whether the rebuild actually
                # CHANGED anything. A cloud flagged only because its seed was
                # never recorded may well rebuild bit-identical, and then nothing
                # downstream needs redoing. That distinction is the whole point.
                old_pts = None
                if os.path.exists(out):
                    try:
                        old_pts = np.load(out)["pts"]
                    except Exception:
                        pass
                pts = process(os.path.join(cfg["stl_folder"], f),
                              cfg["n_points"], cfg["random_seed"])
                if old_pts is not None and old_pts.shape == pts.shape:
                    if float(np.abs(old_pts - pts).max()) < 1e-6:
                        identical.append(os.path.basename(f))
                    else:
                        changed.append(os.path.basename(f))
                elif old_pts is not None:
                    changed.append(os.path.basename(f))
                # Record the parameters INSIDE the file. A cloud that cannot say
                # how it was made cannot be checked, which is exactly how the
                # seed change went unnoticed.
                np.savez_compressed(out, pts=pts,
                                    n_points=np.int32(cfg["n_points"]),
                                    seed=np.int32(cfg["random_seed"]))
                c = float(np.abs(pts.mean(axis=0)).max())
                e = float(np.ptp(pts, axis=0).max())
                ok = "OK" if (c < 1e-3 and 0.999 < e <= 1.001) else "CHECK"
                w.writerow([os.path.basename(f), len(pts), round(c, 6), round(e, 6),
                            round(time.time() - t0, 1), ok])
                vf.flush()
                done += 1
                print(f"[{i}/{len(stls)}] {f}: {ok} ({time.time()-t0:.1f}s)")
            except Exception as ex:
                w.writerow([os.path.basename(f), "", "", "", round(time.time() - t0, 1), f"FAIL {ex}"])
                vf.flush()
                print(f"[{i}/{len(stls)}] {f}: FAIL {ex}")
    if audit and probe_n:
        cands = []
        for c in (cfg["random_seed"], 42, 1, 0):
            if c not in cands:
                cands.append(c)
        unrec = [(n, w) for n, w in stale if "UNRECORDED" in w][:probe_n]
        print(f"\n[PROBE] re-sampling {len(unrec)} cloud(s) at seeds {cands} to find "
              f"which one reproduces the stored points")
        for name, _ in unrec:
            rel = next((x for x in stls if os.path.basename(x) == name), None)
            if rel is None:
                continue
            stem = os.path.splitext(name)[0]
            try:
                stored = np.load(os.path.join(cfg["out_folder"], stem + ".npz"))["pts"]
            except Exception as ex:
                print(f"    {name}: cannot read stored cloud ({ex})"); continue
            print(f"    {name}")
            hit = None
            for c in cands:
                try:
                    fresh = process(os.path.join(cfg["stl_folder"], rel), cfg["n_points"], c)
                except Exception as ex:
                    print(f"        seed {c:<4} FAILED to sample: {ex}"); continue
                if fresh.shape != stored.shape:
                    print(f"        seed {c:<4} shape {fresh.shape} != stored {stored.shape}"); continue
                d = float(np.abs(fresh - stored).max())
                mark = "  <-- MATCH" if d < 1e-6 else ""
                print(f"        seed {c:<4} max coord diff {d:.3e}{mark}")
                if d < 1e-6 and hit is None:
                    hit = c
            print(f"        => stored cloud was built with seed "
                  + (f"{hit}" if hit is not None else "NONE OF THESE"))

    if audit:
        print(f"\n[AUDIT] nothing was written. seed={cfg['random_seed']} "
              f"n_points={cfg['n_points']}")
        print(f"[AUDIT] up to date : {skipped}")
        print(f"[AUDIT] STALE      : {len(stale)}")
        for name, why in stale:
            print(f"           {name:<48} {why}")
        if stale:
            print("\n[AUDIT] re-run WITHOUT --audit (or run the SETUP bat) to rebuild these.")
        return
    print(f"[DONE] new: {done}, skipped(existing): {skipped} -> {cfg['out_folder']}")
    if changed or identical:
        print(f"[DONE] rebuilt {len(changed) + len(identical)} existing cloud(s):")
        print(f"          IDENTICAL to before : {len(identical)}  (already the current seed)")
        print(f"          CHANGED             : {len(changed)}")
        if changed:
            print("\n       ⚠ THESE rows are the ones whose training input moved.")
            print("       ⚠ Any PointNet result computed before now used the OLD points,")
            print("       ⚠ so re-run 10_EXPORT_FORWARD_MODEL.bat (and the seed sweeps)")
            print("       ⚠ before quoting a PointNet number.")
            for n in changed[:40]:
                print(f"           {n}")
            if len(changed) > 40:
                print(f"           ... and {len(changed) - 40} more")
            with open(os.path.join(cfg["out_folder"], "_changed_clouds.txt"), "w") as cf:
                cf.write("\n".join(changed))
            print(f"\n       full list -> {os.path.join(cfg['out_folder'], '_changed_clouds.txt')}")
        else:
            print("\n       ✅ NOTHING CHANGED. Every cloud rebuilt bit-identical, so the")
            print("       ✅ stored clouds were already at the current seed and NO model")
            print("       ✅ needs retraining. The flag was only a missing seed record.")


if __name__ == "__main__":
    main()
